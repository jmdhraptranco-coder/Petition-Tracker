from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory, g, has_request_context, Response
from flask.sessions import SessionInterface, SessionMixin
from functools import wraps
from config import Config
import models
from auth_routes import (
    APIResult,
    InternalAPI,
    handle_first_login_setup,
    handle_forgot_password_request,
    handle_forgot_password_resend_otp,
    handle_forgot_password_set,
    handle_forgot_password_verify,
    handle_login,
    handle_login_verify,
    mask_mobile as auth_mask_mobile,
    normalize_mobile_for_otp as auth_normalize_mobile_for_otp,
)
from datetime import datetime, date, timedelta, timezone
from collections import Counter, deque
import os
import io
import csv
import re
import copy
import random
import json
import base64
import mimetypes
import urllib.parse
import ipaddress
import time
import threading
import hmac
import hashlib
import secrets
from uuid import uuid4
from werkzeug.exceptions import (
    BadGateway,
    BadRequest,
    Forbidden,
    GatewayTimeout,
    InternalServerError,
    MethodNotAllowed,
    NotFound,
    RequestEntityTooLarge,
    ServiceUnavailable,
    TooManyRequests,
    Unauthorized,
)
from werkzeug.datastructures import CallbackDict
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

import psycopg2
from psycopg2 import errors as pg_errors

config = Config()
app = Flask(__name__)
app.config['SECRET_KEY'] = config.SECRET_KEY
app.config['SESSION_COOKIE_NAME'] = config.SESSION_COOKIE_NAME
app.config['TEMPLATES_AUTO_RELOAD'] = bool(config.DEBUG)
app.jinja_env.auto_reload = bool(config.DEBUG)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = config.SESSION_COOKIE_SAMESITE
app.config['SESSION_COOKIE_SECURE'] = config.SESSION_COOKIE_SECURE
app.config['SESSION_COOKIE_DOMAIN'] = config.SESSION_COOKIE_DOMAIN
app.config['SESSION_COOKIE_PATH'] = config.SESSION_COOKIE_PATH
app.config['MAX_CONTENT_LENGTH'] = config.MAX_UPLOAD_SIZE_MB * 1024 * 1024
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=max(15, config.SESSION_INACTIVITY_MINUTES))
if config.TRUST_PROXY_HEADERS:
    # ProxyFix rewrites request.environ so Flask sees the *real* scheme,
    # host, and client IP from the reverse-proxy headers.  x_proto=1 means
    # "trust one hop of X-Forwarded-Proto", which makes request.is_secure
    # return True when the proxy forwards HTTPS traffic.
    #
    # Safari / outer-network resilience:
    #   Safari drops cookies whose Set-Cookie header carries the Secure flag
    #   but whose response was received over what it perceives as an insecure
    #   channel.  When Flask is behind nginx/HAProxy the actual TCP connection
    #   to Flask is plain HTTP; without ProxyFix, request.is_secure is False
    #   and Flask would NOT set the Secure flag even when SESSION_COOKIE_SECURE
    #   is True — or would set it inconsistently.
    #
    #   ProxyFix + PREFERRED_URL_SCHEME ensures url_for() produces https://
    #   links AND that the Secure flag is applied correctly on every response.
    app.wsgi_app = ProxyFix(
        app.wsgi_app,
        x_for=config.PROXY_FIX_X_FOR,
        x_proto=config.PROXY_FIX_X_PROTO,
        x_host=config.PROXY_FIX_X_HOST,
        x_port=config.PROXY_FIX_X_PORT,
    )
    app.config['PREFERRED_URL_SCHEME'] = 'https'
    # Enforce Secure cookie flag when behind a trusted proxy that terminates
    # TLS.  This prevents Safari from silently dropping session cookies on
    # mixed-protocol transitions during high-frequency clicks.
    # Only auto-enable if SESSION_COOKIE_SECURE was not explicitly set in the
    # environment — honouring SESSION_COOKIE_SECURE=0 lets developers run the
    # app over plain HTTP on a local network while still using ProxyFix.
    if os.environ.get('SESSION_COOKIE_SECURE') is None and not app.config.get('SESSION_COOKIE_SECURE'):
        app.config['SESSION_COOKIE_SECURE'] = True

_internal_auth_api = InternalAPI.from_config(config)


# ---------------------------------------------------------------------------
# Session-rotation grace period
# When a session ID is rotated at login the browser still holds the OLD
# cookie for any concurrent in-flight requests.  Without a grace window
# those requests find nothing in the DB and receive a fresh anonymous
# session → the user appears logged out ("2-3 click bug").
#
# Strategy:
#   • _rotate_session_identifier  registers old_sid→new_sid here instead of
#     immediately deleting old_sid from the store.
#   • open_session checks the map: if the cookie SID was recently rotated it
#     transparently loads the NEW session, so concurrent requests stay
#     authenticated.
#   • The old DB row is shortened to ROTATION_GRACE_SECONDS TTL (not deleted
#     outright) so the DB and map stay consistent.
#   • The map is process-local (thread-safe via a Lock).  For multi-process
#     deployments (Gunicorn) each worker maintains its own map; the DB-side
#     short TTL acts as the cross-process safety net.
# ---------------------------------------------------------------------------
_ROTATION_GRACE_SECONDS = 10
_rotation_grace: dict = {}          # {old_sid: (new_sid, expiry_float)}
_rotation_grace_lock = threading.Lock()

# ---------------------------------------------------------------------------
# Per-SID save lock
# Prevents two concurrent threads (or Gunicorn workers sharing a fork) from
# simultaneously writing divergent in-memory session state for the same SID.
# The lock is keyed by SID string, held only for the duration of the DB write,
# and lazily cleaned up whenever the grace-map is pruned.
# ---------------------------------------------------------------------------
_session_sid_locks: dict = {}       # {sid: threading.Lock()}
_session_sid_locks_lock = threading.Lock()


def _get_sid_save_lock(sid: str) -> threading.Lock:
    """Return the per-SID write lock, creating it if necessary."""
    with _session_sid_locks_lock:
        lock = _session_sid_locks.get(sid)
        if lock is None:
            lock = threading.Lock()
            _session_sid_locks[sid] = lock
        return lock


def _prune_sid_save_locks(active_sids: set) -> None:
    """Remove locks for SIDs that are no longer active (called during grace-map pruning)."""
    with _session_sid_locks_lock:
        stale = [k for k in list(_session_sid_locks) if k not in active_sids]
        for k in stale:
            _session_sid_locks.pop(k, None)


def _register_rotation_grace(old_sid: str, new_sid: str) -> None:
    expiry = time.time() + _ROTATION_GRACE_SECONDS
    with _rotation_grace_lock:
        now = time.time()
        stale = [k for k, (_, exp) in list(_rotation_grace.items()) if exp < now]
        for k in stale:
            _rotation_grace.pop(k, None)
        _rotation_grace[old_sid] = (new_sid, expiry)
        # Prune orphaned SID locks while we hold the grace lock.
        active = set(_rotation_grace.keys()) | {v for v, _ in _rotation_grace.values()}
    _prune_sid_save_locks(active)


def _resolve_rotation_grace(sid: str):
    """Return the replacement SID if sid was recently rotated, else None."""
    with _rotation_grace_lock:
        entry = _rotation_grace.get(sid)
    if not entry:
        return None
    new_sid, expiry = entry
    if time.time() > expiry:
        with _rotation_grace_lock:
            _rotation_grace.pop(sid, None)
        return None
    return new_sid


class DatabaseBackedSession(CallbackDict, SessionMixin):
    def __init__(self, initial=None, sid=None, new=False):
        def _mark_modified(_self):
            _self.modified = True

        super().__init__(initial, _mark_modified)
        self.sid = sid
        self.new = new
        self.modified = False
        self.last_persisted_expires_at = None
        self.last_persisted_accessed_at = None


TEST_SERVER_SESSION_STORE = {}
SESSION_DIAGNOSTIC_EVENTS = deque(maxlen=200)


def _record_session_diagnostic(event_type, severity='warning', **details):
    entry = {
        'event_type': event_type,
        'severity': severity,
        'ts': datetime.now(timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z'),
    }
    safe_keys = {
        'reason',
        'path',
        'method',
        'forwarded_proto',
        'request_secure',
        'forwarded_host',
        'forwarded_port',
        'cookie_samesite',
        'cookie_secure',
        'trust_proxy_headers',
        'session_cookie_name',
    }
    for key, value in details.items():
        if key in safe_keys and value is not None:
            entry[key] = value
    SESSION_DIAGNOSTIC_EVENTS.appendleft(entry)


def _load_server_session_record(session_id):
    if app.config.get('TESTING'):
        record = TEST_SERVER_SESSION_STORE.get(session_id)
        if not record:
            return None
        expires_at = record.get('expires_at')
        if expires_at and expires_at <= datetime.now(timezone.utc).replace(tzinfo=None):
            TEST_SERVER_SESSION_STORE.pop(session_id, None)
            return None
        return {
            'session_id': session_id,
            'user_id': record.get('user_id'),
            'data': copy.deepcopy(record.get('data') or {}),
            'expires_at': expires_at,
        }
    if hasattr(models, 'get_server_session'):
        return models.get_server_session(session_id)
    return None


def _save_server_session_record(session_id, data, user_id, expires_at):
    if app.config.get('TESTING'):
        TEST_SERVER_SESSION_STORE[session_id] = {
            'user_id': user_id,
            'data': copy.deepcopy(data or {}),
            'expires_at': expires_at,
        }
        return
    if hasattr(models, 'save_server_session'):
        models.save_server_session(session_id, data, user_id, expires_at)


def _delete_server_session_record(session_id):
    if app.config.get('TESTING'):
        TEST_SERVER_SESSION_STORE.pop(session_id, None)
        return
    if hasattr(models, 'delete_server_session'):
        models.delete_server_session(session_id)


def _expire_server_session_soon(session_id, seconds):
    """Shorten an old session's TTL to `seconds` from now.

    Used after session-ID rotation: rather than hard-deleting the old row
    immediately (which breaks concurrent in-flight requests carrying the old
    cookie), we let it live for a brief grace window so those requests can
    still be authenticated via the grace-map lookup in open_session.
    """
    if not session_id:
        return
    if app.config.get('TESTING'):
        # In the test store the old SID is no longer needed; remove it.
        TEST_SERVER_SESSION_STORE.pop(session_id, None)
        return
    if hasattr(models, 'expire_server_session_soon'):
        try:
            models.expire_server_session_soon(session_id, seconds)
        except Exception:
            # Non-fatal: fall back to hard delete so the old row doesn't linger.
            _delete_server_session_record(session_id)


def _touch_server_session_record(session_id, expires_at):
    if not session_id:
        return False
    if app.config.get('TESTING'):
        record = TEST_SERVER_SESSION_STORE.get(session_id)
        if not record:
            return False
        last_accessed_at = record.get('last_accessed_at')
        now_dt = datetime.now(timezone.utc).replace(tzinfo=None)
        threshold_dt = now_dt - timedelta(seconds=config.SESSION_TOUCH_THRESHOLD_SECONDS)
        if (
            last_accessed_at
            and last_accessed_at > threshold_dt
            and record.get('expires_at')
            and expires_at
            and record.get('expires_at') >= expires_at
        ):
            return False
        record['expires_at'] = expires_at
        record['last_accessed_at'] = now_dt
        return True
    if hasattr(models, 'touch_server_session'):
        return bool(models.touch_server_session(session_id, expires_at, config.SESSION_TOUCH_THRESHOLD_SECONDS))
    return False


def _invalidate_current_session(reason=None, *, revoke_store=True):
    current_sid = getattr(session, 'sid', None)
    session.clear()
    if revoke_store and current_sid:
        _delete_server_session_record(current_sid)
    if has_request_context() and reason:
        g.auth_invalid_reason = reason
        _record_session_diagnostic(
            'auth.session_invalidated',
            severity='warning',
            reason=reason,
            path=request.path,
            method=request.method,
        )


def _queue_proxy_mismatch_diagnostics():
    if not has_request_context() or not config.TRUST_PROXY_HEADERS:
        return
    forwarded_proto = (request.headers.get('X-Forwarded-Proto') or '').strip().lower()
    is_secure = bool(request.is_secure)
    if forwarded_proto and ((forwarded_proto == 'https') != is_secure):
        _record_session_diagnostic(
            'proxy.proto_mismatch',
            severity='warning',
            forwarded_proto=forwarded_proto,
            request_secure=is_secure,
            forwarded_host=(request.headers.get('X-Forwarded-Host') or '').strip() or None,
            forwarded_port=(request.headers.get('X-Forwarded-Port') or '').strip() or None,
            cookie_samesite=config.SESSION_COOKIE_SAMESITE,
            cookie_secure=config.SESSION_COOKIE_SECURE,
            trust_proxy_headers=config.TRUST_PROXY_HEADERS,
            session_cookie_name=config.SESSION_COOKIE_NAME,
        )
        log_security_event(
            'proxy.proto_mismatch',
            severity='warning',
            forwarded_proto=forwarded_proto,
            request_secure=is_secure,
            forwarded_host=(request.headers.get('X-Forwarded-Host') or '').strip() or None,
            forwarded_port=(request.headers.get('X-Forwarded-Port') or '').strip() or None,
        )


def _dedupe_session_cookie_headers(response):
    cookie_name = app.config.get('SESSION_COOKIE_NAME') or 'session'
    set_cookie_headers = response.headers.getlist('Set-Cookie')
    if len(set_cookie_headers) <= 1:
        return response
    session_prefix = f'{cookie_name}='
    kept = []
    latest_session_header = None
    for value in set_cookie_headers:
        if value.startswith(session_prefix):
            latest_session_header = value
        else:
            kept.append(value)
    if latest_session_header is not None:
        kept.append(latest_session_header)
    response.headers.setlist('Set-Cookie', kept)
    return response


def _session_store_health_payload():
    stats = {
        'store': 'database',
        'status': 'ok',
        'trust_proxy_headers': bool(config.TRUST_PROXY_HEADERS),
        'session_cookie_name': config.SESSION_COOKIE_NAME,
        'session_cookie_secure': bool(config.SESSION_COOKIE_SECURE),
        'session_cookie_samesite': config.SESSION_COOKIE_SAMESITE,
        'session_touch_threshold_seconds': int(config.SESSION_TOUCH_THRESHOLD_SECONDS),
        'session_inactivity_minutes': int(config.SESSION_INACTIVITY_MINUTES),
        'session_absolute_hours': int(config.SESSION_ABSOLUTE_HOURS),
    }
    if hasattr(models, 'get_server_session_health_stats'):
        db_stats = models.get_server_session_health_stats() or {}
        stats.update(db_stats)
    total = int(stats.get('total_sessions') or 0)
    expired = int(stats.get('expired_sessions') or 0)
    stats['expired_ratio'] = round((expired / total), 4) if total > 0 else 0.0
    return stats


class DatabaseSessionInterface(SessionInterface):
    session_class = DatabaseBackedSession

    def generate_sid(self):
        return secrets.token_urlsafe(32)

    def open_session(self, app, request):
        cookie_name = self.get_cookie_name(app)
        sid = request.cookies.get(cookie_name)
        if not sid:
            return self.session_class(sid=self.generate_sid(), new=True)
        record = _load_server_session_record(sid)
        if record:
            # Old row still exists (normal path, or within the rotation grace
            # window).  Check the grace map: if this SID was recently rotated
            # AND the new row is already in DB, migrate immediately to the new
            # SID.  This advances the browser's cookie to the new SID on this
            # very response, so the old row expiring at the end of the grace
            # window never causes a logout.
            #
            # If the new row does not exist yet (login's save_session hasn't
            # completed), keep the old SID — the old row is still valid and the
            # browser will receive the new SID cookie from the login response.
            forwarded_sid = _resolve_rotation_grace(sid)
            if forwarded_sid:
                new_record = _load_server_session_record(forwarded_sid)
                if new_record:
                    record = new_record
                    sid = forwarded_sid
        else:
            # Old row is gone (expired or deleted).  Try the grace map: maybe
            # the SID was just rotated and the new row is already in DB.
            forwarded_sid = _resolve_rotation_grace(sid)
            if forwarded_sid:
                record = _load_server_session_record(forwarded_sid)
                if record:
                    sid = forwarded_sid
        if not record:
            return self.session_class(sid=self.generate_sid(), new=True)
        session_obj = self.session_class(initial=record.get('data') or {}, sid=sid, new=False)
        session_obj.last_persisted_expires_at = record.get('expires_at')
        session_obj.last_persisted_accessed_at = record.get('last_accessed_at')
        return session_obj

    def save_session(self, app, session_obj, response):
        cookie_name = self.get_cookie_name(app)
        domain = self.get_cookie_domain(app)
        path = self.get_cookie_path(app)

        if not session_obj:
            # Empty session dict: the session was explicitly cleared (logout /
            # invalidation).  Only delete if it was a real persisted session.
            if session_obj.new and not session_obj.modified:
                return
            sid = getattr(session_obj, 'sid', None)
            if sid:
                lock = _get_sid_save_lock(sid)
                with lock:
                    _delete_server_session_record(sid)
            response.delete_cookie(
                cookie_name,
                domain=domain,
                path=path,
                secure=self.get_cookie_secure(app),
                samesite=self.get_cookie_samesite(app),
                httponly=self.get_cookie_httponly(app),
            )
            return

        if not getattr(session_obj, 'sid', None):
            session_obj.sid = self.generate_sid()

        # _pending_touch is set by _load_current_authenticated_user when it
        # updates auth_last_seen_at without wanting a full UPSERT (see below).
        pending_touch = getattr(session_obj, '_pending_touch', False)

        expires = self.get_expiration_time(app, session_obj)
        persist_expires = expires
        if persist_expires is not None and getattr(persist_expires, 'tzinfo', None) is not None:
            persist_expires = persist_expires.astimezone(timezone.utc).replace(tzinfo=None)
        user_id = session_obj.get('user_id')

        sid = session_obj.sid
        lock = _get_sid_save_lock(sid)
        with lock:
            if session_obj.modified or session_obj.new:
                # Full UPSERT — data actually changed.
                _save_server_session_record(
                    sid,
                    dict(session_obj),
                    user_id,
                    persist_expires or (datetime.now(timezone.utc).replace(tzinfo=None) + app.permanent_session_lifetime),
                )
                session_obj.last_persisted_expires_at = persist_expires
                session_obj.last_persisted_accessed_at = datetime.now(timezone.utc).replace(tzinfo=None)
            elif pending_touch or self.should_set_cookie(app, session_obj):
                # Touch-only path: update expiry/last_accessed without a full
                # data write.  This is the normal path for rapid-click requests
                # that only bumped auth_last_seen_at in-memory.
                refreshed_expires = persist_expires or (datetime.now(timezone.utc).replace(tzinfo=None) + app.permanent_session_lifetime)
                touched = _touch_server_session_record(sid, refreshed_expires)
                if touched:
                    session_obj.last_persisted_expires_at = refreshed_expires
                    session_obj.last_persisted_accessed_at = datetime.now(timezone.utc).replace(tzinfo=None)

        if session_obj.modified or session_obj.new or pending_touch or self.should_set_cookie(app, session_obj):
            response.set_cookie(
                cookie_name,
                sid,
                expires=expires,
                httponly=self.get_cookie_httponly(app),
                secure=self.get_cookie_secure(app),
                samesite=self.get_cookie_samesite(app),
                path=path,
                domain=domain,
            )


app.session_interface = DatabaseSessionInterface()

if os.getenv('SKIP_SCHEMA_UPDATES') != '1':
    models.ensure_schema_updates()

BASE_UPLOAD_DIR = config.UPLOAD_BASE_DIR
ERECEIPT_UPLOAD_DIR = os.path.join(BASE_UPLOAD_DIR, 'e_receipts')
ENQUIRY_UPLOAD_DIR = os.path.join(BASE_UPLOAD_DIR, 'enquiry_reports')
PROFILE_UPLOAD_DIR = os.path.join(BASE_UPLOAD_DIR, 'profile_photos')
HELP_RESOURCE_UPLOAD_DIR = os.path.join(BASE_UPLOAD_DIR, 'help_resources')
MAX_UPLOAD_SIZE_BYTES = config.MAX_UPLOAD_SIZE_MB * 1024 * 1024
PROFILE_PHOTO_MAX_BYTES = 2 * 1024 * 1024
PROFILE_PHOTO_EXTENSIONS = {'jpg', 'jpeg', 'png', 'webp'}
HELP_RESOURCE_ALLOWED_EXTENSIONS = {
    'pdf', 'doc', 'docx', 'ppt', 'pptx', 'xls', 'xlsx',
    'jpg', 'jpeg', 'png', 'webp',
    'mp4', 'webm', 'mov'
}
HELP_RESOURCE_TYPES = {'manual', 'flowchart', 'video', 'office_order', 'news'}
HELP_RESOURCE_STORAGE_KINDS = {'upload', 'external_url'}
LOGIN_ATTEMPTS = {}
PETITION_SUBMISSION_ATTEMPTS = {}
VALID_RECEIVED_AT = {'jmd_office', 'cvo_apspdcl_tirupathi', 'cvo_apepdcl_vizag', 'cvo_apcpdcl_vijayawada'}
VALID_TARGET_CVO = {'apspdcl', 'apepdcl', 'apcpdcl', 'headquarters'}
VALID_ORGANIZATIONS = {'aptransco', 'apgenco'}
VALID_ENQUIRY_TYPES = {'detailed', 'preliminary'}
VALID_SOURCE_OF_PETITION = {'media', 'public_individual', 'govt', 'sumoto', 'cmd_office'}
VALID_GOVT_INSTITUTIONS = {
    'aprc',
    'governor',
    'cs_energy_department',
    'cmd_aptransco',
    'cmo',
    'energy_department',
}
LOGIN_CAPTCHA_TTL_SECONDS = 300
LOGIN_CAPTCHA_LENGTH = 6
LOGIN_CAPTCHA_ALPHABET = '23456789'
LOGIN_CAPTCHA_USED_TOKENS = {}
LOGIN_CAPTCHA_CHALLENGES = {}
VALID_PETITION_TYPES = {
    'bribe',
    'corruption',
    'harassment',
    'electrical_accident',
    'misconduct',
    'works_related',
    'irregularities_in_tenders',
    'illegal_assets',
    'fake_certificates',
    'theft_misappropriation_materials',
    'other',
}
IMPORT_ALLOWED_STATUSES = {
    'received',
    'forwarded_to_cvo',
    'sent_for_permission',
    'permission_approved',
    'permission_rejected',
    'assigned_to_inspector',
    'sent_back_for_reenquiry',
    'enquiry_in_progress',
    'enquiry_report_submitted',
    'forwarded_to_jmd',
    'forwarded_to_po',
    'action_instructed',
    'action_taken',
    'lodged',
    'closed',
}
IMPORT_PETITION_HEADERS = [
    'received_date',
    'received_at',
    'target_cvo',
    'petitioner_name',
    'contact',
    'place',
    'subject',
    'petition_type',
    'source_of_petition',
    'govt_institution_type',
    'enquiry_type',
    'permission_request_type',
    'requires_permission',
    'permission_status',
    'status',
    'efile_no',
    'ereceipt_no',
    'remarks',
    'assigned_inspector_username',
]
IMPORT_HEADER_ALIASES = {
    'date': 'received_date',
    'receiveddate': 'received_date',
    'received date': 'received_date',
    'received on': 'received_date',
    'office': 'received_at',
    'received office': 'received_at',
    'received_at_office': 'received_at',
    'target': 'target_cvo',
    'target office': 'target_cvo',
    'petitioner': 'petitioner_name',
    'petitioner name': 'petitioner_name',
    'mobile': 'contact',
    'phone': 'contact',
    'location': 'place',
    'address': 'place',
    'type': 'petition_type',
    'petition type': 'petition_type',
    'source': 'source_of_petition',
    'source of petition': 'source_of_petition',
    'institution': 'govt_institution_type',
    'government institution type': 'govt_institution_type',
    'enquiry mode': 'enquiry_type',
    'permission type': 'permission_request_type',
    'permission request': 'permission_request_type',
    'permission required': 'requires_permission',
    'permission status': 'permission_status',
    'current status': 'status',
    'eoffice no': 'efile_no',
    'e-office no': 'efile_no',
    'ereceipt no': 'ereceipt_no',
    'e-receipt no': 'ereceipt_no',
    'inspector username': 'assigned_inspector_username',
}
PETITION_TYPE_LABELS = {
    # Current workflow values
    'bribe': 'Bribe',
    'corruption': 'Corruption',
    'harassment': 'Harassment',
    'electrical_accident': 'Electrical Accident',
    'misconduct': 'Misconduct',
    'works_related': 'Works Related',
    'irregularities_in_tenders': 'Irregularities in Tenders',
    'illegal_assets': 'Illegal Assets',
    'fake_certificates': 'Fake Certificates',
    'theft_misappropriation_materials': 'Theft/Misappropriation of Materials',
    'other': 'Other',
    # Legacy DB values kept for display/filter compatibility
    'theft_of_materials': 'Theft of Materials',
    'adverse_news': 'Adverse News',
    'procedural_lapses': 'Procedural Lapses',
}
VALID_PERMISSION_REQUEST_TYPES = {'direct_enquiry', 'permission_required'}
DIRECT_ENQUIRY_EFILE_EDITABLE_STATUSES = {'received', 'forwarded_to_cvo', 'assigned_to_inspector', 'enquiry_in_progress'}
VALID_USER_ROLES = {
    'super_admin', 'data_entry', 'po',
    'cmd_apspdcl', 'cmd_apepdcl', 'cmd_apcpdcl',
    'cgm_hr_transco',
    'dsp', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'inspector'
}
VALID_CVO_OFFICES = {'apspdcl', 'apepdcl', 'apcpdcl', 'headquarters'}
DEO_OFFICE_FLOW = {
    'headquarters': {
        'received_at': 'jmd_office',
        'received_at_label': 'JMD Office',
        'target_cvo': 'headquarters',
        'target_cvo_label': 'Headquarters (DSP)',
        'force_permission_required': True,
    },
    'apspdcl': {
        'received_at': 'cvo_apspdcl_tirupathi',
        'received_at_label': 'CVO/DSP (APSPDCL) - Tirupathi',
        'target_cvo': 'apspdcl',
        'target_cvo_label': 'APSPDCL (Tirupathi)',
        'force_permission_required': False,
    },
    'apepdcl': {
        'received_at': 'cvo_apepdcl_vizag',
        'received_at_label': 'CVO/DSP (APEPDCL) - Vizag',
        'target_cvo': 'apepdcl',
        'target_cvo_label': 'APEPDCL (Vizag)',
        'force_permission_required': False,
    },
    'apcpdcl': {
        'received_at': 'cvo_apcpdcl_vijayawada',
        'received_at_label': 'CVO/DSP (APCPDCL) - Vijayawada',
        'target_cvo': 'apcpdcl',
        'target_cvo_label': 'APCPDCL (Vijayawada)',
        'force_permission_required': False,
    },
}
DEO_COMBINED_TARGET_FLOW = {
    # APSPDCL DEO handles both APSPDCL and APCPDCL entries.
    'apspdcl': [
        {
            'received_at': 'cvo_apspdcl_tirupathi',
            'received_at_label': 'CVO/DSP (APSPDCL) - Tirupathi',
            'target_cvo': 'apspdcl',
            'target_cvo_label': 'APSPDCL (Tirupathi)',
            'force_permission_required': False,
        },
        {
            'received_at': 'cvo_apcpdcl_vijayawada',
            'received_at_label': 'CVO/DSP (APCPDCL) - Vijayawada',
            'target_cvo': 'apcpdcl',
            'target_cvo_label': 'APCPDCL (Vijayawada)',
            'force_permission_required': False,
        },
    ],
}
PHONE_RE = re.compile(r'^[0-9+\-\s()]{7,20}$')
EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
VALID_DYNAMIC_FIELD_TYPES = {'text', 'textarea', 'select', 'date', 'tel', 'email', 'file'}

DEFAULT_FORM_FIELD_CONFIGS = {
    'deo_petition.received_date': {'label': 'Received Date', 'type': 'date', 'required': True, 'options': []},
    'deo_petition.received_at': {
        'label': 'Received At',
        'type': 'select',
        'required': True,
        'options': [
            {'value': 'jmd_office', 'label': 'JMD Office'},
            {'value': 'cvo_apspdcl_tirupathi', 'label': 'CVO/DSP (APSPDCL) - Tirupathi'},
            {'value': 'cvo_apepdcl_vizag', 'label': 'CVO/DSP (APEPDCL) - Vizag'},
            {'value': 'cvo_apcpdcl_vijayawada', 'label': 'CVO/DSP (APCPDCL) - Vijayawada'},
        ]
    },
    'deo_petition.ereceipt_no': {'label': 'E-Receipt No', 'type': 'text', 'required': False, 'options': []},
    'deo_petition.ereceipt_file': {'label': 'Upload E-Receipt (PDF, max 10MB)', 'type': 'file', 'required': False, 'options': []},
    'deo_petition.target_cvo': {
        'label': 'Target CVO/DSP Jurisdiction',
        'type': 'select',
        'required': False,
        'options': [
            {'value': 'apspdcl', 'label': 'APSPDCL (Tirupathi)'},
            {'value': 'apepdcl', 'label': 'APEPDCL (Vizag)'},
            {'value': 'apcpdcl', 'label': 'APCPDCL (Vijayawada)'},
            {'value': 'headquarters', 'label': 'Headquarters (DSP)'},
        ]
    },
    'deo_petition.permission_request_type': {
        'label': 'Enquiry Permission',
        'type': 'select',
        'required': True,
        'options': [
            {'value': 'direct_enquiry', 'label': 'Direct'},
            {'value': 'permission_required', 'label': 'Permission Based (CVO/DSP sends to PO for approval)'},
        ]
    },
    'deo_petition.petitioner_name': {'label': 'Petitioner Name', 'type': 'text', 'required': False, 'options': []},
    'deo_petition.contact': {'label': 'Contact Number', 'type': 'tel', 'required': False, 'options': []},
    'deo_petition.place': {'label': 'Place', 'type': 'text', 'required': False, 'options': []},
    'deo_petition.subject': {'label': 'Subject', 'type': 'textarea', 'required': True, 'options': []},
    'deo_petition.petition_type': {
        'label': 'Type of Petition',
        'type': 'select',
        'required': True,
        'options': [
            {'value': 'bribe', 'label': 'Bribe'},
            {'value': 'corruption', 'label': 'Corruption'},
            {'value': 'harassment', 'label': 'Harassment'},
            {'value': 'electrical_accident', 'label': 'Electrical Accident'},
            {'value': 'misconduct', 'label': 'Misconduct'},
            {'value': 'works_related', 'label': 'Works Related'},
            {'value': 'irregularities_in_tenders', 'label': 'Irregularities in Tenders'},
            {'value': 'illegal_assets', 'label': 'Illegal Assets'},
            {'value': 'fake_certificates', 'label': 'Fake Certificates'},
            {'value': 'theft_misappropriation_materials', 'label': 'Theft/Misappropriation of Materials'},
            {'value': 'other', 'label': 'Other'},
        ]
    },
    'deo_petition.source_of_petition': {
        'label': 'Source of Petition',
        'type': 'select',
        'required': True,
        'options': [
            {'value': 'media', 'label': 'Electronic and Print Media'},
            {'value': 'public_individual', 'label': 'Public (Individual)'},
            {'value': 'govt', 'label': 'Govt'},
            {'value': 'sumoto', 'label': 'Sumoto'},
            {'value': 'cmd_office', 'label': 'O/o CMD'},
        ]
    },
    'deo_petition.remarks': {'label': 'Remarks', 'type': 'textarea', 'required': False, 'options': []},
    'deo_petition.govt_institution_type': {
        'label': 'Type of Institution',
        'type': 'select',
        'required': True,
        'options': [
            {'value': 'aprc', 'label': '1. APRC'},
            {'value': 'governor', 'label': '2. Governor'},
            {'value': 'cs_energy_department', 'label': '3. CS (Energy Department)'},
            {'value': 'cmd_aptransco', 'label': '4. CMD APTRANSCO'},
            {'value': 'cmo', 'label': '5. CMO'},
            {'value': 'energy_department', 'label': '6. Energy Minister'},
        ]
    },
    'inspector_report.report_text': {'label': 'Conclusion of Enquiry Report', 'type': 'textarea', 'required': True, 'options': []},
    'inspector_report.recommendation': {'label': 'Recommendations / Suggestions', 'type': 'textarea', 'required': True, 'options': []},
    'inspector_report.report_file': {'label': 'Enquiry File (PDF, max 10MB)', 'type': 'file', 'required': True, 'options': []},
    'inspector_report.request_detailed_permission': {
        'label': 'Ask permission to convert this preliminary enquiry into detailed enquiry',
        'type': 'text',
        'required': False,
        'options': []
    },
    'inspector_report.detailed_request_reason': {
        'label': 'Reason for Detailed Enquiry Request',
        'type': 'textarea',
        'required': True,
        'options': []
    },
    'cvo_review.cvo_comments': {'label': 'CVO/DSP Comments on Enquiry Report', 'type': 'textarea', 'required': True, 'options': []},
    'cvo_review.consolidated_report_file': {'label': 'Consolidated Report File (PDF, Optional, max 10MB)', 'type': 'file', 'required': False, 'options': []},
    'cmd_action.action_taken': {'label': 'Action Taken Details', 'type': 'textarea', 'required': True, 'options': []},
    'cmd_action.action_report_file': {'label': 'Upload Action Report Copy (PDF, Optional, max 10MB)', 'type': 'file', 'required': False, 'options': []},
    'po_decision.approve_permission_efile_no': {'label': 'E-Office File No', 'type': 'text', 'required': True, 'options': []},
    'po_decision.reject_permission_reason': {'label': 'Reason for Rejection', 'type': 'textarea', 'required': True, 'options': []},
    'po_decision.send_cmd_instructions': {'label': 'CMD/CGM-HR Instructions', 'type': 'textarea', 'required': False, 'options': []},
    'po_decision.send_cmd_efile_no': {'label': 'E-Office File No', 'type': 'text', 'required': True, 'options': []},
    'po_decision.po_lodge_efile_no': {'label': 'E-Office File No', 'type': 'text', 'required': False, 'options': []},
    'po_decision.po_lodge_remarks': {'label': 'PO Lodge Remarks', 'type': 'textarea', 'required': False, 'options': []},
    'po_decision.po_direct_lodge_efile_no': {'label': 'E-Office File No', 'type': 'text', 'required': False, 'options': []},
    'po_decision.po_direct_lodge_remarks': {'label': 'PO Lodge Remarks', 'type': 'textarea', 'required': False, 'options': []},
    'po_decision.close_comments': {'label': 'Closing Remarks', 'type': 'textarea', 'required': False, 'options': []},
}

FORM_MANAGEMENT_GROUPS = {
    'deo_petition': 'DEO Petition Form',
    'inspector_report': 'Inspector Enquiry Form',
    'cvo_review': 'CVO/DSP Review Form',
    'cmd_action': 'CMD/CGM-HR Action Form',
    'po_decision': 'PO Decision Form',
}

SYSTEM_SETTING_DEFINITIONS = {
    'petition_user_rate_limit_window_seconds': {
        'label': 'Per-user window (seconds)',
        'description': 'How long each DEO account submission window stays open.',
        'default': lambda: int(config.PETITION_USER_RATE_LIMIT_WINDOW_SECONDS),
        'min': 30,
        'max': 3600,
    },
    'petition_user_rate_limit_max_submissions': {
        'label': 'Per-user max submissions',
        'description': 'Maximum successful petition submissions allowed inside the per-user window.',
        'default': lambda: int(config.PETITION_USER_RATE_LIMIT_MAX_SUBMISSIONS),
        'min': 1,
        'max': 200,
    },
    'petition_user_rate_limit_block_seconds': {
        'label': 'Per-user block duration (seconds)',
        'description': 'Temporary block duration after a DEO account crosses the user threshold.',
        'default': lambda: int(config.PETITION_USER_RATE_LIMIT_BLOCK_SECONDS),
        'min': 30,
        'max': 3600,
    },
    'petition_ip_rate_limit_window_seconds': {
        'label': 'Per-IP window (seconds)',
        'description': 'How long the shared office IP submission window stays open.',
        'default': lambda: int(config.PETITION_IP_RATE_LIMIT_WINDOW_SECONDS),
        'min': 30,
        'max': 3600,
    },
    'petition_ip_rate_limit_max_submissions': {
        'label': 'Per-IP max submissions',
        'description': 'Maximum successful petition submissions allowed from one office/public IP inside the IP window.',
        'default': lambda: int(config.PETITION_IP_RATE_LIMIT_MAX_SUBMISSIONS),
        'min': 1,
        'max': 1000,
    },
    'petition_ip_rate_limit_block_seconds': {
        'label': 'Per-IP block duration (seconds)',
        'description': 'Temporary block duration after the shared office/public IP crosses the IP threshold.',
        'default': lambda: int(config.PETITION_IP_RATE_LIMIT_BLOCK_SECONDS),
        'min': 30,
        'max': 3600,
    },
}


def parse_optional_int(raw_value):
    value = (raw_value or '').strip()
    if not value:
        return None
    try:
        parsed = int(value)
        return parsed if parsed > 0 else None
    except (TypeError, ValueError):
        return None


def parse_date_input(value):
    text = (value or '').strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, '%Y-%m-%d').date()
    except ValueError:
        return None


def parse_flexible_date(value):
    text = (value or '').strip()
    if not text:
        return None
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_header_key(raw):
    text = str(raw or '').strip().lower()
    text = re.sub(r'[^a-z0-9]+', '_', text).strip('_')
    if not text:
        return ''
    alias_key = text.replace('_', ' ')
    return IMPORT_HEADER_ALIASES.get(alias_key) or IMPORT_HEADER_ALIASES.get(text) or text


def _parse_tabular_upload_rows(upload, required_headers, allowed_headers):
    filename = secure_filename(upload.filename or '')
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in ('xlsx', 'csv'):
        raise ValueError('Only .xlsx or .csv files are allowed.')

    rows = []
    header_map = {}
    if ext == 'csv':
        content = upload.stream.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        raw_headers = reader.fieldnames or []
        for h in raw_headers:
            canonical = _normalize_header_key(h)
            if canonical and canonical in allowed_headers and canonical not in header_map:
                header_map[canonical] = h
        if not set(required_headers).issubset(set(header_map.keys())):
            missing = [h for h in required_headers if h not in header_map]
            raise ValueError(f'Missing required column(s): {", ".join(missing)}')
        for row in reader:
            data = {}
            for canonical, original in header_map.items():
                data[canonical] = str(row.get(original) or '').strip()
            if any(v for v in data.values()):
                rows.append(data)
    else:
        if load_workbook is None:
            raise ValueError('Excel support requires openpyxl dependency.')
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []
        raw_headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        for idx, h in enumerate(raw_headers):
            canonical = _normalize_header_key(h)
            if canonical and canonical in allowed_headers and canonical not in header_map:
                header_map[canonical] = idx
        if not set(required_headers).issubset(set(header_map.keys())):
            missing = [h for h in required_headers if h not in header_map]
            raise ValueError(f'Missing required column(s): {", ".join(missing)}')
        for row in all_rows[1:]:
            data = {}
            for canonical, idx in header_map.items():
                value = row[idx] if idx < len(row) else ''
                data[canonical] = str(value).strip() if value is not None else ''
            if any(v for v in data.values()):
                rows.append(data)
    return rows


def _normalize_received_at(value):
    val = (value or '').strip().lower()
    if not val:
        return None
    alias = {
        'jmd': 'jmd_office',
        'jmd office': 'jmd_office',
        'po office': 'jmd_office',
        'cvo apspdcl tirupathi': 'cvo_apspdcl_tirupathi',
        'apspdcl': 'cvo_apspdcl_tirupathi',
        'tirupathi': 'cvo_apspdcl_tirupathi',
        'cvo apepdcl vizag': 'cvo_apepdcl_vizag',
        'apepdcl': 'cvo_apepdcl_vizag',
        'vizag': 'cvo_apepdcl_vizag',
        'cvo apcpdcl vijayawada': 'cvo_apcpdcl_vijayawada',
        'apcpdcl': 'cvo_apcpdcl_vijayawada',
        'vijayawada': 'cvo_apcpdcl_vijayawada',
    }
    return val if val in VALID_RECEIVED_AT else alias.get(val)


def _normalize_target_cvo(value):
    val = (value or '').strip().lower()
    if not val:
        return None
    alias = {
        'hq': 'headquarters',
        'headquarter': 'headquarters',
        'headquarters dsp': 'headquarters',
        'dsp': 'headquarters',
        'apspdcl tirupathi': 'apspdcl',
        'apepdcl vizag': 'apepdcl',
        'apcpdcl vijayawada': 'apcpdcl',
    }
    return val if val in VALID_TARGET_CVO else alias.get(val)


def _normalize_source(value):
    val = (value or '').strip().lower()
    alias = {
        'electronic and print media': 'media',
        'print media': 'media',
        'electronic media': 'media',
        'public': 'public_individual',
        'public individual': 'public_individual',
        'gov': 'govt',
        'government': 'govt',
        'suo motu': 'sumoto',
        'suomoto': 'sumoto',
        'o/o cmd': 'cmd_office',
        'cmd office': 'cmd_office',
    }
    normalized = alias.get(val, val)
    return normalized if normalized in VALID_SOURCE_OF_PETITION else 'public_individual'


def _normalize_petition_type(value):
    val = (value or '').strip().lower()
    alias = {
        'electrical accident': 'electrical_accident',
        'works related': 'works_related',
        'irregularities in tenders': 'irregularities_in_tenders',
        'illegal assets': 'illegal_assets',
        'fake certificates': 'fake_certificates',
        'theft/misappropriation of materials': 'theft_misappropriation_materials',
        'theft misappropriation materials': 'theft_misappropriation_materials',
    }
    normalized = alias.get(val, val)
    return normalized if normalized in VALID_PETITION_TYPES else 'other'


def _to_bool(value, default=False):
    text = (value or '').strip().lower()
    if text in {'1', 'true', 'yes', 'y', 'required'}:
        return True
    if text in {'0', 'false', 'no', 'n', 'not_required', 'direct'}:
        return False
    return default


def _normalize_petitioner_name(value):
    text = re.sub(r'\s+', ' ', (value or '').strip())
    return text


def _build_petitioner_profile_payload(petitions, petitioner_name):
    target = _normalize_petitioner_name(petitioner_name).lower()
    matches = []
    for p in petitions:
        pname = _normalize_petitioner_name(p.get('petitioner_name') or '')
        if not pname or pname.lower() in {'anonymous', '-'}:
            continue
        if pname.lower() == target:
            matches.append(p)

    total = len(matches)
    status_counts = Counter()
    type_counts = Counter()
    source_counts = Counter()
    month_counts = Counter()
    recent = []
    for p in matches:
        status = (p.get('status') or 'unknown').strip()
        ptype = (p.get('petition_type') or 'other').strip()
        source = (p.get('source_of_petition') or 'public_individual').strip()
        rd = p.get('received_date')
        status_counts[status] += 1
        type_counts[ptype] += 1
        source_counts[source] += 1
        if rd:
            month_counts[rd.strftime('%Y-%m')] += 1
        recent.append({
            'id': int(p.get('id') or 0),
            'sno': p.get('sno') or '-',
            'subject': p.get('subject') or '-',
            'status': status_labels_for_api().get(status, status.replace('_', ' ').title()),
            'received_date': rd.strftime('%d/%m/%Y') if rd else '-',
            '_received_sort': rd.isoformat() if rd else '',
            'view_url': url_for('petition_view', petition_id=int(p.get('id') or 0)) if int(p.get('id') or 0) > 0 else '#',
        })
    recent.sort(key=lambda r: r.get('_received_sort', ''), reverse=True)

    years_months = sorted(month_counts.keys())
    trend_labels = []
    trend_values = []
    for ym in years_months[-12:]:
        try:
            dt = datetime.strptime(ym + '-01', '%Y-%m-%d')
            trend_labels.append(dt.strftime('%b %Y'))
        except Exception:
            trend_labels.append(ym)
        trend_values.append(month_counts[ym])

    def top_counter(counter_obj, limit=8):
        items = sorted(counter_obj.items(), key=lambda x: x[1], reverse=True)[:limit]
        return {
            'labels': [k.replace('_', ' ').title() for k, _ in items],
            'values': [v for _, v in items],
        }

    return {
        'petitioner_name': petitioner_name,
        'total_petitions': total,
        'closed_count': status_counts.get('closed', 0),
        'open_count': max(0, total - status_counts.get('closed', 0)),
        'lodged_count': status_counts.get('lodged', 0),
        'trend': {
            'labels': trend_labels,
            'values': trend_values,
        },
        'status_split': top_counter(status_counts, limit=10),
        'type_split': top_counter(type_counts, limit=8),
        'source_split': top_counter(source_counts, limit=6),
        'recent_petitions': [{k: v for k, v in item.items() if not k.startswith('_')} for item in recent[:12]],
    }


def status_labels_for_api():
    return {
        'received': 'Received',
        'forwarded_to_cvo': 'Forwarded to CVO/DSP',
        'sent_for_permission': 'Sent for Permission',
        'permission_approved': 'Permission Approved',
        'permission_rejected': 'Permission Rejected',
        'assigned_to_inspector': 'Assigned to Inspector',
        'sent_back_for_reenquiry': 'Sent Back for Re-enquiry',
        'enquiry_in_progress': 'Enquiry In Progress',
        'enquiry_report_submitted': 'Enquiry Report Submitted',
        'forwarded_to_po': 'Forwarded to PO',
        'forwarded_to_jmd': 'Forwarded to PO',
        'action_instructed': 'Action Pending at CMD',
        'action_taken': 'Action Taken by CMD',
        'lodged': 'Lodged',
        'closed': 'Closed',
    }

def _clear_legacy_login_captcha_session():
    session.pop('login_captcha_challenges', None)
    session.pop('login_captcha_used_tokens', None)


def _mark_login_captcha_session_dirty():
    # Captcha challenges are stored in the session during login flows only
    # (before a user_id is present).  Setting session.modified ensures the
    # updated challenge data is persisted.  This is intentional here — the
    # captcha session is a pre-auth anonymous session, so there is no risk
    # of contaminating an authenticated session's modified flag.
    if has_request_context():
        session.modified = True


def _normalize_login_captcha_answer(raw_answer):
    value = re.sub(r'[^A-Za-z0-9]', '', (raw_answer or '').strip().upper())
    return value[:32]


def _login_captcha_answer_digest(token, answer):
    normalized_token = (token or '').strip()
    normalized_answer = _normalize_login_captcha_answer(answer)
    secret_key = (app.config.get('SECRET_KEY') or '').encode('utf-8')
    payload = f'{normalized_token}:{normalized_answer}'.encode('utf-8')
    return hmac.new(secret_key, payload, hashlib.sha256).hexdigest()


def _login_captcha_proof_signature(payload_b64):
    secret_key = (app.config.get('SECRET_KEY') or '').encode('utf-8')
    return hmac.new(secret_key, (payload_b64 or '').encode('ascii'), hashlib.sha256).hexdigest()


def _build_login_captcha_proof(token, answer, issued_at):
    payload = {
        'token': (token or '').strip(),
        'issued_at': int(issued_at or 0),
        'answer_digest': _login_captcha_answer_digest(token, answer),
    }
    payload_json = json.dumps(payload, separators=(',', ':')).encode('utf-8')
    payload_b64 = base64.urlsafe_b64encode(payload_json).decode('ascii').rstrip('=')
    return f'{payload_b64}.{_login_captcha_proof_signature(payload_b64)}'


def _parse_login_captcha_proof(captcha_proof):
    proof = (captcha_proof or '').strip()
    if not proof or '.' not in proof:
        return None
    payload_b64, provided_sig = proof.rsplit('.', 1)
    expected_sig = _login_captcha_proof_signature(payload_b64)
    if not hmac.compare_digest(provided_sig, expected_sig):
        return None
    padded = payload_b64 + ('=' * (-len(payload_b64) % 4))
    try:
        payload_bytes = base64.urlsafe_b64decode(padded.encode('ascii'))
        payload = json.loads(payload_bytes.decode('utf-8'))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    token = (payload.get('token') or '').strip()
    answer_digest = (payload.get('answer_digest') or '').strip()
    try:
        issued_at = int(payload.get('issued_at') or 0)
    except (TypeError, ValueError):
        return None
    if not token or not answer_digest or issued_at <= 0:
        return None
    return {
        'token': token,
        'issued_at': issued_at,
        'answer_digest': answer_digest,
    }


def _get_login_captcha_challenges_store():
    if has_request_context():
        challenges = session.get('login_captcha_challenges')
        if not isinstance(challenges, dict):
            challenges = {}
            session['login_captcha_challenges'] = challenges
        return challenges
    return LOGIN_CAPTCHA_CHALLENGES


def _get_login_captcha_used_tokens_store():
    if has_request_context():
        used_tokens = session.get('login_captcha_used_tokens')
        if not isinstance(used_tokens, dict):
            used_tokens = {}
            session['login_captcha_used_tokens'] = used_tokens
        return used_tokens
    return LOGIN_CAPTCHA_USED_TOKENS


def _cleanup_used_login_captcha_tokens(now_ts=None):
    now_ts = int(time.time() if now_ts is None else now_ts)
    stale_before = now_ts - LOGIN_CAPTCHA_TTL_SECONDS - 5
    used_tokens = _get_login_captcha_used_tokens_store()
    challenges = _get_login_captcha_challenges_store()
    stale_tokens = [token for token, seen_at in used_tokens.items() if seen_at < stale_before]
    for token in stale_tokens:
        used_tokens.pop(token, None)
    stale_challenges = [
        token for token, data in challenges.items()
        if int(data.get('issued_at') or 0) < stale_before
    ]
    for token in stale_challenges:
        challenges.pop(token, None)
    if stale_tokens or stale_challenges:
        _mark_login_captcha_session_dirty()


def _mark_login_captcha_token_used(captcha_token, now_ts=None):
    _cleanup_used_login_captcha_tokens(now_ts)
    used_tokens = _get_login_captcha_used_tokens_store()
    used_tokens[(captcha_token or '').strip()] = int(time.time() if now_ts is None else now_ts)
    _mark_login_captcha_session_dirty()


def _captcha_set_pixel(buffer, width, height, x, y, color):
    if 0 <= x < width and 0 <= y < height:
        buffer[y * width + x] = color


def _captcha_fill_rect(buffer, width, height, x, y, w, h, color):
    for yy in range(y, y + h):
        for xx in range(x, x + w):
            _captcha_set_pixel(buffer, width, height, xx, yy, color)


def _captcha_draw_line(buffer, width, height, x1, y1, x2, y2, color):
    dx = abs(x2 - x1)
    dy = -abs(y2 - y1)
    sx = 1 if x1 < x2 else -1
    sy = 1 if y1 < y2 else -1
    err = dx + dy
    x, y = x1, y1
    while True:
        _captcha_set_pixel(buffer, width, height, x, y, color)
        if x == x2 and y == y2:
            break
        e2 = 2 * err
        if e2 >= dy:
            err += dy
            x += sx
        if e2 <= dx:
            err += dx
            y += sy


def _captcha_bmp_bytes(width, height, buffer):
    row_stride = width * 3
    row_padding = (4 - (row_stride % 4)) % 4
    pixel_bytes = bytearray()
    for y in range(height - 1, -1, -1):
        offset = y * width
        for x in range(width):
            r, g, b = buffer[offset + x]
            pixel_bytes.extend((b, g, r))
        pixel_bytes.extend(b'\x00' * row_padding)
    header_size = 14 + 40
    file_size = header_size + len(pixel_bytes)
    bmp = bytearray()
    bmp.extend(b'BM')
    bmp.extend(file_size.to_bytes(4, 'little'))
    bmp.extend((0).to_bytes(4, 'little'))
    bmp.extend(header_size.to_bytes(4, 'little'))
    bmp.extend((40).to_bytes(4, 'little'))
    bmp.extend(width.to_bytes(4, 'little', signed=True))
    bmp.extend(height.to_bytes(4, 'little', signed=True))
    bmp.extend((1).to_bytes(2, 'little'))
    bmp.extend((24).to_bytes(2, 'little'))
    bmp.extend((0).to_bytes(4, 'little'))
    bmp.extend(len(pixel_bytes).to_bytes(4, 'little'))
    bmp.extend((2835).to_bytes(4, 'little', signed=True))
    bmp.extend((2835).to_bytes(4, 'little', signed=True))
    bmp.extend((0).to_bytes(4, 'little'))
    bmp.extend((0).to_bytes(4, 'little'))
    bmp.extend(pixel_bytes)
    return bytes(bmp)


def _build_login_captcha_bmp(challenge_text):
    width = 168
    height = 56
    bg = (16, 24, 40)
    fg = (248, 250, 252)
    accent = (245, 166, 35)
    alt = (125, 211, 252)
    buffer = [bg] * (width * height)
    for _ in range(8):
        _captcha_draw_line(
            buffer,
            width,
            height,
            random.randint(0, width - 1),
            random.randint(0, height - 1),
            random.randint(0, width - 1),
            random.randint(0, height - 1),
            accent if random.randint(0, 1) else alt,
        )
    segments_by_digit = {
        '0': 'abcedf',
        '1': 'bc',
        '2': 'abged',
        '3': 'abgcd',
        '4': 'fgbc',
        '5': 'afgcd',
        '6': 'afgcde',
        '7': 'abc',
        '8': 'abcdefg',
        '9': 'abcfgd',
    }
    digit = challenge_text
    digit_width = 18
    digit_height = 32
    thickness = 3
    spacing = 8
    start_x = 10
    start_y = 12
    for idx, ch in enumerate(digit):
        x = start_x + idx * (digit_width + spacing) + random.randint(-1, 1)
        y = start_y + random.randint(-3, 3)
        segs = segments_by_digit.get(ch, '')
        seg_rects = {
            'a': (x + thickness, y, digit_width - (2 * thickness), thickness),
            'b': (x + digit_width - thickness, y + thickness, thickness, (digit_height // 2) - thickness),
            'c': (x + digit_width - thickness, y + (digit_height // 2), thickness, (digit_height // 2) - thickness),
            'd': (x + thickness, y + digit_height - thickness, digit_width - (2 * thickness), thickness),
            'e': (x, y + (digit_height // 2), thickness, (digit_height // 2) - thickness),
            'f': (x, y + thickness, thickness, (digit_height // 2) - thickness),
            'g': (x + thickness, y + (digit_height // 2) - (thickness // 2), digit_width - (2 * thickness), thickness),
        }
        color = fg if idx % 2 == 0 else accent
        for seg in segs:
            _captcha_fill_rect(buffer, width, height, *seg_rects[seg], color)
    for _ in range(16):
        _captcha_fill_rect(
            buffer,
            width,
            height,
            random.randint(0, width - 4),
            random.randint(0, height - 4),
            random.randint(1, 3),
            random.randint(1, 3),
            alt if random.randint(0, 1) else accent,
        )
    return _captcha_bmp_bytes(width, height, buffer)


def _login_captcha_image_url(token):
    return f'/auth/login-captcha/{urllib.parse.quote((token or "").strip(), safe="")}'


def _login_captcha_image_data_url(token):
    challenges = _get_login_captcha_challenges_store()
    challenge = challenges.get((token or '').strip()) or {}
    image_b64 = (challenge.get('image_b64') or '').strip()
    if not image_b64:
        return ''
    return f'data:image/bmp;base64,{image_b64}'


def _login_captcha_proof(token):
    challenges = _get_login_captcha_challenges_store()
    challenge = challenges.get((token or '').strip()) or {}
    return (challenge.get('proof') or '').strip()


def generate_login_captcha(challenge_text=None, issued_at=None):
    challenge = _normalize_login_captcha_answer(challenge_text)
    if not challenge:
        challenge = ''.join(secrets.choice(LOGIN_CAPTCHA_ALPHABET) for _ in range(LOGIN_CAPTCHA_LENGTH))
    issued_at = int(time.time() if issued_at is None else issued_at)
    token = secrets.token_urlsafe(24)
    image_bytes = _build_login_captcha_bmp(challenge)
    challenges = _get_login_captcha_challenges_store()
    challenges[token] = {
        'answer_digest': _login_captcha_answer_digest(token, challenge),
        'issued_at': issued_at,
        'image_b64': base64.b64encode(image_bytes).decode('ascii'),
        'proof': _build_login_captcha_proof(token, challenge, issued_at),
    }
    _mark_login_captcha_session_dirty()
    _cleanup_used_login_captcha_tokens(issued_at)
    return _login_captcha_image_url(token), token


def reset_login_captcha():
    _clear_legacy_login_captcha_session()
    challenges = _get_login_captcha_challenges_store()
    challenges.clear()
    _mark_login_captcha_session_dirty()
    return generate_login_captcha()


def _get_existing_login_captcha():
    now_ts = int(time.time())
    _cleanup_used_login_captcha_tokens(now_ts)
    used_tokens = _get_login_captcha_used_tokens_store()
    challenges = _get_login_captcha_challenges_store()
    for token, challenge in list(challenges.items()):
        token = (token or '').strip()
        if not token or token in used_tokens:
            continue
        issued_at = int((challenge or {}).get('issued_at') or 0)
        image_b64 = ((challenge or {}).get('image_b64') or '').strip()
        proof = ((challenge or {}).get('proof') or '').strip()
        if issued_at <= 0 or issued_at > now_ts + 5:
            continue
        if now_ts - issued_at > LOGIN_CAPTCHA_TTL_SECONDS:
            continue
        if image_b64 and proof:
            return _login_captcha_image_url(token), token
    return None, None


def get_login_captcha(reuse_existing=False):
    if reuse_existing:
        captcha_image, captcha_token = _get_existing_login_captcha()
        if captcha_image and captcha_token:
            return captcha_image, captcha_token
    return reset_login_captcha()


def validate_login_captcha(raw_answer, captcha_token, captcha_proof=None):
    answer = _normalize_login_captcha_answer(raw_answer)
    token = (captcha_token or '').strip()
    if not answer or not token:
        return False
    now_ts = int(time.time())
    _cleanup_used_login_captcha_tokens(now_ts)
    used_tokens = _get_login_captcha_used_tokens_store()
    challenges = _get_login_captcha_challenges_store()
    if token in used_tokens:
        return False
    proof_payload = _parse_login_captcha_proof(captcha_proof)
    challenge = challenges.get(token)
    if proof_payload:
        if proof_payload.get('token') != token:
            return False
        issued_at = int(proof_payload.get('issued_at') or 0)
        expected_digest = (proof_payload.get('answer_digest') or '').strip()
    else:
        if not challenge:
            return False
        issued_at = int(challenge.get('issued_at') or 0)
        expected_digest = (challenge.get('answer_digest') or '').strip()
    if issued_at > now_ts + 5:
        return False
    if now_ts - issued_at > LOGIN_CAPTCHA_TTL_SECONDS:
        challenges.pop(token, None)
        _mark_login_captcha_session_dirty()
        return False
    if not expected_digest:
        challenges.pop(token, None)
        _mark_login_captcha_session_dirty()
        return False
    provided_digest = _login_captcha_answer_digest(token, answer)
    is_valid = hmac.compare_digest(provided_digest, expected_digest)
    if is_valid:
        _mark_login_captcha_token_used(token, now_ts)
        challenges.pop(token, None)
        _mark_login_captcha_session_dirty()
    return is_valid


def get_deo_office_flow(user_role, cvo_office):
    if user_role != 'data_entry':
        return None
    office = (cvo_office or '').strip().lower()
    return DEO_OFFICE_FLOW.get(office)


def get_deo_target_options(user_role, cvo_office):
    if user_role != 'data_entry':
        return []
    office = (cvo_office or '').strip().lower()
    merged = DEO_COMBINED_TARGET_FLOW.get(office)
    if merged:
        return merged
    flow = DEO_OFFICE_FLOW.get(office)
    return [flow] if flow else []


def validate_pdf_upload(file_obj, label):
    if not file_obj or not file_obj.filename:
        return True, None

    original_name = secure_filename(file_obj.filename or '')
    if not original_name:
        return False, f'{label} filename is invalid.'

    ext = original_name.rsplit('.', 1)[-1].lower() if '.' in original_name else ''
    if ext != 'pdf':
        return False, f'{label} must be PDF format.'

    file_obj.seek(0, os.SEEK_END)
    file_size = file_obj.tell()
    file_obj.seek(0)
    if file_size <= 0:
        return False, f'{label} is empty.'
    if file_size > MAX_UPLOAD_SIZE_BYTES:
        return False, f'{label} must be below {config.MAX_UPLOAD_SIZE_MB} MB.'

    header = file_obj.read(5)
    file_obj.seek(0)
    if header != b'%PDF-':
        return False, f'{label} is not a valid PDF file.'

    return True, original_name


def validate_contact(contact):
    if not contact:
        return True
    return bool(PHONE_RE.match(contact))


def validate_email(email):
    if not email:
        return True
    return bool(EMAIL_RE.match(email))


def ensure_upload_dirs():
    os.makedirs(ERECEIPT_UPLOAD_DIR, exist_ok=True)
    os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
    os.makedirs(PROFILE_UPLOAD_DIR, exist_ok=True)
    os.makedirs(HELP_RESOURCE_UPLOAD_DIR, exist_ok=True)


def _normalize_storage_relpath(path_value):
    raw = (path_value or '').replace('\\', '/').strip('/')
    if not raw:
        return None
    parts = []
    for part in raw.split('/'):
        safe_part = secure_filename(part or '')
        if not safe_part or safe_part in ('.', '..'):
            return None
        parts.append(safe_part)
    return '/'.join(parts) if parts else None


def _storage_abspath(base_dir, relpath):
    rel = _normalize_storage_relpath(relpath)
    if not rel:
        return None
    base_abs = os.path.abspath(base_dir)
    candidate = os.path.abspath(os.path.join(base_abs, rel))
    if not (candidate == base_abs or candidate.startswith(base_abs + os.sep)):
        return None
    return candidate


def _delete_uploaded_file(base_dir, relpath):
    file_path = _storage_abspath(base_dir, relpath)
    if not file_path:
        return
    try:
        if os.path.isfile(file_path):
            os.remove(file_path)
    except Exception:
        pass


def _uploaded_file_exists(base_dir, relpath):
    file_path = _storage_abspath(base_dir, relpath)
    return bool(file_path and os.path.isfile(file_path))


def _save_uploaded_file(file_obj, directory, filename, label, use_date_subdir=True):
    if not file_obj or not filename:
        return False, f'{label} upload payload is missing.'
    try:
        os.makedirs(directory, exist_ok=True)
        safe_name = secure_filename(filename or '')
        if not safe_name:
            return False, f'{label} filename is invalid.'
        relpath = safe_name
        if use_date_subdir:
            date_dir = datetime.now().strftime('%Y-%m-%d')
            relpath = f'{date_dir}/{safe_name}'
        file_path = _storage_abspath(directory, relpath)
        if not file_path:
            return False, f'{label} storage path is invalid.'
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        file_obj.save(file_path)
        if not os.path.isfile(file_path):
            app.logger.error('Upload path missing after save: %s', file_path)
            return False, f'{label} could not be stored on server.'
        return True, relpath
    except Exception:
        app.logger.exception('Failed saving uploaded file (%s) into %s', label, directory)
        return False, f'{label} could not be stored on server.'


def validate_profile_photo_upload(file_obj, user_id=None):
    if not file_obj or not file_obj.filename:
        return True, None, None

    safe_name = secure_filename(file_obj.filename)
    if not safe_name:
        return False, None, 'Profile photo filename is invalid.'

    ext = safe_name.rsplit('.', 1)[-1].lower() if '.' in safe_name else ''
    if ext not in PROFILE_PHOTO_EXTENSIONS:
        return False, None, 'Profile photo must be jpg, jpeg, png, or webp.'

    file_obj.seek(0, os.SEEK_END)
    size = file_obj.tell()
    file_obj.seek(0)
    if size <= 0:
        return False, None, 'Profile photo file is empty.'
    if size > PROFILE_PHOTO_MAX_BYTES:
        return False, None, 'Profile photo must be below 2 MB.'

    photo_user_id = user_id if user_id is not None else session.get('user_id', 'x')
    stored_name = f"user_{photo_user_id}_{uuid4().hex}.{ext}"
    return True, stored_name, None


def validate_help_resource_upload(file_obj):
    if not file_obj or not file_obj.filename:
        return True, None, None, None

    safe_name = secure_filename(file_obj.filename or '')
    if not safe_name:
        return False, None, None, 'Resource filename is invalid.'

    ext = safe_name.rsplit('.', 1)[-1].lower() if '.' in safe_name else ''
    if ext not in HELP_RESOURCE_ALLOWED_EXTENSIONS:
        allowed = ', '.join(sorted(HELP_RESOURCE_ALLOWED_EXTENSIONS))
        return False, None, None, f'Resource file type is not supported. Allowed: {allowed}.'

    file_obj.seek(0, os.SEEK_END)
    size = file_obj.tell()
    file_obj.seek(0)
    if size <= 0:
        return False, None, None, 'Resource file is empty.'
    if size > MAX_UPLOAD_SIZE_BYTES:
        return False, None, None, f'Resource file must be below {config.MAX_UPLOAD_SIZE_MB} MB.'

    stored_name = f"help_{uuid4().hex}.{ext}"
    mime_type = mimetypes.guess_type(safe_name)[0] or 'application/octet-stream'
    return True, stored_name, mime_type, None


def validate_password_strength(password, label='Password'):
    value = password or ''
    if len(value) < 8:
        return False, f'{label} must be at least 8 characters.'
    if not re.search(r'[A-Z]', value):
        return False, f'{label} must include at least one uppercase letter.'
    if not re.search(r'[a-z]', value):
        return False, f'{label} must include at least one lowercase letter.'
    if not re.search(r'[0-9]', value):
        return False, f'{label} must include at least one number.'
    if not re.search(r'[^A-Za-z0-9]', value):
        return False, f'{label} must include at least one special character.'
    return True, None


def flash_internal_error(user_message='Something went wrong. Please contact administrator.'):
    app.logger.exception(user_message)
    flash(user_message, 'danger')


def _request_prefers_json():
    if request.path.startswith('/api/'):
        return True
    best = request.accept_mimetypes.best_match(['application/json', 'text/html'])
    return bool(best == 'application/json' and request.accept_mimetypes[best] > request.accept_mimetypes['text/html'])


def _render_transient_error(status_code, title, message, retry_after_seconds=4, max_retries=4):
    retry_after_seconds = max(1, int(retry_after_seconds))
    max_retries = max(1, int(max_retries))
    if _request_prefers_json():
        payload = {
            'error': title,
            'message': message,
            'status': status_code,
            'temporary': True,
            'retry_after_seconds': retry_after_seconds,
            'max_retries': max_retries,
        }
        response = jsonify(payload)
        response.status_code = status_code
        response.headers['Retry-After'] = str(retry_after_seconds)
        return response
    response = render_template(
        'transient_error.html',
        status_code=status_code,
        title=title,
        message=message,
        retry_after_seconds=retry_after_seconds,
        max_retries=max_retries,
    )
    return response, status_code, {'Retry-After': str(retry_after_seconds)}


def _render_http_error(status_code, title, message):
    if _request_prefers_json():
        payload = {
            'error': title,
            'message': message,
            'status': int(status_code),
        }
        response = jsonify(payload)
        response.status_code = int(status_code)
        return response
    return render_template('http_error.html', status_code=int(status_code), title=title, message=message), int(status_code)


@app.errorhandler(BadRequest)
def handle_bad_request(_error):
    return _render_http_error(400, 'Bad request', 'The request could not be processed. Please verify input and try again.')


@app.errorhandler(Unauthorized)
def handle_unauthorized(_error):
    return _render_http_error(401, 'Authentication required', 'Please login and try again.')


@app.errorhandler(Forbidden)
def handle_forbidden(_error):
    return _render_http_error(403, 'Access denied', 'You do not have permission to access this resource.')


@app.errorhandler(NotFound)
def handle_not_found(_error):
    return _render_http_error(404, 'Not found', 'The requested page or resource was not found.')


@app.errorhandler(MethodNotAllowed)
def handle_method_not_allowed(_error):
    return _render_http_error(405, 'Method not allowed', 'This endpoint does not support the requested method.')


@app.errorhandler(TooManyRequests)
def handle_too_many_requests(_error):
    return _render_http_error(429, 'Too many requests', 'Request rate is too high. Please retry shortly.')


@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(_error):
    if _request_prefers_json():
        payload = {
            'error': 'Payload too large',
            'message': f'Upload exceeds {config.MAX_UPLOAD_SIZE_MB} MB limit.',
            'status': 413,
        }
        response = jsonify(payload)
        response.status_code = 413
        return response
    flash(f'Upload exceeds {config.MAX_UPLOAD_SIZE_MB} MB limit.', 'danger')
    return redirect(_safe_internal_redirect_target(request.headers.get('Referer'), fallback_endpoint='petitions_list'))


@app.errorhandler(InternalServerError)
def handle_internal_server_error(error):
    original_error = getattr(error, 'original_exception', None)
    if original_error:
        app.logger.exception('Unhandled internal server error: %s', original_error)
    else:
        app.logger.exception('Unhandled internal server error')
    return _render_transient_error(
        status_code=500,
        title='Temporary server issue',
        message='The server hit an internal error. Retrying automatically may resolve it.',
        retry_after_seconds=4,
        max_retries=4,
    )


@app.errorhandler(BadGateway)
def handle_bad_gateway(_error):
    app.logger.warning('Transient upstream error: 502 bad gateway for %s', request.path)
    return _render_transient_error(
        status_code=502,
        title='Upstream temporarily unavailable',
        message='A dependency failed to respond correctly. Auto-retrying now.',
        retry_after_seconds=3,
        max_retries=5,
    )


@app.errorhandler(ServiceUnavailable)
def handle_service_unavailable(_error):
    app.logger.warning('Transient service error: 503 service unavailable for %s', request.path)
    return _render_transient_error(
        status_code=503,
        title='Service temporarily unavailable',
        message='The service is temporarily unavailable. Auto-retrying now.',
        retry_after_seconds=3,
        max_retries=5,
    )


@app.errorhandler(GatewayTimeout)
def handle_gateway_timeout(_error):
    app.logger.warning('Transient upstream timeout: 504 gateway timeout for %s', request.path)
    return _render_transient_error(
        status_code=504,
        title='Gateway timeout',
        message='A dependency timed out. Auto-retrying now.',
        retry_after_seconds=3,
        max_retries=5,
    )


def log_security_event(event_type, severity='warning', **details):
    payload = {
        'event_type': event_type,
        'severity': severity,
        'ts': datetime.now(timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z'),
    }
    if has_request_context():
        payload.update({
            'path': request.path,
            'method': request.method,
            'ip': _client_ip(),
            'user_agent': request.headers.get('User-Agent', ''),
            'user_id': session.get('user_id'),
            'user_role': session.get('user_role'),
        })
    for key, value in details.items():
        if value is not None:
            payload[key] = value
    message = json.dumps(payload, ensure_ascii=True, default=str)
    if severity in ('critical', 'error'):
        app.logger.error(message)
    elif severity == 'info':
        app.logger.info(message)
    else:
        app.logger.warning(message)


def delete_profile_photo_file(filename):
    if not filename:
        return
    _delete_uploaded_file(PROFILE_UPLOAD_DIR, filename)


def _session_inactivity_seconds():
    return max(900, int(config.SESSION_INACTIVITY_MINUTES) * 60)


def _session_absolute_seconds():
    return max(_session_inactivity_seconds(), int(config.SESSION_ABSOLUTE_HOURS) * 3600)


def _sync_session_user_fields(user):
    # Write only fields whose value has actually changed.  The CallbackDict
    # fires _mark_modified on every __setitem__ call, even for identical
    # values, which would force a full DB save on every authenticated request
    # and cause write contention under concurrent rapid clicks.
    updates = {
        'username':        user.get('username'),
        'full_name':       user.get('full_name'),
        'user_role':       user.get('role'),
        'cvo_office':      user.get('cvo_office'),
        'phone':           user.get('phone'),
        'email':           user.get('email'),
        'profile_photo':   user.get('profile_photo'),
        'session_version': int(user.get('session_version') or 1),
    }
    for key, value in updates.items():
        if session.get(key) != value:
            session[key] = value


def refresh_session_user():
    user_id = session.get('user_id')
    if not user_id:
        return None
    user = _get_user_by_id_cached(user_id)
    if not user:
        return None
    _sync_session_user_fields(user)
    return user


def _get_user_by_id_cached(user_id):
    if not user_id:
        return None
    if not has_request_context():
        return models.get_user_by_id(user_id)

    cache = getattr(g, '_user_by_id_cache', None)
    if not isinstance(cache, dict):
        cache = {}
        g._user_by_id_cache = cache

    cache_key = int(user_id)
    if cache_key not in cache:
        cache[cache_key] = models.get_user_by_id(cache_key)
    return cache[cache_key]


def _load_current_authenticated_user(refresh_activity=True):
    if has_request_context() and getattr(g, '_current_user_loaded', False):
        return getattr(g, 'current_user', None)

    if has_request_context():
        g._current_user_loaded = True
        g.current_user = None
        g.current_user_role = None
        g.current_user_id = None
        g.auth_invalid_reason = None

    user_id = session.get('user_id')
    if not user_id:
        return None

    now_ts = int(time.time())
    issued_at = session.get('auth_issued_at')
    last_seen_at = session.get('auth_last_seen_at')

    try:
        issued_at = int(issued_at)
        last_seen_at = int(last_seen_at)
    except (TypeError, ValueError):
        _invalidate_current_session('missing_session_metadata')
        return None

    if issued_at > now_ts + 5 or last_seen_at > now_ts + 5:
        _invalidate_current_session('invalid_session_timestamp')
        return None

    if now_ts - last_seen_at > _session_inactivity_seconds():
        _invalidate_current_session('session_inactive')
        return None
    if now_ts - issued_at > _session_absolute_seconds():
        _invalidate_current_session('session_absolute_timeout')
        return None

    user = _get_user_by_id_cached(user_id)
    if not user:
        _invalidate_current_session('missing_user')
        return None
    if user.get('is_active') is False:
        _invalidate_current_session('inactive_user')
        return None

    stored_version = session.get('session_version')
    current_version = int(user.get('session_version') or 1)
    try:
        stored_version = int(stored_version)
    except (TypeError, ValueError):
        _invalidate_current_session('missing_session_version')
        return None
    if stored_version != current_version:
        _invalidate_current_session('credential_change')
        return None

    _sync_session_user_fields(user)
    if refresh_activity and now_ts - last_seen_at >= 30:
        # Update auth_last_seen_at in-memory WITHOUT triggering session.modified.
        # Writing via CallbackDict.__setitem__ fires _mark_modified → full DB
        # UPSERT on every request that crosses the 30-second threshold, which
        # creates unnecessary write contention under rapid concurrent clicks.
        #
        # Instead we bypass the callback (the value is still readable by
        # subsequent calls within this request) and set _pending_touch so
        # save_session uses the cheaper touch_server_session path.
        dict.__setitem__(session, 'auth_last_seen_at', now_ts)
        session._pending_touch = True

    if has_request_context():
        g.current_user = user
        g.current_user_role = user.get('role')
        g.current_user_id = user.get('id')
    return user


ensure_upload_dirs()


def _get_or_create_csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


def _current_csrf_token():
    if 'user_id' in session:
        return _get_or_create_csrf_token()
    return session.get('_csrf_token', '')


def _safe_internal_redirect_target(target, fallback_endpoint='dashboard'):
    fallback_url = url_for(fallback_endpoint)
    candidate = (target or '').strip()
    if not candidate:
        return fallback_url
    parsed = urllib.parse.urlsplit(candidate)
    if parsed.scheme or parsed.netloc:
        return fallback_url
    if not candidate.startswith('/'):
        return fallback_url
    if candidate.startswith('//') or '\\' in candidate:
        return fallback_url
    return candidate


def _client_ip():
    if config.TRUST_PROXY_HEADERS:
        forwarded_for = (request.headers.get('X-Forwarded-For') or '').split(',')[0].strip()
        if forwarded_for:
            return forwarded_for
    return request.remote_addr or 'unknown'


def _cleanup_login_attempts(now_ts):
    stale_before = now_ts - max(config.LOGIN_RATE_LIMIT_WINDOW_SECONDS, config.LOGIN_RATE_LIMIT_BLOCK_SECONDS) - 60
    stale_keys = [k for k, v in LOGIN_ATTEMPTS.items() if (v.get('last_seen') or 0) < stale_before]
    for key in stale_keys:
        LOGIN_ATTEMPTS.pop(key, None)


def _is_login_blocked():
    now_ts = time.time()
    _cleanup_login_attempts(now_ts)
    entry = LOGIN_ATTEMPTS.get(_client_ip(), {})
    blocked_until = float(entry.get('blocked_until') or 0)
    if blocked_until > now_ts:
        return True, int(blocked_until - now_ts)
    return False, 0


def _register_login_failure():
    now_ts = time.time()
    ip = _client_ip()
    _cleanup_login_attempts(now_ts)
    entry = LOGIN_ATTEMPTS.get(ip, {'attempts': [], 'blocked_until': 0, 'last_seen': now_ts})
    window_start = now_ts - config.LOGIN_RATE_LIMIT_WINDOW_SECONDS
    recent = [ts for ts in entry.get('attempts', []) if ts >= window_start]
    recent.append(now_ts)
    entry['attempts'] = recent
    entry['last_seen'] = now_ts
    if len(recent) >= config.LOGIN_RATE_LIMIT_MAX_ATTEMPTS:
        entry['blocked_until'] = now_ts + config.LOGIN_RATE_LIMIT_BLOCK_SECONDS
        log_security_event(
            'auth.login_lockout_triggered',
            severity='warning',
            failed_attempts=len(recent),
            blocked_seconds=config.LOGIN_RATE_LIMIT_BLOCK_SECONDS,
        )
    LOGIN_ATTEMPTS[ip] = entry
    log_security_event('auth.login_failed', severity='warning', failed_attempts=len(recent))


def _clear_login_failures():
    LOGIN_ATTEMPTS.pop(_client_ip(), None)


def _system_setting_defaults():
    return {key: int(meta['default']()) for key, meta in SYSTEM_SETTING_DEFINITIONS.items()}


def get_effective_system_settings():
    cached = getattr(g, '_effective_system_settings', None) if has_request_context() else None
    if cached is not None:
        return cached

    effective = _system_setting_defaults()
    raw_overrides = {}
    if hasattr(models, 'get_system_settings'):
        try:
            raw_overrides = models.get_system_settings(prefix='petition_') or {}
        except Exception:
            app.logger.exception('Unable to load system settings overrides; using defaults')

    for key, raw_value in raw_overrides.items():
        meta = SYSTEM_SETTING_DEFINITIONS.get(key)
        if not meta:
            continue
        try:
            value = int(str(raw_value).strip())
        except (TypeError, ValueError):
            continue
        value = max(int(meta['min']), min(int(meta['max']), value))
        effective[key] = value

    if has_request_context():
        g._effective_system_settings = effective
    return effective


def _system_settings_rows():
    effective = get_effective_system_settings()
    defaults = _system_setting_defaults()
    rows = []
    for key, meta in SYSTEM_SETTING_DEFINITIONS.items():
        rows.append({
            'key': key,
            'label': meta['label'],
            'description': meta['description'],
            'min': int(meta['min']),
            'max': int(meta['max']),
            'default': defaults[key],
            'value': effective[key],
        })
    return rows


def _petition_rate_limit_keys():
    keys = [('ip', f'ip:{_client_ip()}')]
    user_id = session.get('user_id')
    if user_id:
        keys.insert(0, ('user', f'user:{int(user_id)}'))
    return keys


def _petition_rate_limit_settings(scope):
    effective = get_effective_system_settings()
    if scope == 'user':
        return {
            'window_seconds': max(30, int(effective['petition_user_rate_limit_window_seconds'])),
            'max_submissions': max(1, int(effective['petition_user_rate_limit_max_submissions'])),
            'block_seconds': max(30, int(effective['petition_user_rate_limit_block_seconds'])),
        }
    return {
        'window_seconds': max(30, int(effective['petition_ip_rate_limit_window_seconds'])),
        'max_submissions': max(1, int(effective['petition_ip_rate_limit_max_submissions'])),
        'block_seconds': max(30, int(effective['petition_ip_rate_limit_block_seconds'])),
    }


def _petition_rate_limit_scope_entries():
    entries = []
    for scope, key in _petition_rate_limit_keys():
        settings = _petition_rate_limit_settings(scope)
        entries.append({
            'scope_type': scope,
            'scope_key': key,
            'window_seconds': settings['window_seconds'],
            'max_submissions': settings['max_submissions'],
            'block_seconds': settings['block_seconds'],
        })
    return entries


def _cleanup_petition_submission_attempts(now_ts):
    effective = get_effective_system_settings()
    max_window = max(
        int(effective['petition_user_rate_limit_window_seconds']),
        int(effective['petition_ip_rate_limit_window_seconds']),
    )
    max_block = max(
        int(effective['petition_user_rate_limit_block_seconds']),
        int(effective['petition_ip_rate_limit_block_seconds']),
    )
    stale_before = now_ts - max(max_window, max_block) - 60
    stale_keys = [k for k, v in PETITION_SUBMISSION_ATTEMPTS.items() if (v.get('last_seen') or 0) < stale_before]
    for key in stale_keys:
        PETITION_SUBMISSION_ATTEMPTS.pop(key, None)


def _consume_petition_submission_slot():
    scope_entries = _petition_rate_limit_scope_entries()
    if hasattr(models, 'consume_rate_limit'):
        try:
            result = models.consume_rate_limit('petition_submission', scope_entries)
            if isinstance(result, dict):
                return bool(result.get('allowed')), int(result.get('retry_after') or 0), list(result.get('triggered_scopes') or [])
        except Exception:
            app.logger.exception('Persistent petition rate limiter unavailable; falling back to in-memory limiter')
    now_ts = time.time()
    _cleanup_petition_submission_attempts(now_ts)
    retry_after = 0
    blocked_scopes = []
    for scope, key in _petition_rate_limit_keys():
        entry = PETITION_SUBMISSION_ATTEMPTS.get(key, {})
        blocked_until = float(entry.get('blocked_until') or 0)
        if blocked_until > now_ts:
            retry_after = max(retry_after, int(blocked_until - now_ts))
            blocked_scopes.append(scope)
    if retry_after > 0:
        return False, retry_after, blocked_scopes
    triggered = []
    for scope, key in _petition_rate_limit_keys():
        settings = _petition_rate_limit_settings(scope)
        entry = PETITION_SUBMISSION_ATTEMPTS.get(key, {'submissions': [], 'blocked_until': 0, 'last_seen': now_ts})
        window_start = now_ts - settings['window_seconds']
        recent = [ts for ts in entry.get('submissions', []) if ts >= window_start]
        recent.append(now_ts)
        entry['submissions'] = recent
        entry['last_seen'] = now_ts
        if len(recent) >= settings['max_submissions']:
            entry['blocked_until'] = now_ts + settings['block_seconds']
            triggered.append(scope)
        PETITION_SUBMISSION_ATTEMPTS[key] = entry
    return True, 0, triggered


def _can_access_petition(petition_id):
    user_id = session.get('user_id')
    if not user_id:
        return False
    try:
        pid = int(petition_id)
    except (TypeError, ValueError):
        return False
    if not hasattr(models, 'can_user_access_petition'):
        # Test stubs may not expose full data-access surface.
        return True
    user_role = session.get('user_role')
    cvo_office = session.get('cvo_office')
    try:
        return models.can_user_access_petition(user_id, user_role, cvo_office, pid)
    except Exception:
        app.logger.exception('Unable to validate petition access for petition_id=%s', pid)
        return False


def _can_access_cvo_scope(cvo_id):
    role = session.get('user_role')
    try:
        current_user_id = int(session.get('user_id') or 0)
        target_id = int(cvo_id or 0)
    except (TypeError, ValueError):
        return False
    if not current_user_id or not target_id:
        return False
    if role == 'super_admin':
        return True
    if role not in ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
        return False
    if current_user_id == target_id:
        return True

    current_user = _get_user_by_id_cached(current_user_id)
    target_user = _get_user_by_id_cached(target_id)
    if not current_user or not target_user:
        return False
    if target_user.get('role') not in ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
        return False

    current_office = (current_user.get('cvo_office') or '').strip().lower()
    target_office = (target_user.get('cvo_office') or '').strip().lower()
    if role in ('cvo_apspdcl', 'cvo_apcpdcl'):
        return current_office in ('apspdcl', 'apcpdcl') and target_office in ('apspdcl', 'apcpdcl')
    return bool(current_office and current_office == target_office)


def _petition_id_from_filename(filename):
    match = re.search(r'_(\d+)_', filename or '')
    if not match:
        return None
    return int(match.group(1))


def _build_storage_filename(prefix, original_name, petition_id=None):
    safe_prefix = secure_filename(prefix or '').strip('_')
    safe_original = secure_filename(original_name or '')
    if not safe_prefix or not safe_original:
        return None
    max_len = 240
    name_part, ext_part = os.path.splitext(safe_original)
    ext_part = ext_part.lower()[:10]
    if not ext_part:
        ext_part = ''
    token = uuid4().hex
    if len(token) > 32:
        token = token[:32]

    def _finalize(base_value):
        base_value = secure_filename(base_value or '').strip('_')
        if not base_value:
            return None
        if ext_part and not base_value.endswith(ext_part):
            keep = max(1, max_len - len(ext_part))
            base_value = base_value[:keep] + ext_part
        return base_value[:max_len]

    if petition_id is not None:
        try:
            pid = int(petition_id)
        except (TypeError, ValueError):
            pid = None
        if pid and pid > 0:
            base = f'{safe_prefix}_{pid}_{token}_'
            remaining = max_len - len(base) - len(ext_part)
            if remaining < 1:
                remaining = 1
            return _finalize(f'{base}{name_part[:remaining]}{ext_part}')
    base = f'{safe_prefix}_{token}_'
    remaining = max_len - len(base) - len(ext_part)
    if remaining < 1:
        remaining = 1
    return _finalize(f'{base}{name_part[:remaining]}{ext_part}')


def _resolve_petition_id_for_file(filename):
    petition_id = _petition_id_from_filename(filename)
    if petition_id:
        return petition_id
    if not hasattr(models, 'find_petition_id_by_filename'):
        return None
    try:
        resolved = models.find_petition_id_by_filename(filename)
        if resolved:
            return resolved
        base_name = (filename or '').replace('\\', '/').split('/')[-1]
        if base_name and base_name != filename:
            return models.find_petition_id_by_filename(base_name)
        return None
    except Exception:
        app.logger.exception('Unable to resolve petition id for file: %s', filename)
        return None


def _parse_requested_petition_id(raw_value):
    try:
        value = int((raw_value or '').strip())
        return value if value > 0 else None
    except (TypeError, ValueError, AttributeError):
        return None


def _has_pending_inspector_detailed_request(tracking_rows):
    if not tracking_rows:
        return False
    last_report_idx = -1
    for idx, row in enumerate(tracking_rows):
        if (row.get('action') or '').strip() == 'Enquiry Report Submitted':
            last_report_idx = idx
    if last_report_idx < 0:
        return False
    for row in tracking_rows[last_report_idx + 1:]:
        if (row.get('action') or '').strip() == 'Inspector Requested Detailed Enquiry Permission':
            return True
    return False


def _is_conversion_permission_stage(petition, tracking_rows):
    if not petition or (petition.get('status') or '').strip() != 'sent_for_permission':
        return False
    for row in reversed(tracking_rows or []):
        action = (row.get('action') or '').strip().lower()
        if 'requested po permission for detailed enquiry' in action:
            return True
        if 'inspector requested detailed enquiry permission' in action:
            return True
        if action in {'permission approved - sent to cvo', 'permission rejected'}:
            # Older permission-cycle events encountered first: stop scanning.
            break
    return False


def _has_conversion_request_history(tracking_rows):
    for row in tracking_rows or []:
        action = (row.get('action') or '').strip().lower()
        if 'requested po permission for detailed enquiry' in action:
            return True
        if 'inspector requested detailed enquiry permission' in action:
            return True
    return False


@app.before_request
def _security_before_request():
    _queue_proxy_mismatch_diagnostics()
    if request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
        if app.config.get('TESTING'):
            return None
        if 'user_id' in session:
            sent_token = (request.form.get('_csrf_token') or request.headers.get('X-CSRF-Token') or '').strip()
            expected_token = session.get('_csrf_token') or ''
            if not expected_token or not sent_token or not hmac.compare_digest(sent_token, expected_token):
                log_security_event('web.csrf_validation_failed', severity='warning')
                if _request_prefers_json():
                    return jsonify({'error': 'Invalid or missing CSRF token.'}), 403
                flash('Security validation failed. Please refresh and try again.', 'danger')
                return redirect(_safe_internal_redirect_target(request.referrer, fallback_endpoint='dashboard'))
    if 'user_id' in session:
        # Guard: SessionMixin.permanent is stored as session['_permanent'].
        # Calling session.permanent = True unconditionally fires __setitem__ →
        # _mark_modified on every request, forcing a DB write even when the
        # session data has not changed.  Only write when not yet set.
        if not session.get('_permanent'):
            session.permanent = True
        # _get_or_create_csrf_token writes to the session dict only when the
        # token is absent (e.g. first request after login).  Subsequent calls
        # hit the early-return branch and leave session.modified untouched.
        _get_or_create_csrf_token()
    return None


@app.after_request
def _security_after_request(response):
    # Vary: Cookie tells CDNs/proxies not to serve the same cached response
    # to different sessions (important for Safari's aggressive cache).
    response.vary.add('Cookie')
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'DENY')
    response.headers.setdefault('X-XSS-Protection', '0')
    response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    response.headers.setdefault('Permissions-Policy', 'geolocation=(), microphone=(), camera=()')
    response.headers.setdefault('Cross-Origin-Opener-Policy', 'same-origin')
    response.headers.setdefault('Cross-Origin-Resource-Policy', 'same-origin')
    if config.IS_PRODUCTION or config.TRUST_PROXY_HEADERS:
        # HSTS tells Safari (and all browsers) to always use HTTPS for this
        # origin, preventing mixed-content transitions that drop Secure cookies.
        response.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    if response.content_type and 'text/html' in response.content_type:
        response.headers.setdefault(
            'Content-Security-Policy',
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com data:; "
            "img-src 'self' data:; "
            "connect-src 'self'; "
            "base-uri 'self'; "
            "form-action 'self'; "
            "frame-ancestors 'none'"
        )
    # Prevent caching of authenticated / sensitive responses.
    if (
        'user_id' in session
        or request.path.startswith('/api/')
        or request.path == '/login'
        or request.path.startswith('/auth/login-captcha/')
    ):
        response.headers.setdefault('Cache-Control', 'no-store, no-cache, must-revalidate, private')
        response.headers.setdefault('Pragma', 'no-cache')
        response.headers.setdefault('Expires', '0')
    return _dedupe_session_cookie_headers(response)


def get_effective_form_field_configs():
    if has_request_context():
        cached_cfg = getattr(g, '_effective_form_field_configs', None)
        if isinstance(cached_cfg, dict):
            return cached_cfg

    merged = copy.deepcopy(DEFAULT_FORM_FIELD_CONFIGS)
    try:
        overrides = models.get_form_field_configs()
    except Exception:
        overrides = {}

    for key, override in overrides.items():
        if key not in merged:
            continue
        if isinstance(override, dict):
            if override.get('label'):
                merged[key]['label'] = str(override.get('label')).strip() or merged[key]['label']
            field_type = (override.get('type') or '').strip()
            if field_type in VALID_DYNAMIC_FIELD_TYPES:
                merged[key]['type'] = field_type
            merged[key]['required'] = bool(override.get('required'))
            if merged[key]['type'] == 'select':
                options = override.get('options')
                if isinstance(options, list) and options:
                    valid_options = []
                    for opt in options:
                        if not isinstance(opt, dict):
                            continue
                        value = str(opt.get('value', '')).strip()
                        label = str(opt.get('label', '')).strip()
                        if value and label:
                            valid_options.append({'value': value, 'label': label})
                    if valid_options:
                        merged[key]['options'] = valid_options
    if has_request_context():
        g._effective_form_field_configs = merged
    return merged


def get_petitions_for_user_cached(user_id, user_role, cvo_office=None, status_filter=None, enquiry_mode='all'):
    if not has_request_context():
        return models.get_petitions_for_user(user_id, user_role, cvo_office, status_filter, enquiry_mode)

    cache = getattr(g, '_petitions_for_user_cache', None)
    if not isinstance(cache, dict):
        cache = {}
        g._petitions_for_user_cache = cache

    cache_key = (user_id, user_role, cvo_office, status_filter, enquiry_mode)
    if cache_key not in cache:
        cache[cache_key] = models.get_petitions_for_user(
            user_id, user_role, cvo_office, status_filter, enquiry_mode
        )
    return cache[cache_key]


def get_form_field_config(form_key, field_key):
    key = f'{form_key}.{field_key}'
    return get_effective_form_field_configs().get(key, {'label': field_key, 'type': 'text', 'required': False, 'options': []})


def _clear_authenticated_session():
    for key in (
        'user_id',
        'username',
        'full_name',
        'user_role',
        'cvo_office',
        'phone',
        'email',
        'profile_photo',
        'session_version',
        'auth_issued_at',
        'auth_last_seen_at',
        '_csrf_token',
    ):
        session.pop(key, None)


def _begin_forced_password_change(user):
    # Do NOT hard-delete the old SID before the new one is persisted.
    # _invalidate_current_session(revoke_store=True) would delete the old row
    # immediately, but _rotate_session_identifier() registers old→new in the
    # grace map pointing at a new SID that doesn't exist in DB yet.  Any
    # concurrent request using the old cookie would find nothing for old SID,
    # follow the grace map to the new SID, find nothing there either, and get
    # a fresh anonymous session → logged out.
    #
    # Instead: clear in-memory only, rotate SID (which shortens the old row
    # to ROTATION_GRACE_SECONDS via _expire_server_session_soon), and let
    # save_session persist the new row at the end of the request.
    _invalidate_current_session(revoke_store=False)
    _rotate_session_identifier()
    session['force_change_user_id'] = user['id']
    session['force_change_username'] = user['username']
    session['force_change_role'] = user['role']


def _rotate_session_identifier():
    # Single-rotation guard: if this request already rotated the SID, do
    # nothing.  Calling rotate twice (e.g. in _begin_forced_password_change
    # followed by _activate_login_session) would register a stale grace entry
    # and leave the browser with a wrong cookie.
    if has_request_context() and getattr(g, '_session_rotated_this_request', False):
        return
    prior_sid = getattr(session, 'sid', None)
    new_sid = secrets.token_urlsafe(32)
    session.sid = new_sid
    session.modified = True
    if prior_sid and prior_sid != new_sid:
        # Do NOT immediately delete the old session row.  Any concurrent
        # request that already passed open_session with the old cookie would
        # find nothing in the DB and receive a fresh anonymous session →
        # the user gets logged out ("2-3 click bug").
        #
        # Instead:
        #  1. Register old→new in the in-process grace map so open_session
        #     can transparently forward those concurrent requests to the
        #     new SID.
        #  2. Shorten the old row's TTL to the grace window so it expires
        #     cleanly without an immediate hard delete.
        _register_rotation_grace(prior_sid, new_sid)
        _expire_server_session_soon(prior_sid, _ROTATION_GRACE_SECONDS)
    if has_request_context():
        g._session_rotated_this_request = True


def _enforce_concurrent_session_control(user_id):
    if not user_id:
        return
    current_sid = getattr(session, 'sid', None)
    if config.REVOKE_OTHER_SESSIONS_ON_LOGIN and hasattr(models, 'delete_user_server_sessions'):
        models.delete_user_server_sessions(user_id, exclude_session_id=current_sid)
        return
    if hasattr(models, 'prune_user_server_sessions'):
        models.prune_user_server_sessions(user_id, config.MAX_CONCURRENT_SESSIONS, keep_session_id=current_sid)


def _activate_login_session(user):
    # Idempotency guard: a route that calls _activate_login_session should
    # only do so once per request.  A double-call (e.g. from a retry or
    # concurrent POST reaching this code path simultaneously) would rotate the
    # SID a second time, leaving the first response with a stale cookie.
    if has_request_context() and getattr(g, '_login_session_activated', False):
        return
    if has_request_context():
        g._login_session_activated = True

    session.clear()
    _rotate_session_identifier()
    now_ts = int(time.time())
    session['user_id'] = user['id']
    _sync_session_user_fields(user)
    session['auth_issued_at'] = now_ts
    # Write auth_last_seen_at directly (no modified flag needed — the session
    # is already marked modified from the rotation and the data writes above).
    session['auth_last_seen_at'] = now_ts
    session.permanent = True
    _get_or_create_csrf_token()
    _enforce_concurrent_session_control(user['id'])


def _destroy_current_session():
    current_sid = getattr(session, 'sid', None)
    session.clear()
    if current_sid:
        _delete_server_session_record(current_sid)


def _normalize_mobile_for_otp(raw_mobile):
    return auth_normalize_mobile_for_otp(raw_mobile)


def _mask_mobile(mobile):
    return auth_mask_mobile(mobile)


def _check_internal_credentials(username, password):
    # Credentials are always verified against the local database.
    # The external API is used exclusively for OTP delivery/verification.
    user = models.authenticate_user(username, password)
    if user:
        return APIResult(True, message='Credentials verified.', payload={'source': 'local', 'user': user})
    return APIResult(False, reason='invalid_credentials', message='Invalid username or password.')


def _get_user_by_username_for_auth(username):
    getter = getattr(models, 'get_user_by_username', None)
    if callable(getter):
        return getter(username)
    if app.config.get('TESTING'):
        candidate = getattr(models, 'user', None)
        if isinstance(candidate, dict) and candidate.get('username') == username:
            return candidate
    return None


def _send_login_otp(mobile):
    if app.config.get('TESTING') or not _internal_auth_api.is_configured():
        return APIResult(True, message='OTP sent.', payload={'source': 'local'})
    return _internal_auth_api.send_otp(mobile)


def _verify_login_otp(mobile, otp_code):
    if app.config.get('TESTING') or not _internal_auth_api.is_configured():
        return APIResult(True, message='OTP verified.', payload={'source': 'local'})
    return _internal_auth_api.verify_otp(mobile, otp_code)


def _render_login_page(active_tab='secure', **extra_context):
    captcha_image, captcha_token = get_login_captcha(reuse_existing=True)
    captcha_proof = _login_captcha_proof(captcha_token)
    requested_tab = (request.args.get('tab') or active_tab or 'secure').strip().lower()
    if requested_tab not in {'secure', 'recovery'}:
        requested_tab = 'secure'
    return render_template(
        'login.html',
        captcha_image=captcha_image,
        captcha_image_data=_login_captcha_image_data_url(captcha_token),
        captcha_token=captcha_token,
        captcha_proof=captcha_proof,
        active_login_tab=requested_tab,
        **extra_context,
    )


def resolve_efile_no_for_action(petition, incoming_efile_no, required_message=None):
    existing_efile = (petition.get('efile_no') or '').strip() if petition else ''
    incoming = (incoming_efile_no or '').strip()

    if existing_efile:
        # When e-file is already set, downstream actions should always reuse it.
        # Ignore client-posted differences here; dedicated edit action enforces immutability.
        return existing_efile, None

    if not incoming:
        if required_message:
            return None, required_message
        return None, None

    if len(incoming) > 100:
        return None, 'E-Office File No is too long.'

    return incoming, None


# ========================================
# AUTH DECORATORS
# ========================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('force_change_user_id'):
            flash('You must change your password before continuing.', 'warning')
            return redirect(url_for('first_login_setup'))
        user = _load_current_authenticated_user()
        if not user:
            reason = getattr(g, 'auth_invalid_reason', None) if has_request_context() else None
            if reason == 'credential_change':
                log_security_event('auth.session_revoked_after_credential_change', severity='warning')
                flash('Your session has expired due to a credential change. Please login again.', 'warning')
            elif reason == 'inactive_user':
                log_security_event('auth.inactive_user_session_rejected', severity='warning')
                flash('Your account is inactive. Please contact administrator.', 'warning')
            elif reason in ('missing_user',):
                flash('Your session is no longer valid. Please login again.', 'warning')
            elif reason == 'session_absolute_timeout':
                flash('Your session reached the maximum allowed duration. Please login again.', 'warning')
            elif reason in ('session_inactive', 'missing_session_metadata', 'missing_session_version', 'invalid_session_timestamp'):
                flash('Your session has expired. Please login again.', 'warning')
            else:
                log_security_event('access.unauthenticated_request', severity='warning')
                flash('Please login to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            current_user = getattr(g, 'current_user', None) if has_request_context() else None
            current_role = (current_user or {}).get('role') if isinstance(current_user, dict) else None
            if not current_role:
                current_user = _load_current_authenticated_user(refresh_activity=False)
                current_role = (current_user or {}).get('role') if isinstance(current_user, dict) else None
            if not current_role or current_role not in roles:
                log_security_event('access.role_forbidden', severity='warning', required_roles=','.join(roles))
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ========================================
# CONTEXT PROCESSOR
# ========================================

@app.context_processor
def inject_globals():
    current_user = _load_current_authenticated_user(refresh_activity=False) if session.get('user_id') else None
    current_user_id = (current_user or {}).get('id') if isinstance(current_user, dict) else None
    current_user_role = (current_user or {}).get('role') if isinstance(current_user, dict) else None
    role_labels = {
        'super_admin': 'Super Admin',
        'data_entry': 'Data Entry Operator',
        'po': 'Personal Officer (Vigilance)',
        'cmd_apspdcl': 'CMD - APSPDCL',
        'cmd_apepdcl': 'CMD - APEPDCL',
        'cmd_apcpdcl': 'CMD - APCPDCL',
        'cgm_hr_transco': 'CGM/HR TRANSCO (Headquarters)',
        'dsp': 'DSP (Deputy Superintendent of Police) - Headquarters',
        'cvo_apspdcl': 'CVO/DSP - APSPDCL (Tirupathi)',
        'cvo_apepdcl': 'CVO/DSP - APEPDCL (Vizag)',
        'cvo_apcpdcl': 'CVO/DSP - APCPDCL (Vijayawada)',
        'inspector': 'Field Inspector (CI/SI)'
    }
    status_labels = {
        'received': 'Received',
        'forwarded_to_cvo': 'Forwarded to CVO/DSP',
        'sent_for_permission': 'Sent for Permission',
        'permission_approved': 'Permission Approved',
        'permission_rejected': 'Permission Rejected',
        'assigned_to_inspector': 'Assigned to Inspector',
        'sent_back_for_reenquiry': 'Sent Back for Re-enquiry',
        'enquiry_in_progress': 'Enquiry In Progress',
        'enquiry_report_submitted': 'Enquiry Report Submitted',
        'cvo_comments_added': 'CVO/DSP Comments Added',
        'forwarded_to_jmd': 'Forwarded to PO (Legacy)',
        'forwarded_to_po': 'Forwarded to PO',
        'conclusion_given': 'Conclusion Given',
        'action_instructed': 'Sent to CMD for Action',
        'action_taken': 'CMD Action Report Submitted',
        'lodged': 'Lodged',
        'closed': 'Closed'
    }
    status_colors = {
        'received': '#3b82f6',
        'forwarded_to_cvo': '#8b5cf6',
        'sent_for_permission': '#f59e0b',
        'permission_approved': '#10b981',
        'permission_rejected': '#ef4444',
        'assigned_to_inspector': '#6366f1',
        'sent_back_for_reenquiry': '#f97316',
        'enquiry_in_progress': '#0ea5e9',
        'enquiry_report_submitted': '#14b8a6',
        'cvo_comments_added': '#8b5cf6',
        'forwarded_to_jmd': '#f97316',
        'forwarded_to_po': '#ec4899',
        'conclusion_given': '#84cc16',
        'action_instructed': '#06b6d4',
        'action_taken': '#22c55e',
        'lodged': '#0ea5e9',
        'closed': '#6b7280'
    }
    workflow_stage_labels = {
        1: 'Petition Initiated',
        2: 'Enquiry in Progress',
        3: 'Report Finalized & Submitted',
        4: 'Action Pending',
        5: 'Petition Lodged',
        6: 'Petition Closed'
    }
    status_to_stage = {
        'received': 1,
        'forwarded_to_cvo': 1,
        'sent_for_permission': 1,
        'permission_approved': 1,
        'permission_rejected': 1,
        'assigned_to_inspector': 2,
        'sent_back_for_reenquiry': 2,
        'enquiry_in_progress': 2,
        'enquiry_report_submitted': 3,
        'cvo_comments_added': 3,
        'forwarded_to_po': 3,
        'forwarded_to_jmd': 3,
        'action_instructed': 4,
        'action_taken': 4,
        'lodged': 5,
        'closed': 6
    }
    petition_types = PETITION_TYPE_LABELS
    petition_sources = {
        'media': 'Electronic and Print Media',
        'public_individual': 'Public (Individual)',
        'govt': 'Govt',
        'sumoto': 'Sumoto',
        'cmd_office': 'O/o CMD',
    }
    cfg = get_effective_form_field_configs()
    govt_options = cfg.get('deo_petition.govt_institution_type', {}).get('options', [])
    govt_labels = {o.get('value'): o.get('label') for o in govt_options if isinstance(o, dict)}
    profile_photo = (current_user or {}).get('profile_photo') if isinstance(current_user, dict) else None
    current_user_name = (current_user or {}).get('full_name') if isinstance(current_user, dict) else None
    if isinstance(current_user_name, str) and current_user_name:
        normalized_user_name = re.sub(r'APS?CPDCL', 'APCPDCL', current_user_name, flags=re.IGNORECASE)
        if normalized_user_name != current_user_name:
            current_user_name = normalized_user_name
            session['full_name'] = normalized_user_name
    notification = {
        'received_count': 0,
        'pending_count': 0,
        'badge_count': 0,
        'badge_text': '0',
        'items': [],
    }
    user_id = current_user_id
    user_role = current_user_role
    if user_id and user_role:
        try:
            visible_petitions = get_petitions_for_user_cached(
                user_id, user_role, session.get('cvo_office'), status_filter=None
            )
            # Show notifications only for items that are currently in this login's queue.
            pending_in_login = [
                p for p in visible_petitions
                if p.get('status') != 'closed' and p.get('current_handler_id') == user_id
            ]
            received_petitions = [p for p in pending_in_login if p.get('status') == 'received']
            notification['received_count'] = len(received_petitions)
            notification['pending_count'] = len(pending_in_login)
            notification['badge_count'] = notification['pending_count']
            notification['badge_text'] = '9+' if notification['badge_count'] > 9 else str(notification['badge_count'])
            notification['items'] = [
                {
                    'id': p.get('id'),
                    'sno': p.get('sno') or f"#{p.get('id')}",
                    'status_label': status_labels.get(p.get('status'), str(p.get('status') or '-').replace('_', ' ').title()),
                    'subject': p.get('subject') or 'No subject',
                    'received_date': p.get('received_date').strftime('%d/%m/%Y') if p.get('received_date') else '-',
                }
                for p in pending_in_login[:6]
                if p.get('id')
            ]
        except Exception:
            pass
    return dict(
        brand_name=config.BRAND_NAME,
        brand_subtitle=config.BRAND_SUBTITLE,
        brand_logo_file=config.BRAND_LOGO_FILE,
        brand_logo_fallback=config.BRAND_LOGO_FALLBACK,
        role_labels=role_labels,
        status_labels=status_labels,
        status_colors=status_colors,
        petition_types=petition_types,
        petition_sources=petition_sources,
        govt_institution_labels=govt_labels,
        get_field_cfg=lambda form_key, field_key: cfg.get(
            f'{form_key}.{field_key}',
            {'label': field_key, 'type': 'text', 'required': False, 'options': []}
        ),
        workflow_stage_labels=workflow_stage_labels,
        status_to_stage=status_to_stage,
        current_user_role=current_user_role,
        current_user_username=(current_user or {}).get('username') if isinstance(current_user, dict) else None,
        current_user_name=current_user_name,
        current_user_id=current_user_id,
        current_user_phone=(current_user or {}).get('phone') if isinstance(current_user, dict) else None,
        current_user_email=(current_user or {}).get('email') if isinstance(current_user, dict) else None,
        current_user_profile_photo=profile_photo,
        current_user_profile_photo_url=(
            url_for('profile_photo_file', filename=profile_photo) if profile_photo else None
        ),
        csrf_token=_current_csrf_token(),
        notification=notification,
        now=datetime.now()
    )

# ========================================
# AUTH ROUTES
# ========================================

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    landing_stats = {
        'petitions_tracked': 0,
        'offices_covered': 0,
        'resolution_rate': 0,
        'active_monitoring': 0,
    }
    live_status = {
        'resolved_today': 0,
        'under_review': 0,
        'urgent_pending': 0,
    }
    try:
        petitions = models.get_all_petitions()
        today = date.today()
        total_petitions = len(petitions)
        closed_petitions = sum(1 for p in petitions if p.get('status') == 'closed')
        office_keys = set()
        review_statuses = {
            'forwarded_to_cvo',
            'sent_for_permission',
            'permission_approved',
            'assigned_to_inspector',
            'sent_back_for_reenquiry',
            'enquiry_in_progress',
            'enquiry_report_submitted',
            'cvo_comments_added',
            'forwarded_to_po',
            'forwarded_to_jmd',
            'action_instructed',
        }
        urgent_age_days = 30
        resolved_today = 0
        under_review = 0
        urgent_pending = 0

        for petition in petitions:
            for key in ('target_cvo', 'received_at'):
                raw_value = petition.get(key)
                value = str(raw_value).strip() if raw_value is not None else ''
                if value:
                    office_keys.add(value)
            status = petition.get('status')
            if status in review_statuses:
                under_review += 1
            updated_at = petition.get('updated_at')
            if status == 'closed' and updated_at and getattr(updated_at, 'date', None):
                if updated_at.date() == today:
                    resolved_today += 1
            if status != 'closed':
                received_date = petition.get('received_date')
                if received_date and ((today - received_date).days >= urgent_age_days):
                    urgent_pending += 1

        landing_stats = {
            'petitions_tracked': total_petitions,
            'offices_covered': len(office_keys),
            'resolution_rate': int(round((closed_petitions / total_petitions) * 100)) if total_petitions else 0,
            'active_monitoring': max(0, total_petitions - closed_petitions),
        }
        live_status = {
            'resolved_today': resolved_today,
            'under_review': under_review,
            'urgent_pending': urgent_pending,
        }
    except Exception:
        pass

    landing_office_orders = []
    landing_news = []
    try:
        all_resources = models.list_help_resources(active_only=True)
        for r in all_resources:
            if r.get('resource_type') == 'office_order':
                entry = dict(r)
                if r.get('storage_kind') == 'upload' and r.get('file_name'):
                    from flask import url_for as _uf
                    entry['view_url'] = _uf('help_resource_file', filename=r['file_name'])
                elif r.get('storage_kind') == 'external_url':
                    entry['view_url'] = r.get('external_url')
                else:
                    entry['view_url'] = None
                landing_office_orders.append(entry)
            elif r.get('resource_type') == 'news':
                entry = dict(r)
                if r.get('storage_kind') == 'upload' and r.get('file_name'):
                    from flask import url_for as _uf
                    entry['view_url'] = _uf('help_resource_file', filename=r['file_name'])
                elif r.get('storage_kind') == 'external_url':
                    entry['view_url'] = r.get('external_url')
                else:
                    entry['view_url'] = None
                landing_news.append(entry)
    except Exception:
        pass

    return render_template(
        'landing.html',
        landing_stats=landing_stats,
        live_status=live_status,
        landing_office_orders=landing_office_orders,
        landing_news=landing_news,
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    g._log_security_event = log_security_event
    return handle_login(
        {
            'render_login_page': _render_login_page,
            'check_internal_credentials': _check_internal_credentials,
            'send_login_otp': _send_login_otp,
            'verify_login_otp': _verify_login_otp,
            'get_user_by_username': _get_user_by_username_for_auth,
            'clear_legacy_login_captcha_session': _clear_legacy_login_captcha_session,
            'clear_login_failures': _clear_login_failures,
            'register_login_failure': _register_login_failure,
            'is_login_blocked': _is_login_blocked,
            'validate_login_captcha': validate_login_captcha,
            'reset_login_captcha': reset_login_captcha,
            'activate_login_session': _activate_login_session,
            'begin_forced_password_change': _begin_forced_password_change,
        }
    )


@app.route('/auth/login/verify', methods=['GET', 'POST'])
def login_verify_otp():
    g._log_security_event = log_security_event
    return handle_login_verify(
        {
            'send_login_otp': _send_login_otp,
            'verify_login_otp': _verify_login_otp,
            'get_user_by_username': _get_user_by_username_for_auth,
            'clear_legacy_login_captcha_session': _clear_legacy_login_captcha_session,
            'clear_login_failures': _clear_login_failures,
            'activate_login_session': _activate_login_session,
            'begin_forced_password_change': _begin_forced_password_change,
        }
    )


@app.route('/auth/login-captcha/<path:captcha_token>')
def login_captcha_image(captcha_token):
    token = (captcha_token or '').strip()
    _cleanup_used_login_captcha_tokens()
    used_tokens = _get_login_captcha_used_tokens_store()
    challenges = _get_login_captcha_challenges_store()
    if not token or token in used_tokens:
        raise NotFound()
    challenge = challenges.get(token)
    if not challenge:
        raise NotFound()
    issued_at = int(challenge.get('issued_at') or 0)
    now_ts = int(time.time())
    if issued_at > now_ts + 5 or now_ts - issued_at > LOGIN_CAPTCHA_TTL_SECONDS:
        challenges.pop(token, None)
        _mark_login_captcha_session_dirty()
        raise NotFound()
    image_b64 = (challenge.get('image_b64') or '').strip()
    if not image_b64:
        raise NotFound()
    try:
        image_bytes = base64.b64decode(image_b64, validate=True)
    except Exception:
        raise NotFound()
    response = Response(image_bytes, mimetype='image/bmp')
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    return response


@app.route('/auth/request-signup', methods=['POST'])
def request_signup():
    flash('Self signup is disabled. Contact Super Admin to create your account.', 'warning')
    return redirect(url_for('login'))


_DEFAULT_PASSWORD = 'Nigaa@123'
# ── PASSWORD RESET MODULE ────────────────────────────────────────────────────

@app.route('/auth/first-login-setup', methods=['GET', 'POST'])
def first_login_setup():
    g._log_security_event = log_security_event
    return handle_first_login_setup(
        {
            'validate_password_strength': validate_password_strength,
            'flash_internal_error': flash_internal_error,
            'update_password_and_phone': models.update_password_and_phone,
            'get_user_by_id': models.get_user_by_id,
            'activate_login_session': _activate_login_session,
            'invalidate_current_session': _invalidate_current_session,
        }
    )
    """Forced password change on first login.
    All roles, including super admin, must set password + phone.
    The phone is used for account recovery and verification.
    """
    user_id = session.get('force_change_user_id')
    if not user_id:
        return redirect(url_for('login'))

    is_super_admin = session.get('force_change_role') == 'super_admin'

    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        phone = (request.form.get('phone') or '').strip()

        if new_password == _DEFAULT_PASSWORD:
            flash('You cannot keep the default password. Please choose a new one.', 'danger')
            return redirect(url_for('first_login_setup'))

        ok, err = validate_password_strength(new_password, 'New password')
        if not ok:
            flash(err, 'danger')
            return redirect(url_for('first_login_setup'))

        if new_password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return redirect(url_for('first_login_setup'))

        if not re.fullmatch(r'[6-9]\d{9}', phone):
            flash('Please enter a valid 10-digit Indian mobile number (starts with 6–9).', 'danger')
            return redirect(url_for('first_login_setup'))

        try:
            models.update_password_and_phone(user_id, new_password, phone)
        except Exception:
            flash_internal_error('Unable to update credentials. Please try again.')
            return redirect(url_for('first_login_setup'))

        log_security_event('auth.first_login_password_changed', severity='info',
                           target_user_id=user_id)
        # Auto-login: fetch the updated user record (session_version was just incremented)
        updated_user = models.get_user_by_id(user_id)
        if updated_user and updated_user.get('is_active') is not False:
            _activate_login_session(updated_user)
            _clear_login_failures()
            log_security_event('auth.login_success', severity='info', auth_factor='first_login_setup')
            flash(f'Password updated successfully. Welcome, {updated_user["full_name"]}!', 'success')
            return redirect(url_for('dashboard'))
        # Fallback: could not load user — redirect to login
        _invalidate_current_session(revoke_store=True)
        flash('Password updated successfully. Please login with your new credentials.', 'success')
        return redirect(url_for('login'))

    return render_template('first_login_setup.html',
                           username=session.get('force_change_username', ''),
                           is_super_admin=is_super_admin)


def _check_user_for_recovery(username: str, mobile: str):
    """Call APTRANSCO /checkCred to confirm the account exists before issuing a reset OTP.
    Skips gracefully if the API is not configured (dev/test environments)."""
    if app.config.get('TESTING') or not _internal_auth_api.is_configured():
        return APIResult(True, message='User verified (API not configured).', payload={'source': 'local'})
    return _internal_auth_api.check_user_for_recovery(username, mobile)


def _invalidate_all_user_sessions(user_id: int) -> None:
    """Bump the session version so all existing server-sessions are rejected,
    then hard-delete any server-side session rows for the user."""
    if hasattr(models, 'bump_user_session_version'):
        models.bump_user_session_version(user_id)
    if hasattr(models, 'delete_user_server_sessions'):
        models.delete_user_server_sessions(user_id)


@app.route('/auth/forgot-password', methods=['POST'])
@app.route('/auth/request-recovery', methods=['POST'])
def forgot_password_request():
    g._log_security_event = log_security_event
    return handle_forgot_password_request(
        {
            'get_user_by_username': _get_user_by_username_for_auth,
            'send_login_otp': _send_login_otp,
        }
    )


@app.route('/auth/forgot-password/verify', methods=['GET', 'POST'])
def forgot_password_verify():
    g._log_security_event = log_security_event
    return handle_forgot_password_verify(
        {
            'verify_login_otp': _verify_login_otp,
        }
    )


@app.route('/auth/forgot-password/resend-otp', methods=['POST'])
def forgot_password_resend_otp():
    g._log_security_event = log_security_event
    return handle_forgot_password_resend_otp(
        {
            'send_login_otp': _send_login_otp,
        }
    )


@app.route('/auth/forgot-password/set', methods=['GET', 'POST'])
def forgot_password_set():
    g._log_security_event = log_security_event
    return handle_forgot_password_set(
        {
            'validate_password_strength': validate_password_strength,
            'flash_internal_error': flash_internal_error,
            'update_password_only': models.update_password_only,
            'invalidate_user_sessions': _invalidate_all_user_sessions,
            'activate_login_session': _activate_login_session,
            'get_user_by_id': models.get_user_by_id,
        }
    )


# ── END PASSWORD RESET MODULE ────────────────────────────────────────────────


@app.route('/logout')
def logout():
    _destroy_current_session()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/logout/all', methods=['POST'])
@login_required
def logout_all_sessions():
    user_id = session.get('user_id')
    if user_id:
        if hasattr(models, 'bump_user_session_version'):
            models.bump_user_session_version(user_id)
        if hasattr(models, 'delete_user_server_sessions'):
            models.delete_user_server_sessions(user_id)
    _destroy_current_session()
    flash('All active sessions were revoked. Please login again.', 'info')
    return redirect(url_for('login'))

# ========================================
# DASHBOARD
# ========================================

@app.route('/dashboard')
@login_required
def dashboard():
    user_role = session['user_role']
    user_id = session['user_id']
    cvo_office = session.get('cvo_office')

    petitions = get_petitions_for_user_cached(user_id, user_role, cvo_office)
    officer_lookup = {}
    for p in petitions:
        officer_id = p.get('assigned_inspector_id')
        officer_name = (p.get('inspector_name') or '').strip()
        if officer_id and officer_name:
            officer_lookup[int(officer_id)] = officer_name
    officer_options = [
        {'id': oid, 'name': name}
        for oid, name in sorted(officer_lookup.items(), key=lambda x: x[1].lower())
    ]
    dashboard_filter = _extract_dashboard_filters(request.args, officer_lookup)
    filtered_petitions = _apply_dashboard_filters(petitions, dashboard_filter)

    petition_type_labels = PETITION_TYPE_LABELS
    source_labels = {
        'media': 'Electronic and Print Media',
        'public_individual': 'Public (Individual)',
        'govt': 'Govt',
        'sumoto': 'Sumoto',
        'cmd_office': 'O/o CMD',
    }
    office_labels = {
        'jmd_office': 'PO Office',
        'cvo_apspdcl_tirupathi': 'CVO/DSP APSPDCL',
        'cvo_apepdcl_vizag': 'CVO/DSP APEPDCL',
        'cvo_apcpdcl_vijayawada': 'CVO/DSP APCPDCL',
    }
    cvo_labels = {
        'apspdcl': 'APSPDCL',
        'apepdcl': 'APEPDCL',
        'apcpdcl': 'APCPDCL',
        'headquarters': 'Headquarters',
    }
    active_filter_labels = []
    if dashboard_filter['from_date']:
        active_filter_labels.append(f"From: {dashboard_filter['from_date'].strftime('%d %b %Y')}")
    if dashboard_filter['to_date']:
        active_filter_labels.append(f"To: {dashboard_filter['to_date'].strftime('%d %b %Y')}")
    if dashboard_filter['petition_type'] != 'all':
        active_filter_labels.append(f"Type: {petition_type_labels.get(dashboard_filter['petition_type'], dashboard_filter['petition_type'])}")
    if dashboard_filter['source_of_petition'] != 'all':
        active_filter_labels.append(f"Source: {source_labels.get(dashboard_filter['source_of_petition'], dashboard_filter['source_of_petition'])}")
    if dashboard_filter['received_at'] != 'all':
        active_filter_labels.append(f"Received: {office_labels.get(dashboard_filter['received_at'], dashboard_filter['received_at'])}")
    if dashboard_filter['target_cvo'] != 'all':
        active_filter_labels.append(f"Office: {cvo_labels.get(dashboard_filter['target_cvo'], dashboard_filter['target_cvo'])}")
    if dashboard_filter['officer_id']:
        active_filter_labels.append(f"Officer: {officer_lookup.get(dashboard_filter['officer_id'], str(dashboard_filter['officer_id']))}")

    stats = _build_filtered_dashboard_stats(user_role, user_id, petitions, filtered_petitions)
    analytics = _build_dashboard_analytics(filtered_petitions, stats)

    total_items = len(filtered_petitions)
    page_size = min(100, max(10, parse_optional_int(request.args.get('page_size')) or 20))
    total_pages = max(1, (total_items + page_size - 1) // page_size)
    page = min(total_pages, max(1, parse_optional_int(request.args.get('page')) or 1))
    start = (page - 1) * page_size
    end = start + page_size
    paged_petitions = filtered_petitions[start:end]

    return render_template(
        'dashboard.html',
        stats=stats,
        petitions=paged_petitions,
        analytics=analytics,
        officer_options=officer_options,
        dashboard_filter={
            'from_date': dashboard_filter['from_date'].strftime('%Y-%m-%d') if dashboard_filter['from_date'] else '',
            'to_date': dashboard_filter['to_date'].strftime('%Y-%m-%d') if dashboard_filter['to_date'] else '',
            'petition_type': dashboard_filter['petition_type'],
            'source_of_petition': dashboard_filter['source_of_petition'],
            'received_at': dashboard_filter['received_at'],
            'target_cvo': dashboard_filter['target_cvo'],
            'officer_id': str(dashboard_filter['officer_id']) if dashboard_filter['officer_id'] else 'all',
            'page': page,
            'page_size': page_size,
        },
        dashboard_active_filter_count=len(active_filter_labels),
        dashboard_active_filter_labels=active_filter_labels,
        dashboard_pagination={
            'page': page,
            'page_size': page_size,
            'total_items': total_items,
            'total_pages': total_pages,
            'start_item': (start + 1) if total_items else 0,
            'end_item': min(end, total_items),
        }
    )


@app.route('/sla-dashboard')
@login_required
def sla_dashboard():
    user_role = session['user_role']
    user_id = session['user_id']
    cvo_office = session.get('cvo_office')
    sla_data = models.get_sla_dashboard_data_for_user(user_role, user_id, cvo_office)
    employees = sla_data.get('employees', [])
    eval_rows = sla_data.get('petitions', [])
    chart_items = employees[:20]
    employee_chart = {
        'labels': [e.get('officer_name', f"Officer {e.get('officer_id')}") for e in chart_items],
        'keys': [str(e.get('officer_id')) for e in chart_items],
        'within': [int(e.get('within') or 0) for e in chart_items],
        'beyond': [int(e.get('beyond') or 0) for e in chart_items],
        'in_progress': [int(e.get('in_progress') or 0) for e in chart_items],
    }
    sla_drilldown_rows = [
        {
            'id': int(r.get('id') or 0),
            'sno': r.get('sno') or f"#{r.get('id')}",
            'petitioner_name': r.get('petitioner_name') or '-',
            'subject': r.get('subject') or '-',
            'status': r.get('status') or '-',
            'sla_days': int(r.get('sla_days') or 0),
            'elapsed_days': int(r.get('elapsed_days') or 0),
            'sla_state': r.get('sla_state') or 'in_progress',
            'sla_bucket': r.get('sla_bucket') or ('beyond' if (r.get('sla_state') == 'beyond') else 'within'),
            'is_closed': bool(r.get('closed_at')),
            'assigned_date': (r.get('assigned_at').strftime('%Y-%m-%d') if r.get('assigned_at') else None),
            'closed_date': (r.get('closed_at').strftime('%Y-%m-%d') if r.get('closed_at') else None),
            'view_url': url_for('petition_view', petition_id=int(r.get('id') or 0)),
        }
        for r in eval_rows
        if int(r.get('id') or 0) > 0
    ]
    return render_template(
        'sla_dashboard.html',
        sla_summary=sla_data.get('summary', {}),
        sla_employees=employees,
        sla_employee_chart=employee_chart,
        sla_drilldown_rows=sla_drilldown_rows,
    )


@app.route('/sla-dashboard/employee/<int:officer_id>')
@login_required
def sla_employee_profile(officer_id):
    user_role = session['user_role']
    user_id = session['user_id']
    cvo_office = session.get('cvo_office')
    profile_data = models.get_sla_employee_profile_for_user(user_role, user_id, cvo_office, officer_id)
    if profile_data.get('unauthorized'):
        flash('You can view only SLA officers available in your login scope.', 'warning')
        return redirect(url_for('sla_dashboard'))
    officer = profile_data.get('officer') or {'id': officer_id, 'full_name': f'Officer {officer_id}', 'role': 'inspector'}
    petitions = profile_data.get('petitions', [])
    petitions_json = [
        {
            'id': int(p.get('id') or 0),
            'sno': p.get('sno') or f"#{p.get('id')}",
            'petitioner_name': p.get('petitioner_name') or '-',
            'subject': p.get('subject') or '-',
            'status': p.get('status') or '-',
            'sla_days': int(p.get('sla_days') or 0) if p.get('sla_days') is not None else '-',
            'elapsed_days': int(p.get('elapsed_days') or 0) if p.get('elapsed_days') is not None else '-',
            'sla_state': p.get('sla_state') or 'in_progress',
            'sla_bucket': p.get('sla_bucket') or ('beyond' if (p.get('sla_state') == 'beyond') else 'within'),
            'is_closed': bool(p.get('closed_at')),
            'view_url': url_for('petition_view', petition_id=int(p.get('id') or 0)),
        }
        for p in petitions
        if int(p.get('id') or 0) > 0
    ]
    return render_template(
        'sla_employee_profile.html',
        officer=officer,
        sla_summary=profile_data.get('summary', {}),
        sla_petitions=petitions,
        sla_petitions_json=petitions_json,
    )


def _extract_dashboard_filters(args, officer_lookup):
    from_date = parse_date_input(args.get('from_date'))
    to_date = parse_date_input(args.get('to_date'))
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date

    petition_type_filter = (args.get('petition_type') or 'all').strip()
    if petition_type_filter not in PETITION_TYPE_LABELS and petition_type_filter != 'all':
        petition_type_filter = 'all'
    source_filter = (args.get('source_of_petition') or 'all').strip()
    if source_filter not in VALID_SOURCE_OF_PETITION and source_filter != 'all':
        source_filter = 'all'
    received_at_filter = (args.get('received_at') or 'all').strip()
    if received_at_filter not in VALID_RECEIVED_AT and received_at_filter != 'all':
        received_at_filter = 'all'
    target_cvo_filter = (args.get('target_cvo') or 'all').strip()
    if target_cvo_filter not in VALID_TARGET_CVO and target_cvo_filter != 'all':
        target_cvo_filter = 'all'

    officer_filter = None
    officer_filter_raw = (args.get('officer_id') or 'all').strip()
    if officer_filter_raw != 'all':
        parsed_officer = parse_optional_int(officer_filter_raw)
        if parsed_officer in officer_lookup:
            officer_filter = parsed_officer

    return {
        'from_date': from_date,
        'to_date': to_date,
        'petition_type': petition_type_filter,
        'source_of_petition': source_filter,
        'received_at': received_at_filter,
        'target_cvo': target_cvo_filter,
        'officer_id': officer_filter,
    }


def _apply_dashboard_filters(petitions, filters):
    filtered = []
    for p in petitions:
        received_date = p.get('received_date')
        if filters['from_date'] and (not received_date or received_date < filters['from_date']):
            continue
        if filters['to_date'] and (not received_date or received_date > filters['to_date']):
            continue
        if filters['petition_type'] != 'all' and p.get('petition_type') != filters['petition_type']:
            continue
        if filters['source_of_petition'] != 'all' and p.get('source_of_petition') != filters['source_of_petition']:
            continue
        if filters['received_at'] != 'all' and p.get('received_at') != filters['received_at']:
            continue
        if filters['target_cvo'] != 'all' and p.get('target_cvo') != filters['target_cvo']:
            continue
        if filters['officer_id'] and int(p.get('assigned_inspector_id') or 0) != filters['officer_id']:
            continue
        filtered.append(p)
    return filtered


def _build_filtered_dashboard_stats(user_role, user_id, all_petitions, filtered_petitions):
    base_stats = {
        'total_visible': len(all_petitions),
    }
    base_stats.update(models._get_workflow_stage_stats(all_petitions))  # type: ignore[attr-defined]
    base_stats.update(models._get_sla_stats_for_petitions(all_petitions))  # type: ignore[attr-defined]
    base_stats['kpi_cards'] = models._build_role_kpi_cards(user_role, all_petitions, user_id)  # type: ignore[attr-defined]
    filtered_stats = dict(base_stats)
    filtered_stats['total_visible'] = len(filtered_petitions)
    filtered_stats['kpi_cards'] = models._build_role_kpi_cards(user_role, filtered_petitions, user_id)  # type: ignore[attr-defined]
    filtered_stats.update(models._get_workflow_stage_stats(filtered_petitions))  # type: ignore[attr-defined]

    # Keep original SLA stats when no filters are applied; otherwise recompute on filtered set.
    if len(filtered_petitions) != len(all_petitions):
        filtered_stats.update(models._get_sla_stats_for_petitions(filtered_petitions))  # type: ignore[attr-defined]
    return filtered_stats


def _build_dashboard_analytics(petitions, stats):
    status_labels = {
        'received': 'Received',
        'forwarded_to_cvo': 'Forwarded to CVO/DSP',
        'sent_for_permission': 'Sent for Permission',
        'permission_approved': 'Permission Approved',
        'permission_rejected': 'Permission Rejected',
        'assigned_to_inspector': 'Assigned to Field Officer',
        'sent_back_for_reenquiry': 'Sent Back for Re-enquiry',
        'enquiry_in_progress': 'Enquiry In Progress',
        'enquiry_report_submitted': 'Report Submitted',
        'cvo_comments_added': 'CVO/DSP Comments Added',
        'forwarded_to_po': 'Forwarded to PO',
        'action_instructed': 'Action Pending at CMD',
        'action_taken': 'Action Taken by CMD',
        'lodged': 'Lodged',
        'closed': 'Closed'
    }
    petition_type_labels = PETITION_TYPE_LABELS
    source_labels = {
        'media': 'Electronic and Print Media',
        'public_individual': 'Public',
        'govt': 'Govt',
        'sumoto': 'Sumoto',
        'cmd_office': 'O/o CMD',
    }

    now = datetime.now()
    months = []
    for i in range(5, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        key = f"{y:04d}-{m:02d}"
        months.append({
            'key': key,
            'label': datetime(y, m, 1).strftime('%b %Y'),
            'value': 0
        })
    month_index = {m['key']: m for m in months}

    status_counts = Counter()
    type_counts = Counter()
    source_counts = Counter()
    permission_mode_counts = Counter({'Direct': 0, 'Permission': 0})
    office_counts = Counter()
    officer_counts = Counter()
    officer_label_by_id = {}

    for p in petitions:
        status = p.get('status')
        if status:
            status_counts[status_labels.get(status, status.replace('_', ' ').title())] += 1

        ptype = p.get('petition_type')
        if ptype:
            type_counts[petition_type_labels.get(ptype, ptype.replace('_', ' ').title())] += 1

        source = p.get('source_of_petition')
        if source:
            source_counts[source_labels.get(source, source.replace('_', ' ').title())] += 1

        permission_mode_counts['Permission' if p.get('requires_permission') else 'Direct'] += 1

        received_at = p.get('received_at') or 'unknown'
        office_counts[str(received_at)] += 1
        officer_id = p.get('assigned_inspector_id')
        officer_name = (p.get('inspector_name') or '').strip()
        if officer_id and officer_name:
            oid = str(officer_id)
            officer_label_by_id[oid] = officer_name
            officer_counts[oid] += 1

        rd = p.get('received_date')
        if rd:
            month_key = rd.strftime('%Y-%m')
            if month_key in month_index:
                month_index[month_key]['value'] += 1

    def _counter_to_series(counter_obj, limit=8):
        series = sorted(counter_obj.items(), key=lambda x: x[1], reverse=True)
        if limit:
            series = series[:limit]
        return {
            'labels': [x[0] for x in series],
            'values': [x[1] for x in series]
        }

    officer_series = sorted(officer_counts.items(), key=lambda x: x[1], reverse=True)[:8]

    return {
        'monthly_trend': {
            'keys': [m['key'] for m in months],
            'labels': [m['label'] for m in months],
            'values': [m['value'] for m in months]
        },
        'status_split': _counter_to_series(status_counts, limit=10),
        'type_split': _counter_to_series(type_counts, limit=8),
        'source_split': _counter_to_series(source_counts, limit=8),
        'enquiry_mode_split': _counter_to_series(permission_mode_counts, limit=0),
        'office_split': _counter_to_series(office_counts, limit=6),
        'officer_split': {
            'keys': [k for k, _ in officer_series],
            'labels': [officer_label_by_id.get(k, f'Officer {k}') for k, _ in officer_series],
            'values': [v for _, v in officer_series],
        },
        'summary': {
            'total_visible': len(petitions),
            'closed': status_counts.get('Closed', 0),
            'lodged': status_counts.get('Lodged', 0),
            'active': max(0, len(petitions) - status_counts.get('Closed', 0)),
            'sla_within': stats.get('sla_within', 0),
            'sla_breached': stats.get('sla_breached', 0),
        }
    }


def _format_electrical_accident_summary(detail_row):
    if not detail_row:
        return '-'
    accident_type = (detail_row.get('accident_type') or '').strip()
    deceased_category = (detail_row.get('deceased_category') or '').strip()
    departmental_type = (detail_row.get('departmental_type') or '').strip()
    non_departmental_type = (detail_row.get('non_departmental_type') or '').strip()
    deceased_count = int(detail_row.get('deceased_count') or 0)
    general_public_count = int(detail_row.get('general_public_count') or 0)
    animals_count = int(detail_row.get('animals_count') or 0)

    accident_label = 'Fatal' if accident_type == 'fatal' else 'Non Fatal' if accident_type == 'non_fatal' else '-'
    category_label = '-'
    if deceased_category == 'departmental':
        if departmental_type == 'regular':
            category_label = f'Departmental (Regular, {max(deceased_count, 0)})'
        elif departmental_type == 'outsourced':
            category_label = f'Departmental (Outsourced, {max(deceased_count, 0)})'
        else:
            category_label = 'Departmental'
    elif deceased_category == 'non_departmental':
        if non_departmental_type in ('private_electricians', 'private'):
            category_label = f'Non Departmental (Private Electricians, {max(deceased_count, 0)})'
        elif non_departmental_type in ('contract_labour', 'contract'):
            category_label = f'Non Departmental (Contract Labour, {max(deceased_count, 0)})'
        else:
            category_label = 'Non Departmental'
    elif deceased_category == 'general_public':
        category_label = f'General Public ({max(general_public_count or deceased_count, 0)})'
    elif deceased_category == 'animals':
        category_label = f'Animals ({max(animals_count or deceased_count, 0)})'

    return f"{accident_label} | {category_label}"

# ========================================
# ANALYSIS REPORT
# ========================================

def _build_analysis_report_data(petitions):
    """Compute comprehensive analysis data from a (filtered) petition list."""
    TYPE_LABELS = {
        'bribe': 'Bribe', 'corruption': 'Corruption', 'harassment': 'Harassment',
        'electrical_accident': 'Electrical Accident', 'misconduct': 'Misconduct',
        'works_related': 'Works Related', 'irregularities_in_tenders': 'Irregularities in Tenders',
        'illegal_assets': 'Illegal Assets', 'fake_certificates': 'Fake Certificates',
        'theft_misappropriation_materials': 'Theft / Misappropriation', 'other': 'Other',
    }
    SOURCE_LABELS = {
        'media': 'Electronic & Print Media', 'public_individual': 'Public (Individual)',
        'govt': 'Government', 'sumoto': 'Sumoto', 'cmd_office': 'O/o CMD',
    }
    CVO_LABELS = {
        'apspdcl': 'APSPDCL', 'apepdcl': 'APEPDCL',
        'apcpdcl': 'APCPDCL', 'headquarters': 'Headquarters',
    }
    STATUS_LABELS = {
        'received': 'Received', 'forwarded_to_cvo': 'Forwarded to CVO/DSP',
        'sent_for_permission': 'Sent for Permission', 'permission_approved': 'Permission Approved',
        'permission_rejected': 'Permission Rejected', 'assigned_to_inspector': 'Assigned to Field Officer',
        'sent_back_for_reenquiry': 'Sent Back for Re-enquiry', 'enquiry_in_progress': 'Enquiry In Progress',
        'enquiry_report_submitted': 'Report Submitted', 'cvo_comments_added': 'CVO/DSP Comments Added',
        'forwarded_to_po': 'Forwarded to PO', 'forwarded_to_jmd': 'Forwarded to JMD',
        'action_instructed': 'Action Pending at CMD', 'action_taken': 'Action Taken',
        'lodged': 'Lodged', 'closed': 'Closed',
    }

    total = len(petitions)
    if total == 0:
        empty_trend = []
        now = datetime.now()
        for i in range(5, -1, -1):
            m = now.month - i; y = now.year
            while m <= 0: m += 12; y -= 1
            empty_trend.append({'key': f"{y:04d}-{m:02d}", 'label': datetime(y, m, 1).strftime('%b %Y'), 'value': 0})
        return {
            'total': 0, 'closed': 0, 'lodged': 0, 'active': 0,
            'resolution_rate': 0, 'sla_within': 0, 'sla_beyond': 0, 'sla_compliance': 0,
            'overdue_count': 0, 'direct_count': 0, 'permission_count': 0,
            'status_breakdown': [], 'type_breakdown': [], 'source_breakdown': [],
            'dept_stats': [], 'officer_stats': [], 'best_performers': [], 'top_defaulters': [],
            'talking_points': [], 'monthly_trend': empty_trend,
            'enquiry_types': {'preliminary': 0, 'detailed': 0},
            'accident_stats': {'total': 0, 'fatal': 0, 'non_fatal': 0},
            'dept_insights': [], 'type_insights': [], 'source_insights': [],
            'status_insights': [], 'officer_insights': [], 'sla_insights': [],
        }

    status_counter = Counter()
    type_counter = Counter()
    source_counter = Counter()
    dept_map = {}
    officer_map = {}
    direct_count = permission_count = overdue_count = 0
    prelim_count = detailed_count = 0
    accident_total = 0
    monthly_counter = Counter()

    now = datetime.now()
    months = []
    for i in range(5, -1, -1):
        m = now.month - i; y = now.year
        while m <= 0: m += 12; y -= 1
        months.append({'key': f"{y:04d}-{m:02d}", 'label': datetime(y, m, 1).strftime('%b %Y'), 'value': 0})
    month_set = {m['key'] for m in months}

    for p in petitions:
        status = p.get('status') or 'unknown'
        status_counter[status] += 1

        ptype = p.get('petition_type') or 'unknown'
        type_counter[ptype] += 1

        src = p.get('source_of_petition') or 'unknown'
        source_counter[src] += 1

        dept = p.get('target_cvo') or 'unknown'
        if dept not in dept_map:
            dept_map[dept] = {
                'key': dept, 'label': CVO_LABELS.get(dept, dept.upper()),
                'total': 0, 'closed': 0, 'lodged': 0, 'active': 0,
                'sla_within': 0, 'sla_beyond': 0,
            }
        dept_map[dept]['total'] += 1
        if status == 'closed': dept_map[dept]['closed'] += 1
        elif status == 'lodged': dept_map[dept]['lodged'] += 1
        else: dept_map[dept]['active'] += 1

        if p.get('requires_permission'): permission_count += 1
        else: direct_count += 1
        if p.get('is_overdue_escalated'): overdue_count += 1

        eq_type = p.get('enquiry_type') or 'detailed'
        if eq_type == 'preliminary': prelim_count += 1
        else: detailed_count += 1

        if ptype == 'electrical_accident':
            accident_total += 1
            # accident_type comes from enquiry_report join - approximate from petition data
            # (detailed accident breakdown requires enquiry_reports join - use count only here)

        officer_id = p.get('assigned_inspector_id')
        officer_name = (p.get('inspector_name') or '').strip()
        if officer_id and officer_name:
            oid = str(officer_id)
            if oid not in officer_map:
                officer_map[oid] = {
                    'id': officer_id, 'name': officer_name,
                    'total': 0, 'closed': 0, 'lodged': 0, 'active': 0,
                    'overdue': 0, 'sla_within': 0, 'sla_beyond': 0,
                }
            officer_map[oid]['total'] += 1
            if status == 'closed': officer_map[oid]['closed'] += 1
            elif status == 'lodged': officer_map[oid]['lodged'] += 1
            else: officer_map[oid]['active'] += 1
            if p.get('is_overdue_escalated'): officer_map[oid]['overdue'] += 1

        rd = p.get('received_date')
        if rd:
            mk = rd.strftime('%Y-%m')
            if mk in month_set: monthly_counter[mk] += 1

    for m in months:
        m['value'] = monthly_counter.get(m['key'], 0)

    # SLA analysis - join tracking data
    sla_rows = models.get_sla_evaluation_rows(petitions)
    sla_within = sla_beyond = 0
    for r in sla_rows:
        bucket = r.get('sla_bucket') or ''
        dept = r.get('target_cvo') or 'unknown'
        oid = str(r.get('assigned_inspector_id') or '')
        if bucket == 'within':
            sla_within += 1
            if dept in dept_map: dept_map[dept]['sla_within'] += 1
            if oid in officer_map: officer_map[oid]['sla_within'] += 1
        elif bucket == 'beyond':
            sla_beyond += 1
            if dept in dept_map: dept_map[dept]['sla_beyond'] += 1
            if oid in officer_map: officer_map[oid]['sla_beyond'] += 1

    closed_total = status_counter.get('closed', 0)
    lodged_total = status_counter.get('lodged', 0)
    terminal_total = closed_total + lodged_total
    active_total = total - terminal_total
    resolution_rate = round(terminal_total / total * 100, 1) if total > 0 else 0
    sla_tracked = sla_within + sla_beyond
    sla_compliance = round(sla_within / sla_tracked * 100, 1) if sla_tracked > 0 else 0

    # Status breakdown list
    status_breakdown = sorted(
        [{'status': k, 'label': STATUS_LABELS.get(k, k.replace('_', ' ').title()),
          'count': v, 'pct': round(v / total * 100, 1)}
         for k, v in status_counter.items()],
        key=lambda x: x['count'], reverse=True
    )

    # Type breakdown list
    type_breakdown = sorted(
        [{'type': k, 'label': TYPE_LABELS.get(k, k.replace('_', ' ').title()),
          'count': v, 'pct': round(v / total * 100, 1)}
         for k, v in type_counter.items()],
        key=lambda x: x['count'], reverse=True
    )

    # Source breakdown list
    source_breakdown = sorted(
        [{'source': k, 'label': SOURCE_LABELS.get(k, k.replace('_', ' ').title()),
          'count': v, 'pct': round(v / total * 100, 1)}
         for k, v in source_counter.items()],
        key=lambda x: x['count'], reverse=True
    )

    # Department stats list sorted by total
    dept_stats = sorted(dept_map.values(), key=lambda x: x['total'], reverse=True)
    for d in dept_stats:
        dt = d['sla_within'] + d['sla_beyond']
        d['sla_compliance'] = round(d['sla_within'] / dt * 100, 1) if dt > 0 else None
        d['resolution_rate'] = round((d['closed'] + d['lodged']) / d['total'] * 100, 1) if d['total'] > 0 else 0

    # Officer stats enriched
    officer_list = list(officer_map.values())
    for o in officer_list:
        ot = o['sla_within'] + o['sla_beyond']
        o['sla_compliance'] = round(o['sla_within'] / ot * 100, 1) if ot > 0 else None
        o['resolution_rate'] = round((o['closed'] + o['lodged']) / o['total'] * 100, 1) if o['total'] > 0 else 0
        o['score'] = (o['closed'] + o['lodged']) + o['sla_within'] * 0.5

    # Best performers: highest closed/lodged + SLA compliance, min 1 petition
    best_performers = sorted(
        [o for o in officer_list if o['total'] >= 1],
        key=lambda x: (x['closed'] + x['lodged'], x['sla_within'], -(x['sla_beyond'])),
        reverse=True
    )[:8]

    # Top defaulters: most sla_beyond + overdue + active
    top_defaulters = sorted(
        [o for o in officer_list if o['total'] >= 1],
        key=lambda x: (x['sla_beyond'] + x['overdue'], x['active']),
        reverse=True
    )[:8]
    # Filter to only meaningful defaulters
    top_defaulters = [o for o in top_defaulters if o['sla_beyond'] > 0 or o['overdue'] > 0 or o['active'] > 2]

    officer_stats_sorted = sorted(officer_list, key=lambda x: x['total'], reverse=True)

    # Auto-generate talking points (dicts with 'text' and 'severity': neutral/good/warn/bad)
    def _tp(text, severity='neutral'):
        return {'text': text, 'severity': severity}

    talking_points = []
    talking_points.append(_tp(f"A total of <strong>{total}</strong> petition{'s' if total != 1 else ''} {'are' if total != 1 else 'is'} covered in this report period."))
    if resolution_rate > 0:
        sev = 'good' if resolution_rate >= 70 else ('warn' if resolution_rate >= 40 else 'bad')
        talking_points.append(_tp(f"Overall resolution rate stands at <strong>{resolution_rate}%</strong> ({terminal_total} closed/lodged out of {total} petitions).", sev))
    if sla_tracked > 0:
        if sla_compliance >= 80:
            talking_points.append(_tp(f"SLA compliance is <strong>{sla_compliance}%</strong> — {sla_within} of {sla_tracked} tracked petitions resolved within SLA timelines.", 'good'))
        elif sla_compliance >= 50:
            talking_points.append(_tp(f"SLA compliance is <strong>{sla_compliance}%</strong> — {sla_beyond} petition{'s' if sla_beyond != 1 else ''} beyond SLA and need attention.", 'warn'))
        else:
            talking_points.append(_tp(f"SLA compliance is critically low at <strong>{sla_compliance}%</strong> — {sla_beyond} of {sla_tracked} tracked petitions are beyond SLA deadlines.", 'bad'))
    if overdue_count > 0:
        talking_points.append(_tp(f"<strong>{overdue_count}</strong> petition{'s' if overdue_count != 1 else ''} {'have' if overdue_count != 1 else 'has'} been escalated as overdue and require immediate action.", 'bad'))
    if active_total > 0:
        talking_points.append(_tp(f"<strong>{active_total}</strong> petition{'s are' if active_total != 1 else ' is'} currently active in the workflow pipeline."))
    if dept_stats:
        top_dept = dept_stats[0]
        talking_points.append(_tp(f"<strong>{top_dept['label']}</strong> has the highest workload with <strong>{top_dept['total']}</strong> petitions ({round(top_dept['total']/total*100,1)}% of total)."))
        if len(dept_stats) > 1:
            best_dept = max(dept_stats, key=lambda x: x.get('sla_compliance') or -1)
            if (best_dept.get('sla_compliance') or 0) > 0:
                talking_points.append(_tp(f"Best SLA performance by department: <strong>{best_dept['label']}</strong> at {best_dept['sla_compliance']}% compliance.", 'good'))
    if best_performers:
        bp = best_performers[0]
        talking_points.append(_tp(f"Top performing officer: <strong>{bp['name']}</strong> — {bp['closed'] + bp['lodged']} cases resolved out of {bp['total']} assigned.", 'good'))
    if top_defaulters:
        td = top_defaulters[0]
        talking_points.append(_tp(f"Attention required: <strong>{td['name']}</strong> has {td['sla_beyond']} beyond-SLA case{'s' if td['sla_beyond'] != 1 else ''} and {td['overdue']} overdue escalation{'s' if td['overdue'] != 1 else ''}.", 'bad'))
    if type_breakdown:
        top_type = type_breakdown[0]
        talking_points.append(_tp(f"Most common petition type: <strong>{top_type['label']}</strong> ({top_type['count']} petitions, {top_type['pct']}% of total)."))
    if source_breakdown:
        top_src = source_breakdown[0]
        talking_points.append(_tp(f"Primary source of petitions: <strong>{top_src['label']}</strong> ({top_src['count']} petitions, {top_src['pct']}% of total)."))
    if permission_count > 0 or direct_count > 0:
        talking_points.append(_tp(f"Enquiry mode: <strong>{direct_count}</strong> direct {'enquiry' if direct_count == 1 else 'enquiries'} vs <strong>{permission_count}</strong> requiring prior permission."))
    if prelim_count > 0 or detailed_count > 0:
        talking_points.append(_tp(f"Enquiry type: <strong>{prelim_count}</strong> preliminary (15-day SLA) and <strong>{detailed_count}</strong> detailed enquiries."))

    # ── Department insights
    dept_insights = []
    if dept_stats:
        if len(dept_stats) > 1:
            best_res_dept = max(dept_stats, key=lambda d: d['resolution_rate'])
            worst_res_dept = min(dept_stats, key=lambda d: d['resolution_rate'])
            most_active_dept = max(dept_stats, key=lambda d: d['active'])
            most_beyond_dept = max(dept_stats, key=lambda d: d['sla_beyond'])
            if best_res_dept['resolution_rate'] > 0:
                dept_insights.append(_tp(f"<strong>{best_res_dept['label']}</strong> leads in resolution with a <strong>{best_res_dept['resolution_rate']}%</strong> closure rate ({best_res_dept['closed'] + best_res_dept['lodged']} of {best_res_dept['total']} petitions resolved).", 'good'))
            if worst_res_dept['resolution_rate'] < 50 and worst_res_dept['total'] >= 2:
                dept_insights.append(_tp(f"<strong>{worst_res_dept['label']}</strong> has the lowest resolution rate at <strong>{worst_res_dept['resolution_rate']}%</strong> — {worst_res_dept['active']} petitions still active.", 'warn'))
            if most_active_dept['active'] > 0:
                dept_insights.append(_tp(f"<strong>{most_active_dept['label']}</strong> carries the highest active workload with <strong>{most_active_dept['active']}</strong> open cases.", 'neutral'))
            if most_beyond_dept['sla_beyond'] > 0:
                dept_insights.append(_tp(f"<strong>{most_beyond_dept['label']}</strong> has the most SLA breaches: <strong>{most_beyond_dept['sla_beyond']}</strong> petitions beyond deadline.", 'bad'))
        no_closure_depts = [d for d in dept_stats if d['closed'] + d['lodged'] == 0 and d['total'] >= 2]
        if no_closure_depts:
            names = ', '.join(d['label'] for d in no_closure_depts)
            dept_insights.append(_tp(f"No closures recorded in: <strong>{names}</strong>. All assigned petitions are still active.", 'warn'))
        perfect_sla_depts = [d for d in dept_stats if d.get('sla_compliance') == 100.0 and (d['sla_within'] + d['sla_beyond']) >= 2]
        if perfect_sla_depts:
            names = ', '.join(d['label'] for d in perfect_sla_depts)
            dept_insights.append(_tp(f"Perfect SLA compliance (100%) achieved by: <strong>{names}</strong>.", 'good'))

    # ── Type insights
    type_insights = []
    if type_breakdown:
        top_type = type_breakdown[0]
        type_insights.append(_tp(f"<strong>{top_type['label']}</strong> is the most reported category with <strong>{top_type['count']}</strong> petitions ({top_type['pct']}% of total).", 'neutral'))
        if top_type['pct'] > 50:
            type_insights.append(_tp(f"Over half of all petitions are of type <strong>{top_type['label']}</strong> — consider focused mitigation measures for this category.", 'warn'))
        if len(type_breakdown) > 1:
            second = type_breakdown[1]
            type_insights.append(_tp(f"Second most common: <strong>{second['label']}</strong> ({second['count']} petitions, {second['pct']}%).", 'neutral'))
        rare_types = [t for t in type_breakdown if t['count'] == 1]
        if rare_types:
            type_insights.append(_tp(f"{len(rare_types)} petition type{'s' if len(rare_types) != 1 else ''} ({', '.join(t['label'] for t in rare_types[:4])}) {'each have' if len(rare_types) != 1 else 'has'} only 1 petition in this period.", 'neutral'))

    # ── Source insights
    source_insights = []
    if source_breakdown:
        top_src = source_breakdown[0]
        source_insights.append(_tp(f"<strong>{top_src['label']}</strong> is the primary source with <strong>{top_src['count']}</strong> petitions ({top_src['pct']}%).", 'neutral'))
        if top_src['pct'] > 60:
            source_insights.append(_tp(f"<strong>{top_src['pct']}%</strong> of petitions originate from a single source (<strong>{top_src['label']}</strong>). Consider monitoring other channels.", 'warn'))
        if len(source_breakdown) > 1:
            second_src = source_breakdown[1]
            source_insights.append(_tp(f"Second source: <strong>{second_src['label']}</strong> with {second_src['count']} petitions ({second_src['pct']}%).", 'neutral'))
        media_count = source_counter.get('media', 0)
        public_count = source_counter.get('public_individual', 0)
        if media_count > 0 and public_count > 0:
            source_insights.append(_tp(f"Media-reported petitions: <strong>{media_count}</strong>. Public (individual) petitions: <strong>{public_count}</strong>.", 'neutral'))

    # ── Status pipeline insights
    EARLY_STAGES = {'received', 'forwarded_to_cvo', 'sent_for_permission', 'permission_approved'}
    ENQUIRY_STAGES = {'assigned_to_inspector', 'enquiry_in_progress', 'sent_back_for_reenquiry'}
    REPORTING_STAGES = {'enquiry_report_submitted', 'cvo_comments_added'}
    DECISION_STAGES = {'forwarded_to_po', 'forwarded_to_jmd', 'action_instructed', 'action_taken'}
    early_count_s = sum(status_counter.get(s, 0) for s in EARLY_STAGES)
    enquiry_count_s = sum(status_counter.get(s, 0) for s in ENQUIRY_STAGES)
    reporting_count_s = sum(status_counter.get(s, 0) for s in REPORTING_STAGES)
    decision_count_s = sum(status_counter.get(s, 0) for s in DECISION_STAGES)
    reenquiry_count = status_counter.get('sent_back_for_reenquiry', 0)
    perm_rejected_count = status_counter.get('permission_rejected', 0)
    action_instructed_count = status_counter.get('action_instructed', 0)

    status_insights = []
    if early_count_s > 0:
        sev = 'warn' if early_count_s > total * 0.3 else 'neutral'
        status_insights.append(_tp(f"<strong>{early_count_s}</strong> petition{'s are' if early_count_s != 1 else ' is'} in early intake/forwarding stages — these haven't yet reached field investigation.", sev))
    if enquiry_count_s > 0:
        status_insights.append(_tp(f"<strong>{enquiry_count_s}</strong> petition{'s are' if enquiry_count_s != 1 else ' is'} currently under active field enquiry.", 'neutral'))
    if reenquiry_count > 0:
        status_insights.append(_tp(f"<strong>{reenquiry_count}</strong> petition{'s were' if reenquiry_count != 1 else ' was'} sent back for re-enquiry — indicates initial report was insufficient.", 'warn'))
    if perm_rejected_count > 0:
        status_insights.append(_tp(f"<strong>{perm_rejected_count}</strong> petition{'s had' if perm_rejected_count != 1 else ' had'} permission rejected. These may need review for alternate handling.", 'warn'))
    if reporting_count_s > 0:
        status_insights.append(_tp(f"<strong>{reporting_count_s}</strong> petition{'s have' if reporting_count_s != 1 else ' has'} field reports submitted and are awaiting CVO/senior review.", 'neutral'))
    if action_instructed_count > 0:
        status_insights.append(_tp(f"<strong>{action_instructed_count}</strong> petition{'s are' if action_instructed_count != 1 else ' is'} pending action at CMD level.", 'warn'))
    if decision_count_s > 0:
        status_insights.append(_tp(f"<strong>{decision_count_s}</strong> petition{'s are' if decision_count_s != 1 else ' is'} at the final decision/action stage.", 'good'))
    if terminal_total > 0:
        status_insights.append(_tp(f"<strong>{closed_total}</strong> petition{'s' if closed_total != 1 else ''} closed and <strong>{lodged_total}</strong> lodged — total terminal: {terminal_total}.", 'good'))

    # ── Officer insights
    officer_insights = []
    if officer_list:
        avg_petitions = round(sum(o['total'] for o in officer_list) / len(officer_list), 1)
        officer_insights.append(_tp(f"<strong>{len(officer_list)}</strong> officers have assigned petitions. Average workload: <strong>{avg_petitions}</strong> petitions per officer.", 'neutral'))
        overdue_officers = [o for o in officer_list if o['overdue'] > 0]
        if overdue_officers:
            names = ', '.join(o['name'] for o in overdue_officers[:4])
            officer_insights.append(_tp(f"<strong>{len(overdue_officers)}</strong> officer{'s have' if len(overdue_officers) != 1 else ' has'} overdue escalations: <strong>{names}</strong>{'...' if len(overdue_officers) > 4 else ''}.", 'bad'))
        all_resolved_officers = [o for o in officer_list if o['total'] >= 1 and o['active'] == 0 and (o['closed'] + o['lodged']) > 0]
        if all_resolved_officers:
            names = ', '.join(o['name'] for o in all_resolved_officers[:4])
            officer_insights.append(_tp(f"{len(all_resolved_officers)} officer{'s have' if len(all_resolved_officers) != 1 else ' has'} resolved all assigned petitions with zero active cases: <strong>{names}</strong>.", 'good'))
        no_resolution_officers = [o for o in officer_list if o['total'] >= 2 and (o['closed'] + o['lodged']) == 0]
        if no_resolution_officers:
            names = ', '.join(o['name'] for o in no_resolution_officers[:4])
            officer_insights.append(_tp(f"<strong>{len(no_resolution_officers)}</strong> officer{'s have' if len(no_resolution_officers) != 1 else ' has'} no closures despite having 2+ assigned cases: <strong>{names}</strong>.", 'warn'))
        beyond_sla_officers = [o for o in officer_list if o['sla_beyond'] > 0]
        if beyond_sla_officers:
            total_beyond = sum(o['sla_beyond'] for o in beyond_sla_officers)
            officer_insights.append(_tp(f"<strong>{len(beyond_sla_officers)}</strong> officer{'s account' if len(beyond_sla_officers) != 1 else ' accounts'} for all {total_beyond} beyond-SLA cases.", 'warn'))

    # ── SLA insights
    sla_insights = []
    untracked = total - sla_tracked
    if sla_tracked == 0:
        sla_insights.append(_tp("No SLA tracking data is available for this report period. SLA compliance cannot be assessed.", 'neutral'))
    else:
        if untracked > 0:
            sla_insights.append(_tp(f"<strong>{untracked}</strong> petition{'s' if untracked != 1 else ''} ({round(untracked/total*100,1)}%) do not yet have SLA tracking data (may still be in early stages).", 'neutral'))
        if sla_compliance == 100.0:
            sla_insights.append(_tp(f"Perfect SLA compliance — all <strong>{sla_tracked}</strong> tracked petitions were resolved within their SLA deadlines.", 'good'))
        elif sla_beyond > 0:
            sla_insights.append(_tp(f"<strong>{sla_beyond}</strong> petition{'s have' if sla_beyond != 1 else ' has'} breached SLA deadlines. Immediate follow-up is recommended.", 'bad'))
        if prelim_count > 0:
            sla_insights.append(_tp(f"<strong>{prelim_count}</strong> preliminary enquiry case{'s' if prelim_count != 1 else ''} subject to the 15-day SLA window.", 'neutral'))
        if detailed_count > 0:
            sla_insights.append(_tp(f"<strong>{detailed_count}</strong> detailed enquiry case{'s' if detailed_count != 1 else ''} subject to the 45–90 day SLA window.", 'neutral'))
        if dept_stats:
            worst_sla_dept = max(dept_stats, key=lambda d: d['sla_beyond'])
            if worst_sla_dept['sla_beyond'] > 0:
                sla_insights.append(_tp(f"Department with most SLA breaches: <strong>{worst_sla_dept['label']}</strong> with {worst_sla_dept['sla_beyond']} beyond-deadline case{'s' if worst_sla_dept['sla_beyond'] != 1 else ''}.", 'bad'))

    return {
        'total': total,
        'closed': closed_total,
        'lodged': lodged_total,
        'active': active_total,
        'terminal': terminal_total,
        'resolution_rate': resolution_rate,
        'sla_within': sla_within,
        'sla_beyond': sla_beyond,
        'sla_tracked': sla_tracked,
        'sla_compliance': sla_compliance,
        'overdue_count': overdue_count,
        'direct_count': direct_count,
        'permission_count': permission_count,
        'enquiry_types': {'preliminary': prelim_count, 'detailed': detailed_count},
        'status_breakdown': status_breakdown,
        'type_breakdown': type_breakdown,
        'source_breakdown': source_breakdown,
        'dept_stats': dept_stats,
        'officer_stats': officer_stats_sorted,
        'best_performers': best_performers,
        'top_defaulters': top_defaulters,
        'talking_points': talking_points,
        'monthly_trend': months,
        'dept_insights': dept_insights,
        'type_insights': type_insights,
        'source_insights': source_insights,
        'status_insights': status_insights,
        'officer_insights': officer_insights,
        'sla_insights': sla_insights,
    }


@app.route('/analysis-report')
@login_required
def analysis_report():
    user_role = session['user_role']
    user_id = session['user_id']
    cvo_office = session.get('cvo_office')

    petitions = get_petitions_for_user_cached(user_id, user_role, cvo_office)
    officer_lookup = {}
    for p in petitions:
        oid = p.get('assigned_inspector_id')
        oname = (p.get('inspector_name') or '').strip()
        if oid and oname:
            officer_lookup[int(oid)] = oname
    officer_options = [
        {'id': oid, 'name': name}
        for oid, name in sorted(officer_lookup.items(), key=lambda x: x[1].lower())
    ]

    dashboard_filter = _extract_dashboard_filters(request.args, officer_lookup)
    filtered = _apply_dashboard_filters(petitions, dashboard_filter)

    report_data = _build_analysis_report_data(filtered)

    # Build active filter labels (same as dashboard)
    source_labels = {'media': 'Electronic and Print Media', 'public_individual': 'Public (Individual)',
                     'govt': 'Govt', 'sumoto': 'Sumoto', 'cmd_office': 'O/o CMD'}
    office_labels = {'jmd_office': 'PO Office', 'cvo_apspdcl_tirupathi': 'CVO/DSP APSPDCL',
                     'cvo_apepdcl_vizag': 'CVO/DSP APEPDCL', 'cvo_apcpdcl_vijayawada': 'CVO/DSP APCPDCL'}
    cvo_labels = {'apspdcl': 'APSPDCL', 'apepdcl': 'APEPDCL', 'apcpdcl': 'APCPDCL', 'headquarters': 'Headquarters'}
    active_filter_labels = []
    if dashboard_filter['from_date']:
        active_filter_labels.append(f"From: {dashboard_filter['from_date'].strftime('%d %b %Y')}")
    if dashboard_filter['to_date']:
        active_filter_labels.append(f"To: {dashboard_filter['to_date'].strftime('%d %b %Y')}")
    if dashboard_filter['petition_type'] != 'all':
        active_filter_labels.append(f"Type: {PETITION_TYPE_LABELS.get(dashboard_filter['petition_type'], dashboard_filter['petition_type'])}")
    if dashboard_filter['source_of_petition'] != 'all':
        active_filter_labels.append(f"Source: {source_labels.get(dashboard_filter['source_of_petition'], dashboard_filter['source_of_petition'])}")
    if dashboard_filter['received_at'] != 'all':
        active_filter_labels.append(f"Received at: {office_labels.get(dashboard_filter['received_at'], dashboard_filter['received_at'])}")
    if dashboard_filter['target_cvo'] != 'all':
        active_filter_labels.append(f"Office: {cvo_labels.get(dashboard_filter['target_cvo'], dashboard_filter['target_cvo'])}")
    if dashboard_filter['officer_id']:
        active_filter_labels.append(f"Officer: {officer_lookup.get(dashboard_filter['officer_id'], str(dashboard_filter['officer_id']))}")

    generated_at = datetime.now().strftime('%d %b %Y, %I:%M %p')
    period_label = 'All Time'
    if dashboard_filter['from_date'] and dashboard_filter['to_date']:
        period_label = f"{dashboard_filter['from_date'].strftime('%d %b %Y')} – {dashboard_filter['to_date'].strftime('%d %b %Y')}"
    elif dashboard_filter['from_date']:
        period_label = f"From {dashboard_filter['from_date'].strftime('%d %b %Y')}"
    elif dashboard_filter['to_date']:
        period_label = f"Until {dashboard_filter['to_date'].strftime('%d %b %Y')}"

    return render_template(
        'analysis_report.html',
        report=report_data,
        officer_options=officer_options,
        active_filter_labels=active_filter_labels,
        period_label=period_label,
        generated_at=generated_at,
        dashboard_filter={
            'from_date': dashboard_filter['from_date'].strftime('%Y-%m-%d') if dashboard_filter['from_date'] else '',
            'to_date': dashboard_filter['to_date'].strftime('%Y-%m-%d') if dashboard_filter['to_date'] else '',
            'petition_type': dashboard_filter['petition_type'],
            'source_of_petition': dashboard_filter['source_of_petition'],
            'received_at': dashboard_filter['received_at'],
            'target_cvo': dashboard_filter['target_cvo'],
            'officer_id': str(dashboard_filter['officer_id']) if dashboard_filter['officer_id'] else 'all',
        },
    )


# ========================================
# PETITION ROUTES
# ========================================

@app.route('/petitions')
@login_required
def petitions_list():
    status_filter = request.args.get('status', 'all')
    enquiry_mode = request.args.get('mode', 'all')
    user_role = session['user_role']
    user_id = session['user_id']
    if status_filter == 'beyond_sla':
        enquiry_mode = 'all'
    
    if user_role == 'super_admin':
        petitions = models.get_all_petitions(status_filter, enquiry_mode)
    else:
        petitions = models.get_petitions_for_user(user_id, user_role, session.get('cvo_office'), status_filter, enquiry_mode)

    sla_eval_map = {}
    if petitions:
        try:
            sla_eval_map = {
                int(row.get('id')): row
                for row in models.get_sla_evaluation_rows(petitions)
                if row.get('id')
            }
        except Exception:
            sla_eval_map = {}
    if status_filter == 'beyond_sla' and sla_eval_map:
        petitions = sorted(
            petitions,
            key=lambda p: (
                -(int((sla_eval_map.get(int(p.get('id') or 0)) or {}).get('elapsed_days') or 0)),
                -int(p.get('id') or 0),
            )
        )

    accident_detail_map = {}
    petition_ids = [int(p.get('id')) for p in petitions if p.get('id')]
    if petition_ids:
        try:
            accident_detail_map = models.get_latest_enquiry_report_accident_details(petition_ids) or {}
        except Exception:
            accident_detail_map = {}

    return render_template(
        'petitions_list.html',
        petitions=petitions,
        status_filter=status_filter,
        enquiry_mode=enquiry_mode,
        accident_detail_map=accident_detail_map,
        sla_eval_map=sla_eval_map,
        show_beyond_sla_tab=(user_role in ('po', 'super_admin'))
    )

@app.route('/petitions/new', methods=['GET', 'POST'])
@login_required
@role_required('super_admin', 'data_entry')
def petition_new():
    deo_flow = get_deo_office_flow(session.get('user_role'), session.get('cvo_office'))
    deo_target_options = get_deo_target_options(session.get('user_role'), session.get('cvo_office'))
    deo_target_map = {opt.get('target_cvo'): opt for opt in deo_target_options if isinstance(opt, dict)}
    show_cmd_source_option = (
        session.get('user_role') == 'data_entry' and
        (session.get('cvo_office') or '').strip().lower() in ('apepdcl', 'apspdcl')
    )

    def render_petition_form():
        return render_template(
            'petition_form.html',
            deo_flow=deo_flow,
            deo_target_options=deo_target_options,
            show_cmd_source_option=show_cmd_source_option,
        )

    if request.method == 'POST':
        ereceipt_no = request.form.get('ereceipt_no', '').strip() or None
        ereceipt_file = request.files.get('ereceipt_file')
        ereceipt_filename = None

        petitioner_name = request.form.get('petitioner_name', '').strip()
        contact = request.form.get('contact', '').strip()
        place = request.form.get('place', '').strip()
        petitioner_identity_type = (request.form.get('petitioner_identity_type') or 'identified').strip().lower()
        subject = request.form.get('subject', '').strip()
        petition_type = (request.form.get('petition_type') or '').strip()
        source_of_petition = (request.form.get('source_of_petition') or '').strip()
        govt_institution_type = (request.form.get('govt_institution_type') or '').strip()
        received_at = (request.form.get('received_at') or '').strip()
        target_cvo = (request.form.get('target_cvo') or '').strip()
        permission_request_type = (request.form.get('permission_request_type') or '').strip()

        if session.get('user_role') == 'data_entry':
            if not deo_target_options:
                flash('DEO office mapping is missing. Please contact admin.', 'danger')
                return render_petition_form()
            selected_target = (target_cvo or '').strip()
            if len(deo_target_options) == 1 and not selected_target:
                selected_target = deo_target_options[0].get('target_cvo')
            selected_flow = deo_target_map.get(selected_target)
            if not selected_flow:
                flash('Please select a valid target CVO/DSP office.', 'warning')
                return render_petition_form()
            received_at = selected_flow['received_at']
            target_cvo = selected_flow['target_cvo']
            permission_request_type = 'permission_required' if selected_flow.get('force_permission_required') else 'direct_enquiry'

        is_jmd_received = (received_at == 'jmd_office')
        received_date_raw = request.form.get('received_date')
        received_date = parse_date_input(received_date_raw)
        remarks = request.form.get('remarks', '').strip()
        petition_cfg = get_effective_form_field_configs()
        cfg_received_date = petition_cfg.get('deo_petition.received_date', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.received_date'])
        cfg_received_at = petition_cfg.get('deo_petition.received_at', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.received_at'])
        cfg_ereceipt_no = petition_cfg.get('deo_petition.ereceipt_no', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.ereceipt_no'])
        cfg_ereceipt_file = petition_cfg.get('deo_petition.ereceipt_file', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.ereceipt_file'])
        cfg_target_cvo = petition_cfg.get('deo_petition.target_cvo', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.target_cvo'])
        cfg_permission_request = petition_cfg.get('deo_petition.permission_request_type', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.permission_request_type'])
        cfg_petitioner = petition_cfg.get('deo_petition.petitioner_name', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.petitioner_name'])
        cfg_contact = petition_cfg.get('deo_petition.contact', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.contact'])
        cfg_place = petition_cfg.get('deo_petition.place', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.place'])
        cfg_subject = petition_cfg.get('deo_petition.subject', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.subject'])
        cfg_petition_type = petition_cfg.get('deo_petition.petition_type', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.petition_type'])
        cfg_source = petition_cfg.get('deo_petition.source_of_petition', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.source_of_petition'])
        cfg_remarks = petition_cfg.get('deo_petition.remarks', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.remarks'])
        cfg_govt_institution = petition_cfg.get('deo_petition.govt_institution_type', DEFAULT_FORM_FIELD_CONFIGS['deo_petition.govt_institution_type'])
        if petitioner_identity_type not in ('identified', 'anonymous'):
            petitioner_identity_type = 'identified'
        if petitioner_identity_type == 'anonymous':
            petitioner_name = ''
            contact = ''
            place = ''

        if cfg_received_date.get('required') and not received_date:
            flash(f"{cfg_received_date.get('label', 'Received Date')} is required.", 'warning')
            return render_petition_form()
        if received_date_raw and not received_date:
            flash(f"Please provide a valid {cfg_received_date.get('label', 'Received Date').lower()}.", 'warning')
            return render_petition_form()
        if cfg_received_at.get('required') and not received_at:
            flash(f"{cfg_received_at.get('label', 'Received At')} is required.", 'warning')
            return render_petition_form()
        if received_at not in VALID_RECEIVED_AT:
            flash(f"Please select a valid {cfg_received_at.get('label', 'Received At')}.", 'warning')
            return render_petition_form()
        if not subject:
            flash(f"{cfg_subject.get('label', 'Subject')} is required.", 'warning')
            return render_petition_form()
        if cfg_petition_type.get('required') and not petition_type:
            flash(f"{cfg_petition_type.get('label', 'Type of Petition')} is required.", 'warning')
            return render_petition_form()
        if petition_type not in VALID_PETITION_TYPES:
            flash(f"Please select a valid {cfg_petition_type.get('label', 'Type of Petition')}.", 'warning')
            return render_petition_form()
        if cfg_source.get('required') and not source_of_petition:
            flash(f"{cfg_source.get('label', 'Source of Petition')} is required.", 'warning')
            return render_petition_form()
        if source_of_petition not in VALID_SOURCE_OF_PETITION:
            flash(f"Please select a valid {cfg_source.get('label', 'Source of Petition')}.", 'warning')
            return render_petition_form()
        if source_of_petition == 'cmd_office' and not show_cmd_source_option:
            flash('O/o CMD source is allowed only for APSPDCL/APEPDCL DEO login.', 'warning')
            return render_petition_form()
        govt_option_values = {o.get('value') for o in cfg_govt_institution.get('options', []) if isinstance(o, dict)}
        if source_of_petition == 'govt' and cfg_govt_institution.get('required') and not govt_institution_type:
            flash(f"Please select {cfg_govt_institution.get('label', 'Type of Institution')}.", 'warning')
            return render_petition_form()
        if source_of_petition == 'govt' and govt_institution_type and govt_institution_type not in govt_option_values:
            flash('Please select a valid Govt institution type.', 'warning')
            return render_petition_form()
        if petitioner_identity_type == 'identified' and cfg_petitioner.get('required') and not petitioner_name:
            flash(f"{cfg_petitioner.get('label', 'Petitioner Name')} is required.", 'warning')
            return render_petition_form()
        if petitioner_identity_type == 'identified' and cfg_contact.get('required') and not contact:
            flash(f"{cfg_contact.get('label', 'Contact Number')} is required.", 'warning')
            return render_petition_form()
        if petitioner_identity_type == 'identified' and cfg_place.get('required') and not place:
            flash(f"{cfg_place.get('label', 'Place')} is required.", 'warning')
            return render_petition_form()
        if cfg_remarks.get('required') and not remarks:
            flash(f"{cfg_remarks.get('label', 'Remarks')} is required.", 'warning')
            return render_petition_form()
        if not is_jmd_received:
            if cfg_permission_request.get('required') and not permission_request_type:
                flash(f"{cfg_permission_request.get('label', 'Permission Request')} is required.", 'warning')
                return render_petition_form()
            if permission_request_type not in VALID_PERMISSION_REQUEST_TYPES:
                flash(f"Please select a valid {cfg_permission_request.get('label', 'Permission Request')}.", 'warning')
                return render_petition_form()
            if cfg_target_cvo.get('required') and not target_cvo:
                flash(f"{cfg_target_cvo.get('label', 'Target CVO/DSP Jurisdiction')} is required.", 'warning')
                return render_petition_form()
            if target_cvo not in VALID_TARGET_CVO:
                flash(f"Please select a valid {cfg_target_cvo.get('label', 'Target CVO/DSP Jurisdiction')}.", 'warning')
                return render_petition_form()
        if petitioner_name and len(petitioner_name) > 255:
            flash('Petitioner name is too long.', 'warning')
            return render_petition_form()
        if len(subject) > 5000:
            flash('Subject is too long.', 'warning')
            return render_petition_form()
        if len(place) > 255:
            flash('Place is too long.', 'warning')
            return render_petition_form()
        if petitioner_identity_type == 'identified' and not validate_contact(contact):
            flash('Please provide a valid contact number.', 'warning')
            return render_petition_form()
        if cfg_ereceipt_file.get('required') and (not ereceipt_file or not ereceipt_file.filename):
            flash(f"{cfg_ereceipt_file.get('label', 'E-Receipt File')} is required.", 'warning')
            return render_petition_form()
        if ereceipt_no and len(ereceipt_no) > 100:
            flash(f"{cfg_ereceipt_no.get('label', 'E-Receipt No')} is too long.", 'warning')
            return render_petition_form()
        if ereceipt_no and (not ereceipt_file or not ereceipt_file.filename):
            flash('E-Receipt file is required when E-Receipt No is provided. Please choose the PDF again and submit.', 'warning')
            return render_petition_form()
        if (ereceipt_file and ereceipt_file.filename) and not ereceipt_no:
            flash('E-Receipt No is required when uploading E-Receipt file.', 'warning')
            return render_petition_form()
        if len(remarks) > 5000:
            flash('Remarks are too long.', 'warning')
            return render_petition_form()

        if ereceipt_file and ereceipt_file.filename:
            ok, upload_result = validate_pdf_upload(ereceipt_file, 'DEO e-receipt file')
            if not ok:
                flash(upload_result, 'danger')
                return render_petition_form()
            original_name = upload_result

        allowed_submission, retry_after, blocked_scopes = _consume_petition_submission_slot()
        if not allowed_submission:
            log_security_event(
                'petition.rate_limit_blocked',
                severity='warning',
                retry_after_seconds=retry_after,
                rate_limit_scope=','.join(blocked_scopes) if blocked_scopes else 'unknown',
            )
            flash(f'Too many petition submissions. Please wait {retry_after} seconds before trying again.', 'danger')
            return render_petition_form(), 429
        if blocked_scopes:
            log_security_event(
                'petition.rate_limit_triggered',
                severity='warning',
                rate_limit_scope=','.join(blocked_scopes),
            )

        if ereceipt_file and ereceipt_file.filename:
            os.makedirs(ERECEIPT_UPLOAD_DIR, exist_ok=True)
            ereceipt_filename = _build_storage_filename('deo_ereceipt', original_name)
            if not ereceipt_filename:
                flash('Unable to prepare e-receipt filename.', 'danger')
                return redirect(url_for('petition_new'))
            saved_ok, save_result = _save_uploaded_file(ereceipt_file, ERECEIPT_UPLOAD_DIR, ereceipt_filename, 'E-receipt file')
            if not saved_ok:
                flash(save_result, 'danger')
                return redirect(url_for('petition_new'))
            ereceipt_filename = save_result

        data = {
            'efile_no': None,
            'petitioner_name': petitioner_name or 'Anonymous',
            'contact': contact,
            'place': place,
            'subject': subject,
            'petition_type': petition_type,
            'source_of_petition': source_of_petition,
            'govt_institution_type': govt_institution_type if source_of_petition == 'govt' else None,
            'received_at': received_at,
            'target_cvo': None if is_jmd_received else target_cvo,
            'permission_request_type': 'permission_required' if is_jmd_received else permission_request_type,
            'received_date': received_date or date.today(),
            'remarks': remarks,
            'ereceipt_no': ereceipt_no,
            'ereceipt_file': ereceipt_filename
        }
        
        try:
            if session.get('user_role') == 'data_entry' and not is_jmd_received:
                # DEO no longer decides enquiry mode; CVO decides at forwarded_to_cvo stage.
                data['requires_permission'] = True
                data['permission_status'] = 'pending'
            elif data['permission_request_type'] == 'direct_enquiry':
                data['requires_permission'] = False
                data['permission_status'] = 'not_required'
            else:
                data['requires_permission'] = True
                data['permission_status'] = 'pending'

            result = models.create_petition(data, session['user_id'])
            if is_jmd_received:
                models.send_for_permission(
                    result['id'],
                    session['user_id'],
                    comments='Auto-routed to PO from JMD Office receipt'
                )
                flash(f'Petition {result["sno"]} created and routed to PO successfully!', 'success')
            else:
                models.forward_petition_to_cvo(
                    result['id'],
                    session['user_id'],
                    data['target_cvo'],
                    comments='Auto-forwarded to concerned CVO/DSP from Data Entry'
                )
                flash(f'Petition {result["sno"]} created and auto-forwarded to CVO/DSP successfully!', 'success')
            return redirect(url_for('petition_view', petition_id=result['id']))
        except pg_errors.UniqueViolation:
            app.logger.warning('Duplicate petition SNO on create', exc_info=True)
            flash('A petition with this serial number already exists. Please try again.', 'danger')
        except psycopg2.DatabaseError:
            app.logger.exception('Database error creating petition')
            flash('A database error occurred while creating the petition. Please contact the administrator.', 'danger')
        except Exception:
            app.logger.exception('Error creating petition')
            flash('Unable to create petition. Please contact administrator.', 'danger')

    return render_petition_form()


@app.route('/petitions/import')
@login_required
@role_required('po', 'super_admin')
def petitions_import():
    return render_template(
        'petitions_import.html',
        template_headers=IMPORT_PETITION_HEADERS,
    )


@app.route('/petitions/import/template')
@login_required
@role_required('po', 'super_admin')
def petitions_import_template():
    sample = {
        'received_date': date.today().strftime('%Y-%m-%d'),
        'received_at': 'jmd_office',
        'target_cvo': 'headquarters',
        'petitioner_name': 'Sample Petitioner',
        'contact': '9876543210',
        'place': 'Vijayawada',
        'subject': 'Historical petition import sample row',
        'petition_type': 'other',
        'source_of_petition': 'public_individual',
        'govt_institution_type': '',
        'enquiry_type': 'detailed',
        'permission_request_type': 'direct_enquiry',
        'requires_permission': 'no',
        'permission_status': 'not_required',
        'status': 'received',
        'efile_no': '',
        'ereceipt_no': '',
        'remarks': 'Optional remarks',
        'assigned_inspector_username': '',
    }
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=IMPORT_PETITION_HEADERS)
    writer.writeheader()
    writer.writerow(sample)
    csv_content = out.getvalue()
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=petition_bulk_import_template.csv'}
    )


@app.route('/petitions/import/upload', methods=['POST'])
@login_required
@role_required('po', 'super_admin')
def petitions_import_upload():
    _return_to = (request.form.get('_return_to') or '').strip()
    _import_back = url_for('help_page') if _return_to == 'help' else url_for('petitions_import')
    upload = request.files.get('petitions_file')
    if not upload or not upload.filename:
        flash('Please choose an Excel/CSV file to upload.', 'warning')
        return redirect(_import_back)

    try:
        rows = _parse_tabular_upload_rows(
            upload,
            required_headers={'subject'},
            allowed_headers=set(IMPORT_PETITION_HEADERS),
        )
    except Exception:
        app.logger.exception('Unable to parse petition import upload file')
        flash('Unable to parse upload file. Please verify format and retry.', 'danger')
        return redirect(_import_back)

    if not rows:
        flash('Uploaded file is empty.', 'warning')
        return redirect(_import_back)

    active_users = [u for u in models.get_all_users() if u.get('is_active')]
    user_by_username = {(u.get('username') or '').strip().lower(): u for u in active_users if u.get('username')}
    first_role_user = {}
    for u in active_users:
        role = (u.get('role') or '').strip()
        if role and role not in first_role_user:
            first_role_user[role] = u

    def role_user_id(role_name):
        user = first_role_user.get(role_name)
        return int(user['id']) if user and user.get('id') else None

    target_for_received = {
        'jmd_office': 'headquarters',
        'cvo_apspdcl_tirupathi': 'apspdcl',
        'cvo_apepdcl_vizag': 'apepdcl',
        'cvo_apcpdcl_vijayawada': 'apcpdcl',
    }
    received_for_target = {
        'headquarters': 'jmd_office',
        'apspdcl': 'cvo_apspdcl_tirupathi',
        'apepdcl': 'cvo_apepdcl_vizag',
        'apcpdcl': 'cvo_apcpdcl_vijayawada',
    }
    cvo_role_for_target = {
        'apspdcl': 'cvo_apspdcl',
        'apcpdcl': 'cvo_apspdcl',
        'apepdcl': 'cvo_apepdcl',
        'headquarters': 'dsp',
    }
    cmd_role_for_target = {
        'apspdcl': 'cmd_apspdcl',
        'apcpdcl': 'cmd_apcpdcl',
        'apepdcl': 'cmd_apepdcl',
        'headquarters': 'cgm_hr_transco',
    }

    created = 0
    failed = 0
    warnings = []
    errors = []
    actor_user_id = session['user_id']
    po_handler_id = actor_user_id

    for idx, row in enumerate(rows, start=2):
        try:
            received_at = _normalize_received_at(row.get('received_at'))
            target_cvo = _normalize_target_cvo(row.get('target_cvo'))
            if not received_at and target_cvo:
                received_at = received_for_target.get(target_cvo)
            if not target_cvo and received_at:
                target_cvo = target_for_received.get(received_at)
            if not received_at:
                received_at = 'jmd_office'
            if not target_cvo:
                target_cvo = target_for_received.get(received_at, 'headquarters')

            received_date = parse_flexible_date(row.get('received_date')) or date.today()
            petitioner_name = (row.get('petitioner_name') or '').strip()
            contact = (row.get('contact') or '').strip()
            place = (row.get('place') or '').strip()
            subject = (row.get('subject') or '').strip()
            remarks = (row.get('remarks') or '').strip()
            if not subject:
                subject = (remarks[:240] if remarks else '').strip() or 'Imported historical petition'

            petition_type = _normalize_petition_type(row.get('petition_type'))
            source_of_petition = _normalize_source(row.get('source_of_petition'))
            govt_institution_type = (row.get('govt_institution_type') or '').strip().lower() or None
            if source_of_petition != 'govt':
                govt_institution_type = None
            elif govt_institution_type not in VALID_GOVT_INSTITUTIONS:
                warnings.append(f'Row {idx}: invalid govt_institution_type mapped to blank.')
                govt_institution_type = None

            enquiry_type_raw = (row.get('enquiry_type') or '').strip().lower()
            enquiry_type = enquiry_type_raw if enquiry_type_raw in VALID_ENQUIRY_TYPES else 'detailed'

            perm_req_type = (row.get('permission_request_type') or '').strip().lower()
            if perm_req_type in ('direct', 'direct_enquiry', 'not_required'):
                requires_permission = False
            elif perm_req_type in ('permission', 'permission_required', 'required'):
                requires_permission = True
            else:
                requires_permission = _to_bool(row.get('requires_permission'), default=(source_of_petition != 'media'))

            permission_status_raw = (row.get('permission_status') or '').strip().lower()
            if permission_status_raw not in {'pending', 'approved', 'rejected', 'not_required'}:
                permission_status_raw = ''
            permission_status = permission_status_raw
            if not permission_status:
                permission_status = 'pending' if requires_permission else 'not_required'

            status_raw = (row.get('status') or '').strip().lower()
            status_alias = {
                'open': 'received',
                'in_progress': 'enquiry_in_progress',
                'beyond_sla': 'enquiry_in_progress',
                'within_sla': 'enquiry_in_progress',
            }
            status = status_alias.get(status_raw, status_raw or 'received')
            if status not in IMPORT_ALLOWED_STATUSES:
                warnings.append(f'Row {idx}: invalid status "{status_raw}" mapped to "received".')
                status = 'received'

            assigned_inspector_id = None
            inspector_username = (row.get('assigned_inspector_username') or '').strip().lower()
            if inspector_username:
                inspector = user_by_username.get(inspector_username)
                if inspector and inspector.get('role') == 'inspector':
                    assigned_inspector_id = int(inspector['id'])
                else:
                    warnings.append(f'Row {idx}: assigned_inspector_username "{inspector_username}" not found/invalid.')

            data = {
                'efile_no': (row.get('efile_no') or '').strip() or None,
                'petitioner_name': petitioner_name or 'Anonymous',
                'contact': contact or None,
                'place': place or None,
                'subject': subject,
                'petition_type': petition_type,
                'source_of_petition': source_of_petition,
                'received_at': received_at,
                'target_cvo': target_cvo,
                'requires_permission': requires_permission,
                'received_date': received_date,
                'govt_institution_type': govt_institution_type,
                'permission_status': permission_status,
                'enquiry_type': enquiry_type,
                'remarks': remarks or None,
                'ereceipt_no': (row.get('ereceipt_no') or '').strip() or None,
                'ereceipt_file': None,
            }
            created_petition = models.create_petition(data, actor_user_id)
            petition_id = int(created_petition['id'])

            cvo_handler_id = role_user_id(cvo_role_for_target.get(target_cvo, ''))
            cmd_handler_id = role_user_id(cmd_role_for_target.get(target_cvo, ''))
            if status in {'forwarded_to_cvo', 'permission_approved', 'sent_back_for_reenquiry'}:
                current_handler_id = cvo_handler_id or po_handler_id
            elif status in {'sent_for_permission', 'permission_rejected', 'forwarded_to_po', 'forwarded_to_jmd', 'action_taken', 'lodged', 'closed'}:
                current_handler_id = po_handler_id
            elif status in {'action_instructed'}:
                current_handler_id = cmd_handler_id or po_handler_id
            elif status in {'assigned_to_inspector', 'enquiry_in_progress', 'enquiry_report_submitted'}:
                current_handler_id = assigned_inspector_id or cvo_handler_id or po_handler_id
            else:
                current_handler_id = po_handler_id

            models.update_imported_petition_state(
                petition_id,
                actor_user_id,
                status=status,
                current_handler_id=current_handler_id,
                assigned_inspector_id=assigned_inspector_id,
                target_cvo=target_cvo,
                requires_permission=requires_permission,
                permission_status=permission_status,
                enquiry_type=enquiry_type,
                received_date=received_date,
                remarks=(remarks or None),
            )
            created += 1
        except pg_errors.UniqueViolation:
            failed += 1
            app.logger.warning('Duplicate SNO on import row %s', idx, exc_info=True)
            errors.append(f'Row {idx}: duplicate serial number — already exists in database.')
        except psycopg2.DatabaseError:
            failed += 1
            app.logger.exception('Database error on import row %s', idx)
            errors.append(f'Row {idx}: database error during import.')
        except Exception:
            failed += 1
            app.logger.exception('Petition import row failed at row %s', idx)
            errors.append(f'Row {idx}: internal processing error.')

    if created:
        flash(f'Petition import complete. Imported: {created}, Failed: {failed}.', 'success')
    if warnings:
        preview = '; '.join(warnings[:4]) + ('; ...' if len(warnings) > 4 else '')
        flash(f'Import warnings: {preview}', 'warning')
    if errors:
        preview = '; '.join(errors[:5]) + ('; ...' if len(errors) > 5 else '')
        flash(f'Import errors: {preview}', 'danger')
    return redirect(_import_back)


@app.route('/petitions/<int:petition_id>')
@login_required
def petition_view(petition_id):
    if not _can_access_petition(petition_id):
        log_security_event('access.petition_forbidden', severity='warning', petition_id=petition_id)
        flash('You do not have access to this petition.', 'danger')
        return redirect(url_for('petitions_list'))

    petition = models.get_petition_by_id(petition_id)
    if not petition:
        flash('Petition not found.', 'danger')
        return redirect(url_for('petitions_list'))
    
    tracking = models.get_petition_tracking(petition_id)
    report = models.get_enquiry_report(petition_id)
    ereceipt_file_available = (
        bool(petition.get('ereceipt_file'))
        and _uploaded_file_exists(ERECEIPT_UPLOAD_DIR, petition.get('ereceipt_file'))
    )
    conclusion_file_available = (
        bool(petition.get('conclusion_file'))
        and _uploaded_file_exists(ENQUIRY_UPLOAD_DIR, petition.get('conclusion_file'))
    )
    report_file_availability = {
        'report_file': False,
        'cvo_consolidated_report_file': False,
        'cmd_action_report_file': False,
    }
    if report:
        report_file_availability['report_file'] = bool(report.get('report_file')) and _uploaded_file_exists(
            ENQUIRY_UPLOAD_DIR, report.get('report_file')
        )
        report_file_availability['cvo_consolidated_report_file'] = bool(report.get('cvo_consolidated_report_file')) and _uploaded_file_exists(
            ENQUIRY_UPLOAD_DIR, report.get('cvo_consolidated_report_file')
        )
        report_file_availability['cmd_action_report_file'] = bool(report.get('cmd_action_report_file')) and _uploaded_file_exists(
            ENQUIRY_UPLOAD_DIR, report.get('cmd_action_report_file')
        )
    tracking_file_availability = {}
    for row in tracking or []:
        attachment_name = row.get('attachment_file')
        if attachment_name and attachment_name not in tracking_file_availability:
            tracking_file_availability[attachment_name] = _uploaded_file_exists(ENQUIRY_UPLOAD_DIR, attachment_name)
    ci_assignment_memo = None
    for row in reversed(tracking or []):
        if (row.get('action') or '').strip() == 'Assigned to Inspector' and row.get('attachment_file'):
            ci_assignment_memo = {
                'filename': row.get('attachment_file'),
                'from_name': row.get('from_name'),
                'created_at': row.get('created_at'),
                'comments': row.get('comments'),
                'is_available': tracking_file_availability.get(row.get('attachment_file'), False),
            }
            break
    inspector_conversion_request_pending = (
        petition.get('status') == 'enquiry_report_submitted'
        and (petition.get('enquiry_type') or '').strip().lower() == 'preliminary'
        and _has_pending_inspector_detailed_request(tracking)
    )
    conversion_permission_stage = _is_conversion_permission_stage(petition, tracking)
    conversion_reassignment_locked = (
        petition.get('status') == 'permission_approved'
        and _has_conversion_request_history(tracking)
        and bool(petition.get('assigned_inspector_id'))
    )
    petition_sla_eval = None
    try:
        petition_sla_rows = models.get_sla_evaluation_rows([petition])
        if petition_sla_rows:
            petition_sla_eval = petition_sla_rows[0]
    except Exception:
        petition_sla_eval = None
    po_beyond_sla_permission_allowed = (
        session['user_role'] in ('po', 'super_admin')
        and bool(petition_sla_eval and petition_sla_eval.get('is_beyond_sla_for_po'))
        and not petition_sla_eval.get('closed_at')
        and petition.get('status') not in ('permission_approved', 'lodged', 'closed')
    )
    
    # Get inspectors mapped to the relevant CVO/DSP officer
    inspectors = []
    cvo_users = []
    cmd_cgm_users = []
    cvo_like_roles = ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp')
    if session['user_role'] in cvo_like_roles:
        inspectors = models.get_inspectors_by_cvo(session['user_id'])
    elif session['user_role'] == 'super_admin':
        handler_id = petition.get('current_handler_id')
        if handler_id:
            handler_user = _get_user_by_id_cached(handler_id)
            if handler_user and handler_user.get('role') in cvo_like_roles:
                inspectors = models.get_inspectors_by_cvo(handler_id)
    if session['user_role'] in ('po', 'super_admin'):
        cvo_users = models.get_cvo_users()
        cmd_cgm_users = models.get_cmd_cgm_users()
    
    return render_template('petition_view.html', 
                         petition=petition, tracking=tracking, report=report,
                         inspectors=inspectors, cvo_users=cvo_users, cmd_cgm_users=cmd_cgm_users,
                         inspector_conversion_request_pending=inspector_conversion_request_pending,
                         conversion_permission_stage=conversion_permission_stage,
                         conversion_reassignment_locked=conversion_reassignment_locked,
                         petition_sla_eval=petition_sla_eval,
                         po_beyond_sla_permission_allowed=po_beyond_sla_permission_allowed,
                         ci_assignment_memo=ci_assignment_memo,
                         tracking_file_availability=tracking_file_availability,
                         ereceipt_file_available=ereceipt_file_available,
                         conclusion_file_available=conclusion_file_available,
                         report_file_availability=report_file_availability)

# ========================================
# WORKFLOW ACTION ROUTES
# ========================================

@app.route('/petitions/<int:petition_id>/action', methods=['POST'])
@login_required
def petition_action(petition_id):
    if not _can_access_petition(petition_id):
        log_security_event('access.petition_action_forbidden', severity='warning', petition_id=petition_id)
        flash('You do not have access to this petition.', 'danger')
        return redirect(url_for('petitions_list'))

    action = (request.form.get('action') or '').strip()
    comments = request.form.get('comments', '').strip()
    user_id = session['user_id']
    user_role = session['user_role']
    try:
        petition_for_guard = models.get_petition_by_id(petition_id)
    except Exception:
        petition_for_guard = None
    pending_inspector_conversion_request = False
    if petition_for_guard and petition_for_guard.get('status') == 'enquiry_report_submitted' and (petition_for_guard.get('enquiry_type') or '').strip().lower() == 'preliminary':
        try:
            pending_inspector_conversion_request = _has_pending_inspector_detailed_request(models.get_petition_tracking(petition_id))
        except Exception:
            pending_inspector_conversion_request = False
    
    try:
        form_cfg = get_effective_form_field_configs()
        cfg_po_approve_efile = form_cfg.get('po_decision.approve_permission_efile_no', DEFAULT_FORM_FIELD_CONFIGS['po_decision.approve_permission_efile_no'])
        cfg_po_reject_reason = form_cfg.get('po_decision.reject_permission_reason', DEFAULT_FORM_FIELD_CONFIGS['po_decision.reject_permission_reason'])
        cfg_po_send_cmd_instructions = form_cfg.get('po_decision.send_cmd_instructions', DEFAULT_FORM_FIELD_CONFIGS['po_decision.send_cmd_instructions'])
        cfg_po_send_cmd_efile = form_cfg.get('po_decision.send_cmd_efile_no', DEFAULT_FORM_FIELD_CONFIGS['po_decision.send_cmd_efile_no'])
        cfg_po_lodge_efile = form_cfg.get('po_decision.po_lodge_efile_no', DEFAULT_FORM_FIELD_CONFIGS['po_decision.po_lodge_efile_no'])
        cfg_po_lodge_remarks = form_cfg.get('po_decision.po_lodge_remarks', DEFAULT_FORM_FIELD_CONFIGS['po_decision.po_lodge_remarks'])
        cfg_po_direct_lodge_efile = form_cfg.get('po_decision.po_direct_lodge_efile_no', DEFAULT_FORM_FIELD_CONFIGS['po_decision.po_direct_lodge_efile_no'])
        cfg_po_direct_lodge_remarks = form_cfg.get('po_decision.po_direct_lodge_remarks', DEFAULT_FORM_FIELD_CONFIGS['po_decision.po_direct_lodge_remarks'])
        cfg_po_close_comments = form_cfg.get('po_decision.close_comments', DEFAULT_FORM_FIELD_CONFIGS['po_decision.close_comments'])

        if not action:
            flash('Invalid action request.', 'warning')
            return redirect(url_for('petition_view', petition_id=petition_id))

        if pending_inspector_conversion_request and user_role in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
            if action in {'cvo_comments', 'cvo_send_back_reenquiry', 'cvo_direct_lodge'}:
                flash('Inspector requested detailed conversion. Please use "Request Detailed Enquiry" action only.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

        if action == 'forward_to_cvo':
            if user_role not in ('super_admin', 'data_entry'):
                flash('You are not allowed to forward petitions to CVO/DSP.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            target_cvo = (request.form.get('target_cvo') or '').strip()
            if target_cvo not in VALID_TARGET_CVO:
                flash('Please select a valid target CVO/DSP.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.forward_petition_to_cvo(petition_id, user_id, target_cvo, comments)
            flash('Petition forwarded to CVO/DSP successfully.', 'success')
            
        elif action == 'send_for_permission':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can send petitions for permission routing.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.send_for_permission(petition_id, user_id, comments)
            flash('Petition sent to PO for permission.', 'success')

        elif action in ('cvo_set_enquiry_mode', 'send_receipt_to_po', 'cvo_route_petition'):
            if user_role not in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Only CVO/DSP can decide enquiry mode.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            if petition.get('status') != 'forwarded_to_cvo':
                flash('Enquiry mode can be decided only when petition is with CVO/DSP.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            permission_request_type = (request.form.get('permission_request_type') or '').strip()
            if action == 'send_receipt_to_po':
                # Backward-compatible path for older form payload.
                permission_request_type = 'permission_required'
            if permission_request_type not in VALID_PERMISSION_REQUEST_TYPES:
                flash('Please select enquiry permission (Direct/Permission Based).', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            if permission_request_type == 'permission_required':
                permission_file = request.files.get('permission_file')
                permission_filename = None
                if permission_file and permission_file.filename:
                    ok, upload_result = validate_pdf_upload(permission_file, 'Permission document')
                    if not ok:
                        flash(upload_result, 'danger')
                        return redirect(url_for('petition_view', petition_id=petition_id))
                    original_name = upload_result
                    os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
                    permission_filename = _build_storage_filename('cvo_permission', original_name, petition_id)
                    if not permission_filename:
                        flash('Unable to prepare permission document filename.', 'danger')
                        return redirect(url_for('petition_view', petition_id=petition_id))
                    saved_ok, save_result = _save_uploaded_file(permission_file, ENQUIRY_UPLOAD_DIR, permission_filename, 'Permission document')
                    if not saved_ok:
                        flash(save_result, 'danger')
                        return redirect(url_for('petition_view', petition_id=petition_id))
                    permission_filename = save_result
                try:
                    models.cvo_send_receipt_to_po(petition_id, user_id, comments, permission_filename)
                except Exception:
                    if permission_filename:
                        _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, permission_filename)
                    raise
                flash('Permission route selected and receipt sent to PO.', 'success')
            else:
                enquiry_type_decision = (request.form.get('enquiry_type_decision') or '').strip() or 'detailed'
                if enquiry_type_decision not in VALID_ENQUIRY_TYPES:
                    enquiry_type_decision = 'detailed'
                if action == 'cvo_route_petition':
                    inspector_id = parse_optional_int(request.form.get('inspector_id'))
                    if not inspector_id:
                        flash('Please select a valid field inspector.', 'warning')
                        return redirect(url_for('petition_view', petition_id=petition_id))
                    memo_file = request.files.get('assignment_memo_file')
                    memo_filename = None
                    if memo_file and memo_file.filename:
                        ok, upload_result = validate_pdf_upload(memo_file, 'Upload memo/instructions')
                        if not ok:
                            flash(upload_result, 'danger')
                            return redirect(url_for('petition_view', petition_id=petition_id))
                        original_name = upload_result
                        os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
                        memo_filename = _build_storage_filename('assign_memo', original_name, petition_id)
                        if not memo_filename:
                            flash('Unable to prepare memo filename.', 'danger')
                            return redirect(url_for('petition_view', petition_id=petition_id))
                        saved_ok, save_result = _save_uploaded_file(memo_file, ENQUIRY_UPLOAD_DIR, memo_filename, 'Memo/instructions file')
                        if not saved_ok:
                            flash(save_result, 'danger')
                            return redirect(url_for('petition_view', petition_id=petition_id))
                        memo_filename = save_result
                    try:
                        models.cvo_mark_direct_enquiry(petition_id, user_id, comments, enquiry_type_decision)
                        models.assign_to_inspector(
                            petition_id, user_id, inspector_id, comments, enquiry_type_decision, memo_filename
                        )
                    except Exception:
                        if memo_filename:
                            _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, memo_filename)
                        raise
                    flash('Direct enquiry selected and petition assigned to inspector.', 'success')
                else:
                    models.cvo_mark_direct_enquiry(petition_id, user_id, comments, enquiry_type_decision)
                    flash('Direct enquiry mode selected. You can now assign inspector.', 'success')

        elif action == 'cvo_direct_lodge':
            if user_role not in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Only CVO/DSP can directly lodge this petition.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            if petition.get('source_of_petition') != 'media':
                flash('Direct lodge at CVO is allowed only for Electronic and Print Media source petitions.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if petition.get('status') != 'enquiry_report_submitted':
                flash('Direct lodge at CVO is allowed only after enquiry report is submitted by field officer.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            lodge_remarks = (request.form.get('lodge_remarks') or '').strip()
            if not lodge_remarks:
                flash('Remarks are required for direct lodge by CVO.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(lodge_remarks) > 5000:
                flash('Lodge remarks are too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.cvo_direct_lodge_petition(petition_id, user_id, lodge_remarks)
            flash('Petition directly lodged by CVO for media source.', 'success')
            
        elif action == 'approve_permission':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can approve permission.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            target_cvo = (request.form.get('target_cvo') or '').strip()
            organization = (request.form.get('organization') or '').strip().lower()
            enquiry_type_decision = (request.form.get('enquiry_type_decision') or '').strip()
            efile_no_input = request.form.get('efile_no', '').strip()
            conversion_permission_stage = False
            try:
                conversion_permission_stage = _is_conversion_permission_stage(
                    petition,
                    models.get_petition_tracking(petition_id)
                )
            except Exception:
                conversion_permission_stage = False
            if target_cvo not in VALID_TARGET_CVO:
                flash('Please select a valid target CVO/DSP.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if petition.get('received_at') == 'jmd_office' and organization not in VALID_ORGANIZATIONS:
                flash('Please select organization (APTRANSCO/APGENCO).', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if conversion_permission_stage:
                enquiry_type_decision = 'detailed'
            elif enquiry_type_decision not in VALID_ENQUIRY_TYPES:
                flash('Please select enquiry type decision (Detailed/Preliminary).', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            efile_no, efile_error = resolve_efile_no_for_action(
                petition,
                efile_no_input,
                required_message=f"{cfg_po_approve_efile.get('label', 'E-Office File No')} is required to approve permission." if cfg_po_approve_efile.get('required') else None
            )
            if efile_error:
                flash(efile_error, 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.approve_permission(
                petition_id,
                user_id,
                target_cvo,
                efile_no,
                comments,
                enquiry_type_decision,
                organization=organization if organization in VALID_ORGANIZATIONS else None,
            )
            flash('Permission granted and pushed to respective CVO/DSP.', 'success')

        elif action == 'po_beyond_sla_send_to_cvo':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can process beyond-SLA petitions.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            sla_rows = models.get_sla_evaluation_rows([petition])
            sla_row = sla_rows[0] if sla_rows else None
            if not sla_row or not sla_row.get('is_beyond_sla_for_po') or sla_row.get('closed_at'):
                flash('Beyond-SLA routing is available only after the petition crosses 90 days.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if petition.get('status') in ('permission_approved', 'lodged', 'closed'):
                flash('This petition is no longer eligible for beyond-SLA PO routing.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            target_cvo = (request.form.get('target_cvo') or petition.get('target_cvo') or '').strip()
            if target_cvo not in VALID_TARGET_CVO:
                flash('Please select a valid target CVO/DSP.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            organization = (request.form.get('organization') or '').strip().lower()
            if petition.get('received_at') == 'jmd_office' and organization not in VALID_ORGANIZATIONS:
                flash('Please select organization (APTRANSCO/APGENCO).', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            permission_copy = request.files.get('permission_file')
            if not permission_copy or not permission_copy.filename:
                flash('Permission copy PDF is required for beyond-SLA PO routing.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            ok, upload_result = validate_pdf_upload(permission_copy, 'Permission copy')
            if not ok:
                flash(upload_result, 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            original_name = upload_result
            os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
            permission_filename = _build_storage_filename('po_beyond_sla_permission', original_name, petition_id)
            if not permission_filename:
                flash('Unable to prepare permission copy filename.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            saved_ok, save_result = _save_uploaded_file(permission_copy, ENQUIRY_UPLOAD_DIR, permission_filename, 'Permission copy')
            if not saved_ok:
                flash(save_result, 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            permission_filename = save_result
            efile_no_input = request.form.get('efile_no', '').strip()
            efile_no, efile_error = resolve_efile_no_for_action(
                petition,
                efile_no_input,
                required_message=f"{cfg_po_approve_efile.get('label', 'E-Office File No')} is required for beyond-SLA PO routing." if cfg_po_approve_efile.get('required') else None
            )
            if efile_error:
                _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, permission_filename)
                flash(efile_error, 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            try:
                models.approve_permission(
                    petition_id,
                    user_id,
                    target_cvo,
                    efile_no,
                    comments,
                    petition.get('enquiry_type'),
                    organization=organization if organization in VALID_ORGANIZATIONS else None,
                    attachment_file=permission_filename,
                    tracking_action='Beyond SLA Permission Copy Uploaded - Sent to CVO',
                    mark_overdue_escalated=True,
                )
            except Exception:
                if permission_filename:
                    _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, permission_filename)
                raise
            flash('Beyond-SLA petition routed to concerned CVO with permission copy.', 'success')

        elif action == 'reject_permission':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can reject permission.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if cfg_po_reject_reason.get('required') and not comments:
                flash(f"{cfg_po_reject_reason.get('label', 'Reason for rejection')} is required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.reject_permission(petition_id, user_id, comments)
            flash('Permission rejected.', 'warning')
            
        elif action == 'assign_inspector':
            if user_role not in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Only CVO/DSP can assign inspectors.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if petition and petition.get('requires_permission') and petition.get('status') != 'permission_approved':
                flash('Permission is compulsory. PO approval required before assigning inspector.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if petition and not petition.get('requires_permission') and petition.get('status') != 'forwarded_to_cvo':
                flash('For Direct Enquiry, inspector can be assigned only when petition is at CVO/DSP.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            inspector_id = parse_optional_int(request.form.get('inspector_id'))
            if not inspector_id:
                flash('Please select a valid field inspector.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            conversion_reassignment_locked = False
            try:
                conversion_reassignment_locked = (
                    petition is not None
                    and petition.get('status') == 'permission_approved'
                    and _has_conversion_request_history(models.get_petition_tracking(petition_id))
                    and bool(petition.get('assigned_inspector_id'))
                )
            except Exception:
                conversion_reassignment_locked = False
            if conversion_reassignment_locked:
                locked_inspector_id = int(petition.get('assigned_inspector_id') or 0)
                if not locked_inspector_id:
                    flash('Conversion workflow requires previously assigned inspector.', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if int(inspector_id) != locked_inspector_id:
                    flash('For preliminary-to-detailed conversion, only previously assigned inspector is allowed.', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
            enquiry_type_decision = (request.form.get('enquiry_type_decision') or '').strip().lower() or None
            if petition and not petition.get('requires_permission'):
                if enquiry_type_decision not in VALID_ENQUIRY_TYPES:
                    flash('Please select enquiry type decision (Detailed/Preliminary).', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
            else:
                enquiry_type_decision = None
            memo_file = request.files.get('assignment_memo_file')
            memo_filename = None
            if memo_file and memo_file.filename:
                ok, upload_result = validate_pdf_upload(memo_file, 'Upload memo/instructions')
                if not ok:
                    flash(upload_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                original_name = upload_result
                os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
                memo_filename = _build_storage_filename('assign_memo', original_name, petition_id)
                if not memo_filename:
                    flash('Unable to prepare memo filename.', 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                saved_ok, save_result = _save_uploaded_file(memo_file, ENQUIRY_UPLOAD_DIR, memo_filename, 'Memo/instructions file')
                if not saved_ok:
                    flash(save_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                memo_filename = save_result
            try:
                models.assign_to_inspector(
                    petition_id, user_id, inspector_id, comments, enquiry_type_decision, memo_filename
                )
            except Exception:
                if memo_filename:
                    _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, memo_filename)
                raise
            flash('Petition assigned to inspector.', 'success')

        elif action == 'submit_report':
            if user_role not in ('super_admin', 'inspector'):
                flash('Only inspectors can upload enquiry report.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            form_cfg = get_effective_form_field_configs()
            cfg_report_text = form_cfg.get('inspector_report.report_text', DEFAULT_FORM_FIELD_CONFIGS['inspector_report.report_text'])
            cfg_recommendation = form_cfg.get('inspector_report.recommendation', DEFAULT_FORM_FIELD_CONFIGS['inspector_report.recommendation'])
            cfg_report_file = form_cfg.get('inspector_report.report_file', DEFAULT_FORM_FIELD_CONFIGS['inspector_report.report_file'])
            cfg_request_detailed = form_cfg.get(
                'inspector_report.request_detailed_permission',
                DEFAULT_FORM_FIELD_CONFIGS['inspector_report.request_detailed_permission']
            )
            cfg_detailed_reason = form_cfg.get(
                'inspector_report.detailed_request_reason',
                DEFAULT_FORM_FIELD_CONFIGS['inspector_report.detailed_request_reason']
            )
            report_text = request.form.get('report_text', '').strip()
            recommendation = request.form.get('recommendation', '').strip()
            report_next_step = (request.form.get('report_next_step') or '').strip().lower()
            if report_next_step not in ('send_to_cvo', 'ask_detailed_permission'):
                # Backward compatible path for older checkbox payloads.
                report_next_step = 'ask_detailed_permission' if (request.form.get('request_detailed_permission') or '').strip() == '1' else 'send_to_cvo'
            request_detailed_permission = report_next_step == 'ask_detailed_permission'
            detailed_request_reason = (request.form.get('detailed_request_reason') or '').strip()
            accident_type = None
            deceased_category = None
            departmental_type = None
            non_departmental_type = None
            deceased_count = None
            general_public_count = None
            animals_count = None
            if not request_detailed_permission and (petition.get('petition_type') or '').strip() == 'electrical_accident':
                accident_type = (request.form.get('accident_type') or '').strip()
                deceased_category = (request.form.get('deceased_category') or '').strip()
                departmental_type = (request.form.get('departmental_type') or '').strip()
                non_departmental_type = (request.form.get('non_departmental_type') or '').strip()
                deceased_count_raw = (request.form.get('deceased_count') or '').strip()
                general_public_raw = (request.form.get('general_public_count') or '').strip()
                animals_raw = (request.form.get('animals_count') or '').strip()
                if accident_type not in ('fatal', 'non_fatal'):
                    flash('Please select type of accident (Fatal / Non Fatal).', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if deceased_category not in ('departmental', 'non_departmental', 'general_public', 'animals'):
                    flash('Please select deceased person/category.', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                try:
                    deceased_count = int(deceased_count_raw)
                except (TypeError, ValueError):
                    deceased_count = 0
                if deceased_count <= 0:
                    flash('Please enter valid No. of Deceased (greater than 0).', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if deceased_category == 'departmental':
                    if departmental_type not in ('regular', 'outsourced'):
                        flash('Please select departmental type (Regular / Outsourced).', 'warning')
                        return redirect(url_for('petition_view', petition_id=petition_id))
                    non_departmental_type = None
                if deceased_category == 'non_departmental':
                    if non_departmental_type not in ('private_electricians', 'contract_labour'):
                        flash('Please select non-departmental type (Private Electricians / Contract Labour).', 'warning')
                        return redirect(url_for('petition_view', petition_id=petition_id))
                    departmental_type = None
                else:
                    non_departmental_type = None
                if deceased_category != 'departmental':
                    departmental_type = None
                if deceased_category == 'general_public':
                    try:
                        general_public_count = int(general_public_raw or deceased_count)
                    except (TypeError, ValueError):
                        general_public_count = 0
                    if general_public_count <= 0:
                        flash('Please enter valid No. of General Public (greater than 0).', 'warning')
                        return redirect(url_for('petition_view', petition_id=petition_id))
                if deceased_category == 'animals':
                    try:
                        animals_count = int(animals_raw or deceased_count)
                    except (TypeError, ValueError):
                        animals_count = 0
                    if animals_count <= 0:
                        flash('Please enter valid No. of Animals (greater than 0).', 'warning')
                        return redirect(url_for('petition_view', petition_id=petition_id))
            if request_detailed_permission:
                if (petition.get('enquiry_type') or '').strip().lower() != 'preliminary':
                    flash('Detailed enquiry conversion request is allowed only for preliminary enquiry petitions.', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if cfg_detailed_reason.get('required') and not detailed_request_reason:
                    flash(f"{cfg_detailed_reason.get('label', 'Reason for Detailed Enquiry Request')} is required.", 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if len(detailed_request_reason) > 2000:
                    flash(f"{cfg_detailed_reason.get('label', 'Reason for Detailed Enquiry Request')} is too long.", 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
            else:
                if cfg_report_text.get('required') and not report_text:
                    flash(f"{cfg_report_text.get('label', 'Conclusion of enquiry report')} is required.", 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if len(report_text) > 20000:
                    flash('Conclusion of enquiry report is too long.', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if cfg_recommendation.get('required') and not recommendation:
                    flash(f"{cfg_recommendation.get('label', 'Recommendations/Suggestions')} are required.", 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                if len(recommendation) > 5000:
                    flash('Recommendations/Suggestions text is too long.', 'warning')
                    return redirect(url_for('petition_view', petition_id=petition_id))
            report_file = request.files.get('report_file')
            if (not request_detailed_permission) and cfg_report_file.get('required') and (not report_file or not report_file.filename):
                flash('Enquiry report file (PDF) is compulsory.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if request_detailed_permission:
                report_file = None
            if not report_file or not report_file.filename:
                report_filename = None
                stored_report_text = report_text
                if request_detailed_permission and not stored_report_text:
                    stored_report_text = 'Inspector requested permission to convert preliminary enquiry to detailed enquiry.'
                models.submit_enquiry_report(
                    petition_id, user_id, stored_report_text, '', recommendation if not request_detailed_permission else '', report_filename,
                    request_detailed_permission=request_detailed_permission,
                    detailed_request_reason=detailed_request_reason,
                    accident_type=accident_type,
                    deceased_category=deceased_category,
                    departmental_type=departmental_type,
                    non_departmental_type=non_departmental_type,
                    deceased_count=deceased_count,
                    general_public_count=general_public_count,
                    animals_count=animals_count
                )
                if request_detailed_permission:
                    flash(
                        f'"{cfg_request_detailed.get("label", "Detailed enquiry conversion request")}" sent to CVO/DSP.',
                        'success'
                    )
                else:
                    flash('Enquiry report uploaded successfully.', 'success')
                return redirect(url_for('petition_view', petition_id=petition_id))
            ok, upload_result = validate_pdf_upload(report_file, 'Enquiry report attachment')
            if not ok:
                flash(upload_result, 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            original_name = upload_result
            os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
            report_filename = _build_storage_filename('enquiry', original_name, petition_id)
            if not report_filename:
                flash('Unable to prepare enquiry report filename.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            saved_ok, save_result = _save_uploaded_file(report_file, ENQUIRY_UPLOAD_DIR, report_filename, 'Enquiry report file')
            if not saved_ok:
                flash(save_result, 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            report_filename = save_result
            stored_report_text = report_text
            if request_detailed_permission and not stored_report_text:
                stored_report_text = 'Inspector requested permission to convert preliminary enquiry to detailed enquiry.'
            try:
                models.submit_enquiry_report(
                    petition_id, user_id, stored_report_text, '', recommendation if not request_detailed_permission else '', report_filename,
                    request_detailed_permission=request_detailed_permission,
                    detailed_request_reason=detailed_request_reason,
                    accident_type=accident_type,
                    deceased_category=deceased_category,
                    departmental_type=departmental_type,
                    non_departmental_type=non_departmental_type,
                    deceased_count=deceased_count,
                    general_public_count=general_public_count,
                    animals_count=animals_count
                )
            except Exception:
                _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, report_filename)
                raise
            if request_detailed_permission:
                flash(
                    f'"{cfg_request_detailed.get("label", "Detailed enquiry conversion request")}" sent to CVO/DSP.',
                    'success'
                )
            else:
                flash('Enquiry report uploaded successfully.', 'success')
            
        elif action == 'cvo_comments':
            if user_role not in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Only CVO/DSP can enter remarks.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            form_cfg = get_effective_form_field_configs()
            cfg_cvo_comments = form_cfg.get('cvo_review.cvo_comments', DEFAULT_FORM_FIELD_CONFIGS['cvo_review.cvo_comments'])
            cfg_cvo_file = form_cfg.get('cvo_review.consolidated_report_file', DEFAULT_FORM_FIELD_CONFIGS['cvo_review.consolidated_report_file'])
            cvo_comments = request.form.get('cvo_comments', '').strip()
            if cfg_cvo_comments.get('required') and not cvo_comments:
                flash(f"{cfg_cvo_comments.get('label', 'CVO/DSP comments')} are required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            consolidated_file = request.files.get('consolidated_report_file')
            if cfg_cvo_file.get('required') and (not consolidated_file or not consolidated_file.filename):
                flash(f"{cfg_cvo_file.get('label', 'Consolidated report file')} is required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            consolidated_filename = None
            if consolidated_file and consolidated_file.filename:
                ok, upload_result = validate_pdf_upload(consolidated_file, 'Consolidated report upload')
                if not ok:
                    flash(upload_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                original_name = upload_result

                os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
                consolidated_filename = _build_storage_filename('cvo_consolidated', original_name, petition_id)
                if not consolidated_filename:
                    flash('Unable to prepare consolidated report filename.', 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                saved_ok, save_result = _save_uploaded_file(consolidated_file, ENQUIRY_UPLOAD_DIR, consolidated_filename, 'Consolidated report file')
                if not saved_ok:
                    flash(save_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                consolidated_filename = save_result
            try:
                if consolidated_filename:
                    models.cvo_upload_consolidated_report(petition_id, user_id, consolidated_filename)
                models.cvo_add_comments(petition_id, user_id, cvo_comments)
            except Exception:
                if consolidated_filename:
                    _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, consolidated_filename)
                raise
            flash('Forwarded to PO for conclusion.', 'success')

        elif action == 'cvo_send_back_reenquiry':
            if user_role not in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Only CVO/DSP can send back for re-enquiry.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            if petition.get('status') != 'enquiry_report_submitted':
                flash('Re-enquiry send back is allowed only after inspector report submission.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            inspector_id = parse_optional_int(request.form.get('inspector_id'))
            if not inspector_id:
                flash('Please select a valid field inspector.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            reenquiry_reason = request.form.get('comments', '').strip()
            if not reenquiry_reason:
                flash('Reason is required to send back for re-enquiry.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(reenquiry_reason) > 5000:
                flash('Reason is too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.cvo_send_back_to_inspector_for_reenquiry(
                petition_id, user_id, inspector_id, reenquiry_reason
            )
            flash('Sent back to field level for re-enquiry.', 'success')

        elif action == 'upload_consolidated_report':
            if user_role not in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Only CVO/DSP can upload consolidated report.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if petition and petition.get('status') != 'enquiry_report_submitted':
                flash('Consolidated report can be uploaded only after inspector report submission.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            consolidated_file = request.files.get('consolidated_report_file')
            if not consolidated_file or not consolidated_file.filename:
                flash('Please choose consolidated report PDF to upload.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            ok, upload_result = validate_pdf_upload(consolidated_file, 'Consolidated report upload')
            if not ok:
                flash(upload_result, 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            original_name = upload_result

            os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
            consolidated_filename = _build_storage_filename('cvo_consolidated', original_name, petition_id)
            if not consolidated_filename:
                flash('Unable to prepare consolidated report filename.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            saved_ok, save_result = _save_uploaded_file(consolidated_file, ENQUIRY_UPLOAD_DIR, consolidated_filename, 'Consolidated report file')
            if not saved_ok:
                flash(save_result, 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            consolidated_filename = save_result
            try:
                models.cvo_upload_consolidated_report(petition_id, user_id, consolidated_filename)
            except Exception:
                _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, consolidated_filename)
                raise
            flash('Consolidated report uploaded successfully.', 'success')

        elif action == 'request_detailed_enquiry':
            if user_role not in ('super_admin', 'cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Only CVO/DSP can request detailed enquiry.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if petition and petition.get('enquiry_type') != 'preliminary':
                flash('Detailed enquiry request is allowed only for preliminary petitions.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            cvo_comments = request.form.get('cvo_comments', '').strip()
            if not cvo_comments:
                flash('Remarks are required to request detailed enquiry.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            prima_facie_file = request.files.get('prima_facie_file')
            prima_facie_filename = None
            if prima_facie_file and prima_facie_file.filename:
                ok, upload_result = validate_pdf_upload(prima_facie_file, 'Prima Facie document')
                if not ok:
                    flash(upload_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                original_name = upload_result
                os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
                prima_facie_filename = _build_storage_filename('cvo_prima_facie', original_name, petition_id)
                if not prima_facie_filename:
                    flash('Unable to prepare prima facie filename.', 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                saved_ok, save_result = _save_uploaded_file(prima_facie_file, ENQUIRY_UPLOAD_DIR, prima_facie_filename, 'Prima facie file')
                if not saved_ok:
                    flash(save_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                prima_facie_filename = save_result
            try:
                models.cvo_request_detailed_enquiry(petition_id, user_id, cvo_comments, prima_facie_filename)
            except Exception:
                if prima_facie_filename:
                    _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, prima_facie_filename)
                raise
            flash('Detailed enquiry requested. Workflow restarted at PO permission stage.', 'success')
            
        elif action == 'give_conclusion':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can give conclusion.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            efile_no_input = request.form.get('efile_no', '').strip()
            final_conclusion = request.form.get('final_conclusion', '').strip()
            instructions = request.form.get('instructions', '').strip()
            conclusion_file = request.files.get('conclusion_file')
            conclusion_filename = None
            efile_no, efile_error = resolve_efile_no_for_action(
                petition,
                efile_no_input,
                required_message='E-Office File No is required for final conclusion.'
            )
            if efile_error:
                flash(efile_error, 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if not final_conclusion:
                flash('Final conclusion is required.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(final_conclusion) > 10000:
                flash('Final conclusion is too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(instructions) > 5000:
                flash('Instructions are too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            if conclusion_file and conclusion_file.filename:
                ok, upload_result = validate_pdf_upload(conclusion_file, 'Conclusion upload')
                if not ok:
                    flash(upload_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                original_name = upload_result
                os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
                conclusion_filename = _build_storage_filename('po_conclusion', original_name, petition_id)
                if not conclusion_filename:
                    flash('Unable to prepare conclusion filename.', 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                saved_ok, save_result = _save_uploaded_file(conclusion_file, ENQUIRY_UPLOAD_DIR, conclusion_filename, 'Conclusion file')
                if not saved_ok:
                    flash(save_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                conclusion_filename = save_result

            try:
                models.po_give_conclusion(petition_id, user_id, efile_no, final_conclusion, instructions, conclusion_filename)
            except Exception:
                if conclusion_filename:
                    _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, conclusion_filename)
                raise
            flash('Final conclusion submitted and petition closed.', 'success')

        elif action == 'send_to_cmd':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can forward petition to CMD/CGM-HR.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            efile_no_input = request.form.get('efile_no', '').strip()
            efile_no, efile_error = resolve_efile_no_for_action(
                petition,
                efile_no_input,
                required_message=f"{cfg_po_send_cmd_efile.get('label', 'E-Office File No')} is compulsory before sending to CMD/CGM-HR." if cfg_po_send_cmd_efile.get('required') else None
            )
            if efile_error:
                flash(efile_error, 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            cmd_instructions = request.form.get('cmd_instructions', '').strip()
            if cfg_po_send_cmd_instructions.get('required') and not cmd_instructions:
                flash(f"{cfg_po_send_cmd_instructions.get('label', 'CMD/CGM-HR Instructions')} is required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(cmd_instructions) > 5000:
                flash('CMD/CGM-HR instructions are too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            cmd_handler_id = parse_optional_int(request.form.get('cmd_handler_id'))
            if not cmd_handler_id:
                flash('Please select CMD/CGM-HR assignee for action.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.po_send_to_cmd(petition_id, user_id, cmd_instructions, efile_no, cmd_handler_id)
            flash('Petition forwarded to concerned CMD/CGM-HR for action.', 'success')

        elif action == 'po_send_back_reenquiry':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can send back for re-enquiry.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            if petition.get('status') != 'forwarded_to_po':
                flash('Re-enquiry send back is allowed only when report is pending with PO.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            reenquiry_reason = request.form.get('comments', '').strip()
            if not reenquiry_reason:
                flash('Reason is required to send back for re-enquiry.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(reenquiry_reason) > 5000:
                flash('Reason is too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.po_send_back_to_cvo_for_reenquiry(petition_id, user_id, reenquiry_reason)
            flash('Sent back to CVO/DSP for re-enquiry routing.', 'success')

        elif action == 'update_efile_no':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can update E-Office File No.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            efile_no = request.form.get('efile_no', '').strip()
            if not efile_no:
                flash('E-Office File No is required.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(efile_no) > 100:
                flash('E-Office File No is too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            if petition.get('requires_permission'):
                flash('This action is allowed only for direct enquiry petitions.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if petition.get('status') not in DIRECT_ENQUIRY_EFILE_EDITABLE_STATUSES:
                flash('E-Office File No can be updated only before enquiry report completion.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if (petition.get('efile_no') or '').strip():
                flash('E-Office File No is already set. Editing is not allowed.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            updated = models.po_update_efile_no(petition_id, user_id, efile_no)
            if not updated:
                flash('E-Office File No is already set. Editing is not allowed.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            flash('E-Office File No updated successfully.', 'success')

        elif action == 'cmd_submit_action_report':
            if user_role not in ('super_admin', 'cmd_apspdcl', 'cmd_apepdcl', 'cmd_apcpdcl', 'cgm_hr_transco'):
                flash('Only CMD can upload action taken report.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            form_cfg = get_effective_form_field_configs()
            cfg_action_taken = form_cfg.get('cmd_action.action_taken', DEFAULT_FORM_FIELD_CONFIGS['cmd_action.action_taken'])
            cfg_action_file = form_cfg.get('cmd_action.action_report_file', DEFAULT_FORM_FIELD_CONFIGS['cmd_action.action_report_file'])
            action_taken = request.form.get('action_taken', '').strip()
            if cfg_action_taken.get('required') and not action_taken:
                flash(f"{cfg_action_taken.get('label', 'Action taken details')} are required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(action_taken) > 10000:
                flash('Action taken details are too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))

            action_report_file = request.files.get('action_report_file')
            action_report_filename = None
            if cfg_action_file.get('required') and (not action_report_file or not action_report_file.filename):
                flash(f"{cfg_action_file.get('label', 'Action report copy')} is required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if action_report_file and action_report_file.filename:
                ok, upload_result = validate_pdf_upload(action_report_file, 'Action report upload')
                if not ok:
                    flash(upload_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                original_name = upload_result
                os.makedirs(ENQUIRY_UPLOAD_DIR, exist_ok=True)
                action_report_filename = _build_storage_filename('cmd_action', original_name, petition_id)
                if not action_report_filename:
                    flash('Unable to prepare action report filename.', 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                saved_ok, save_result = _save_uploaded_file(action_report_file, ENQUIRY_UPLOAD_DIR, action_report_filename, 'Action report file')
                if not saved_ok:
                    flash(save_result, 'danger')
                    return redirect(url_for('petition_view', petition_id=petition_id))
                action_report_filename = save_result

            try:
                models.cmd_submit_action_report(petition_id, user_id, action_taken, action_report_filename)
            except Exception:
                if action_report_filename:
                    _delete_uploaded_file(ENQUIRY_UPLOAD_DIR, action_report_filename)
                raise
            flash('Action taken recorded and copy sent to PO for closure.', 'success')

        elif action == 'po_lodge':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can lodge petition.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if not petition:
                flash('Petition not found.', 'danger')
                return redirect(url_for('petitions_list'))
            lodge_remarks = request.form.get('lodge_remarks', '').strip()
            if cfg_po_lodge_remarks.get('required') and not lodge_remarks:
                flash(f"{cfg_po_lodge_remarks.get('label', 'PO Lodge Remarks')} is required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(lodge_remarks) > 5000:
                flash('Lodge remarks are too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            efile_no_input = request.form.get('efile_no', '').strip()
            efile_no, efile_error = resolve_efile_no_for_action(
                petition,
                efile_no_input,
                required_message=f"{cfg_po_lodge_efile.get('label', 'E-Office File No')} is required." if cfg_po_lodge_efile.get('required') else None
            )
            if efile_error:
                flash(efile_error, 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.po_lodge_petition(petition_id, user_id, lodge_remarks, efile_no)
            flash('Petition lodged in PO login.', 'success')

        elif action == 'po_direct_lodge':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can directly lodge petition.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if petition and petition.get('status') not in ('sent_for_permission',):
                flash('Direct lodge without enquiry is allowed only at permission stage.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            lodge_remarks = request.form.get('lodge_remarks', '').strip()
            if cfg_po_direct_lodge_remarks.get('required') and not lodge_remarks:
                flash(f"{cfg_po_direct_lodge_remarks.get('label', 'PO Lodge Remarks')} is required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(lodge_remarks) > 5000:
                flash('Lodge remarks are too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            efile_no_input = request.form.get('efile_no', '').strip()
            efile_no, efile_error = resolve_efile_no_for_action(
                petition,
                efile_no_input,
                required_message=f"{cfg_po_direct_lodge_efile.get('label', 'E-Office File No')} is required." if cfg_po_direct_lodge_efile.get('required') else None
            )
            if efile_error:
                flash(efile_error, 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.po_direct_lodge_no_enquiry(petition_id, user_id, lodge_remarks, efile_no)
            flash('Petition directly lodged by PO (no enquiry/action required).', 'success')
            
        elif action == 'close':
            if user_role not in ('super_admin', 'po'):
                flash('Only PO can close petition.', 'danger')
                return redirect(url_for('petition_view', petition_id=petition_id))
            petition = models.get_petition_by_id(petition_id)
            if petition and petition.get('status') != 'lodged':
                flash('Petition can be closed only after Lodged stage.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if cfg_po_close_comments.get('required') and not comments:
                flash(f"{cfg_po_close_comments.get('label', 'Closing Remarks')} is required.", 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            if len(comments) > 5000:
                flash('Closing remarks are too long.', 'warning')
                return redirect(url_for('petition_view', petition_id=petition_id))
            models.close_petition(petition_id, user_id, comments)
            flash('Petition closed.', 'success')
        else:
            flash('Unsupported action.', 'warning')
            
    except Exception:
        flash_internal_error('Unable to complete action. Please contact administrator.')
    
    return redirect(url_for('petition_view', petition_id=petition_id))

@app.route('/e-receipts/<path:filename>')
@login_required
def ereceipt_file(filename):
    filename = _normalize_storage_relpath(filename)
    if not filename:
        return Response(status=404)
    requested_petition_id = _parse_requested_petition_id(request.args.get('petition_id'))
    petition_id = _resolve_petition_id_for_file(filename)
    if requested_petition_id and petition_id and requested_petition_id != petition_id:
        log_security_event(
            'access.file_mismatch',
            severity='warning',
            file_type='e_receipt',
            petition_id=petition_id,
            requested_petition_id=requested_petition_id,
        )
        return Response(status=404)
    if requested_petition_id and not petition_id:
        petition_id = requested_petition_id
    if not petition_id:
        log_security_event('access.file_unresolved', severity='warning', file_type='e_receipt')
        return Response(status=404)
    if not _can_access_petition(petition_id):
        log_security_event('access.file_forbidden', severity='warning', petition_id=petition_id, file_type='e_receipt')
        flash('You do not have access to this file.', 'danger')
        return redirect(url_for('petitions_list'))
    if not _uploaded_file_exists(ERECEIPT_UPLOAD_DIR, filename):
        log_security_event('access.file_missing', severity='info', petition_id=petition_id, file_type='e_receipt')
        flash('No file uploaded.', 'warning')
        return redirect(url_for('petition_view', petition_id=petition_id))
    return send_from_directory(ERECEIPT_UPLOAD_DIR, filename, as_attachment=False)

@app.route('/enquiry-files/<path:filename>')
@login_required
def enquiry_file(filename):
    filename = _normalize_storage_relpath(filename)
    if not filename:
        return Response(status=404)
    requested_petition_id = _parse_requested_petition_id(request.args.get('petition_id'))
    petition_id = _resolve_petition_id_for_file(filename)
    if requested_petition_id and petition_id and requested_petition_id != petition_id:
        log_security_event(
            'access.file_mismatch',
            severity='warning',
            file_type='enquiry',
            petition_id=petition_id,
            requested_petition_id=requested_petition_id,
        )
        return Response(status=404)
    if requested_petition_id and not petition_id:
        petition_id = requested_petition_id
    if not petition_id:
        log_security_event('access.file_unresolved', severity='warning', file_type='enquiry')
        return Response(status=404)
    if not _can_access_petition(petition_id):
        log_security_event('access.file_forbidden', severity='warning', petition_id=petition_id, file_type='enquiry')
        flash('You do not have access to this file.', 'danger')
        return redirect(url_for('petitions_list'))
    if not _uploaded_file_exists(ENQUIRY_UPLOAD_DIR, filename):
        log_security_event('access.file_missing', severity='info', petition_id=petition_id, file_type='enquiry')
        flash('No file uploaded.', 'warning')
        return redirect(url_for('petition_view', petition_id=petition_id))
    return send_from_directory(ENQUIRY_UPLOAD_DIR, filename, as_attachment=False)


@app.route('/profile-photos/<path:filename>')
@login_required
def profile_photo_file(filename):
    filename = _normalize_storage_relpath(filename)
    if not filename:
        return Response(status=404)
    owner_match = re.match(r'^user_(\d+)_', os.path.basename(filename))
    if owner_match:
        owner_id = int(owner_match.group(1))
        if session.get('user_role') != 'super_admin' and owner_id != int(session.get('user_id') or 0):
            log_security_event('access.profile_photo_forbidden', severity='warning', owner_id=owner_id)
            return Response(status=403)
    return send_from_directory(PROFILE_UPLOAD_DIR, filename, as_attachment=False)


@app.route('/petition-search')
def petition_search_public():
    raise NotFound()
    """Public petition status lookup — returns minimal info (no PII)."""
    q = (request.args.get('q') or '').strip()
    field = (request.args.get('field') or 'efile_no').strip()
    office = (request.args.get('office') or '').strip()
    if not q or len(q) < 3:
        return jsonify({'results': [], 'message': 'Please enter at least 3 characters.'})
    if field not in ('efile_no', 'ereceipt_no'):
        field = 'efile_no'
    if office and office not in VALID_RECEIVED_AT:
        office = ''
    OFFICE_LABELS = {
        'jmd_office': 'JMD Office, Hyderabad',
        'cvo_apspdcl_tirupathi': 'CVO/DSP APSPDCL – Tirupathi',
        'cvo_apepdcl_vizag': 'CVO/DSP APEPDCL – Vizag',
        'cvo_apcpdcl_vijayawada': 'CVO/DSP APCPDCL – Vijayawada',
    }
    STATUS_LABELS = {
        'received': 'Received',
        'forwarded_to_cvo': 'Forwarded to CVO/DSP',
        'sent_for_permission': 'Sent for Permission',
        'permission_approved': 'Permission Approved',
        'permission_rejected': 'Permission Rejected',
        'assigned_to_inspector': 'Assigned to Inspector',
        'sent_back_for_reenquiry': 'Sent Back for Re-Enquiry',
        'enquiry_in_progress': 'Enquiry in Progress',
        'enquiry_report_submitted': 'Enquiry Report Submitted',
        'cvo_comments_added': 'CVO Comments Added',
        'forwarded_to_po': 'Forwarded to PO',
        'forwarded_to_jmd': 'Forwarded to JMD',
        'action_instructed': 'Action Instructed',
        'closed': 'Closed',
    }
    try:
        results = models.public_petition_status_lookup(q, field, office or None)
        out = []
        for r in results:
            out.append({
                'sno': r.get('sno') or '—',
                'status': STATUS_LABELS.get(r.get('status'), str(r.get('status') or '').replace('_', ' ').title()),
                'received_date': r['received_date'].strftime('%d %b %Y') if r.get('received_date') else '—',
                'office': OFFICE_LABELS.get(r.get('received_at'), str(r.get('received_at') or '').replace('_', ' ').title()),
            })
        return jsonify({'results': out})
    except Exception:
        return jsonify({'results': [], 'message': 'Could not connect to database. Please try again.'})


@app.route('/help-resources/files/<path:filename>')
@login_required
def help_resource_file(filename):
    filename = _normalize_storage_relpath(filename)
    if not filename:
        return Response(status=404)
    resource = models.get_help_resource_by_file_name(filename)
    if not resource:
        return Response(status=404)
    if session.get('user_role') not in ('super_admin', 'po') and not resource.get('is_active'):
        log_security_event('access.help_resource_inactive_forbidden', severity='warning', resource_id=resource.get('id'))
        return Response(status=404)
    if not _uploaded_file_exists(HELP_RESOURCE_UPLOAD_DIR, filename):
        return Response(status=404)
    lower_name = filename.lower()
    mime_type = str(resource.get('mime_type') or '').strip().lower()
    force_download = lower_name.endswith('.svg') or mime_type == 'image/svg+xml'
    return send_from_directory(HELP_RESOURCE_UPLOAD_DIR, filename, as_attachment=force_download)


@app.route('/help-center')
@app.route('/help-management')
@login_required
def help_center():
    return redirect(url_for('help_page'))


def _build_grouped_resources(active_only=True):
    resources = models.list_help_resources(active_only=active_only)
    grouped_resources = {key: [] for key in ('manual', 'flowchart', 'video', 'office_order', 'news')}
    for resource in resources:
        entry = dict(resource)
        mime_type = (entry.get('mime_type') or '').strip().lower()
        file_name = (entry.get('file_name') or '').strip().lower()
        if entry.get('storage_kind') == 'upload' and entry.get('file_name'):
            entry['view_url'] = url_for('help_resource_file', filename=entry['file_name'])
        elif entry.get('storage_kind') == 'external_url':
            entry['view_url'] = entry.get('external_url')
        else:
            entry['view_url'] = None
        entry['preview_kind'] = None
        if entry.get('storage_kind') == 'upload' and entry.get('view_url'):
            if mime_type == 'application/pdf' or file_name.endswith('.pdf'):
                entry['preview_kind'] = 'pdf'
            elif mime_type.startswith('image/') or file_name.endswith(('.jpg', '.jpeg', '.png', '.webp')):
                entry['preview_kind'] = 'image'
            elif mime_type.startswith('video/') or file_name.endswith(('.mp4', '.webm', '.mov')):
                entry['preview_kind'] = 'video'
        grouped_resources.setdefault(entry.get('resource_type') or 'manual', []).append(entry)
    return grouped_resources


@app.route('/help', methods=['GET', 'POST'])
@login_required
def help_page():
    if request.method == 'POST':
        if session.get('user_role') not in ('super_admin', 'po'):
            return Response(status=403)
        action = (request.form.get('action') or 'upload').strip()
        if action == 'toggle_active':
            resource_id = parse_optional_int(request.form.get('resource_id'))
            should_activate = request.form.get('activate') == '1'
            if not resource_id:
                flash('Invalid help resource selection.', 'warning')
                return redirect(url_for('help_page'))
            try:
                models.set_help_resource_active(resource_id, should_activate)
                flash('Help resource visibility updated.', 'success')
            except Exception:
                flash_internal_error('Unable to update help resource visibility. Please contact administrator.')
            return redirect(url_for('help_page'))

        title = (request.form.get('title') or '').strip()
        resource_type = (request.form.get('resource_type') or '').strip()
        storage_kind = (request.form.get('storage_kind') or 'upload').strip()
        external_url = (request.form.get('external_url') or '').strip() or None
        display_order_raw = (request.form.get('display_order') or '0').strip()
        upload = request.files.get('resource_file')

        if not title:
            flash('Resource title is required.', 'warning')
            return redirect(url_for('help_page'))
        if resource_type not in HELP_RESOURCE_TYPES:
            flash('Please select a valid resource type.', 'warning')
            return redirect(url_for('help_page'))
        if storage_kind not in HELP_RESOURCE_STORAGE_KINDS:
            flash('Please select a valid resource source.', 'warning')
            return redirect(url_for('help_page'))
        try:
            display_order = int(display_order_raw or '0')
        except ValueError:
            flash('Display order must be a number.', 'warning')
            return redirect(url_for('help_page'))

        file_name = None
        mime_type = None
        if storage_kind == 'upload':
            ok_upload, stored_name, detected_mime_type, upload_error = validate_help_resource_upload(upload)
            if not ok_upload:
                flash(upload_error, 'warning')
                return redirect(url_for('help_page'))
            if not stored_name or not upload:
                flash('Please choose a file to upload.', 'warning')
                return redirect(url_for('help_page'))
            ensure_upload_dirs()
            saved_ok, save_result = _save_uploaded_file(upload, HELP_RESOURCE_UPLOAD_DIR, stored_name, 'Help resource', use_date_subdir=True)
            if not saved_ok:
                flash(save_result, 'danger')
                return redirect(url_for('help_page'))
            file_name = save_result
            mime_type = detected_mime_type
        else:
            parsed = urllib.parse.urlparse(external_url or '')
            if parsed.scheme not in {'http', 'https'} or not parsed.netloc:
                flash('Please provide a valid external URL.', 'warning')
                return redirect(url_for('help_page'))

        try:
            models.create_help_resource(
                title=title,
                resource_type=resource_type,
                storage_kind=storage_kind,
                file_name=file_name,
                external_url=external_url,
                mime_type=mime_type,
                display_order=display_order,
                uploaded_by=session['user_id'],
            )
            flash('Help resource added successfully.', 'success')
        except Exception:
            if file_name:
                _delete_uploaded_file(HELP_RESOURCE_UPLOAD_DIR, file_name)
            flash_internal_error('Unable to save help resource. Please contact administrator.')
        return redirect(url_for('help_page'))

    is_admin = session.get('user_role') in ('super_admin', 'po')
    resources = models.list_help_resources(active_only=False) if is_admin else []
    grouped_resources = _build_grouped_resources(active_only=True)
    return render_template('help_management.html', resources=resources, grouped_resources=grouped_resources, is_admin=is_admin, bulk_import_headers=IMPORT_PETITION_HEADERS)


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user = _get_user_by_id_cached(session['user_id'])
    if not user:
        flash('User profile not found.', 'danger')
        return redirect(url_for('logout'))

    if request.method == 'POST':
        full_name = (request.form.get('full_name') or '').strip()
        username = (request.form.get('username') or '').strip()
        phone = (request.form.get('phone') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        current_password = (request.form.get('current_password') or '').strip()
        new_password = (request.form.get('new_password') or '').strip()
        confirm_password = (request.form.get('confirm_password') or '').strip()
        remove_photo = request.form.get('remove_photo') == 'on'
        photo_upload = request.files.get('profile_photo')

        if not full_name or len(full_name) < 3:
            flash('Name must be at least 3 characters.', 'warning')
            return redirect(url_for('profile'))
        if not username or len(username) < 3:
            flash('Username must be at least 3 characters.', 'warning')
            return redirect(url_for('profile'))
        if not re.match(r'^[A-Za-z0-9_.-]+$', username):
            flash('Username can only contain letters, numbers, dot, underscore, and hyphen.', 'warning')
            return redirect(url_for('profile'))
        if not validate_contact(phone):
            flash('Please provide a valid phone number.', 'warning')
            return redirect(url_for('profile'))
        if not validate_email(email):
            flash('Please provide a valid email address.', 'warning')
            return redirect(url_for('profile'))

        if new_password:
            if not current_password:
                flash('Current password is required to set a new password.', 'warning')
                return redirect(url_for('profile'))
            if not models.authenticate_user(user.get('username') or '', current_password):
                log_security_event('auth.profile_password_change_rejected', severity='warning')
                flash('Current password is incorrect.', 'warning')
                return redirect(url_for('profile'))
            ok_password, password_error = validate_password_strength(new_password, 'New password')
            if not ok_password:
                flash(password_error, 'warning')
                return redirect(url_for('profile'))
            if new_password != confirm_password:
                flash('Password confirmation does not match.', 'warning')
                return redirect(url_for('profile'))

        ok_photo, stored_photo_name, photo_error = validate_profile_photo_upload(photo_upload, session['user_id'])
        if not ok_photo:
            flash(photo_error, 'warning')
            return redirect(url_for('profile'))

        old_photo = user.get('profile_photo')
        photo_changed = False
        try:
            if username != user.get('username'):
                models.set_username(session['user_id'], username)

            models.update_user_profile_info(session['user_id'], full_name, phone, email)

            if new_password:
                models.set_user_password(session['user_id'], new_password)

            if stored_photo_name and photo_upload:
                ensure_upload_dirs()
                saved_ok, save_result = _save_uploaded_file(photo_upload, PROFILE_UPLOAD_DIR, stored_photo_name, 'Profile photo')
                if not saved_ok:
                    flash(save_result, 'danger')
                    return redirect(url_for('profile'))
                stored_photo_name = save_result
                models.set_user_profile_photo(session['user_id'], stored_photo_name)
                photo_changed = True
            elif remove_photo and old_photo:
                models.set_user_profile_photo(session['user_id'], None)
                photo_changed = True

            if photo_changed and old_photo and old_photo != stored_photo_name:
                delete_profile_photo_file(old_photo)

            if new_password:
                # Re-fetch user to get the incremented session_version, then
                # activate a fresh session.
                #
                # Idempotency guard: double-clicking Save would bump
                # session_version twice.  The browser receives two responses
                # with cookies carrying different versions; whichever cookie
                # wins the race has the wrong version → version-mismatch
                # logout on the next click.
                #
                # We bind the activation to the version we just wrote.  If
                # another concurrent request already bumped the version again,
                # that request "owns" the final session; this one just exits.
                version_at_change = int(user.get('session_version') or 1) + 1
                updated_user = models.get_user_by_id(session['user_id'])
                if updated_user and updated_user.get('is_active') is not False:
                    actual_version = int(updated_user.get('session_version') or 1)
                    if actual_version == version_at_change:
                        # We own this version bump — activate session.
                        _activate_login_session(updated_user)
                        log_security_event('auth.profile_password_changed', severity='info')
                        flash('Password updated successfully.', 'success')
                    else:
                        # A concurrent request already claimed the new version.
                        # Do nothing extra — the concurrent response carries the
                        # correct cookie; just flash and let the browser follow
                        # the redirect normally (session is still valid via the
                        # concurrent response's cookie).
                        log_security_event('auth.profile_password_changed', severity='info')
                        flash('Password updated successfully.', 'success')
                    return redirect(url_for('profile'))
                session.clear()
                flash('Password updated successfully. Please login again.', 'success')
                return redirect(url_for('login'))

            refresh_session_user()
            flash('Profile updated successfully.', 'success')
            return redirect(url_for('profile'))
        except Exception as e:
            if stored_photo_name:
                delete_profile_photo_file(stored_photo_name)
            error_text = str(e).lower()
            if 'unique' in error_text or 'duplicate key' in error_text:
                flash('Username already exists. Choose a different username.', 'danger')
            else:
                flash_internal_error('Unable to update profile. Please contact administrator.')
            return redirect(url_for('profile'))

    return render_template('profile.html', user=user)

# ========================================
# USER MANAGEMENT (SUPER ADMIN)
# ========================================

@app.route('/users')
@login_required
@role_required('super_admin')
def users_list():
    users = models.get_all_users()
    cvo_users = models.get_cvo_users()
    role_login_users = models.get_role_login_users()
    inspector_mappings = models.get_inspector_mappings()
    try:
        reset_requests = models.get_pending_password_reset_requests()
    except Exception:
        reset_requests = []
    return render_template(
        'users.html',
        users=users,
        cvo_users=cvo_users,
        role_login_users=role_login_users,
        inspector_mappings=inspector_mappings,
        reset_requests=reset_requests,
    )

@app.route('/users/password-reset-requests/<int:request_id>/approve', methods=['POST'])
@login_required
@role_required('super_admin')
def approve_password_reset_request(request_id):
    try:
        models.approve_password_reset_request(request_id, session['user_id'])
        flash('Password reset request approved.', 'success')
    except Exception:
        flash_internal_error('Unable to approve password reset request. Please contact administrator.')
    return redirect(url_for('users_list'))


@app.route('/users/password-reset-requests/<int:request_id>/reject', methods=['POST'])
@login_required
@role_required('super_admin')
def reject_password_reset_request(request_id):
    note = (request.form.get('decision_notes') or '').strip()
    try:
        models.reject_password_reset_request(request_id, session['user_id'], note)
        flash('Password reset request rejected.', 'success')
    except Exception:
        flash_internal_error('Unable to reject password reset request. Please contact administrator.')
    return redirect(url_for('users_list'))


@app.route('/form-management', methods=['GET', 'POST'])
@login_required
@role_required('super_admin')
def form_management():
    if request.method == 'POST':
        action = (request.form.get('action') or 'update_field').strip()
        if action == 'add_field':
            form_key = (request.form.get('new_form_key') or '').strip()
            field_key = (request.form.get('new_field_key') or '').strip().lower()
            label = (request.form.get('new_label') or '').strip()
            field_type = (request.form.get('new_field_type') or 'text').strip()
            is_required = request.form.get('new_is_required') == 'on'
            if form_key not in FORM_MANAGEMENT_GROUPS:
                flash('Please select a valid form group.', 'danger')
                return redirect(url_for('form_management'))
            if not re.fullmatch(r'[a-z][a-z0-9_]{1,80}', field_key):
                flash('Field key must use lowercase letters, numbers, underscore (2-81 chars).', 'warning')
                return redirect(url_for('form_management'))
            if not label:
                flash('Field label is required.', 'warning')
                return redirect(url_for('form_management'))
            if field_type not in VALID_DYNAMIC_FIELD_TYPES:
                flash('Invalid field type.', 'danger')
                return redirect(url_for('form_management'))
            config_key = f'{form_key}.{field_key}'
            if config_key in get_effective_form_field_configs():
                flash('Field key already exists for this form group.', 'warning')
                return redirect(url_for('form_management'))
            try:
                models.upsert_form_field_config(
                    form_key=form_key,
                    field_key=field_key,
                    label=label,
                    field_type=field_type,
                    is_required=is_required,
                    options=[],
                    updated_by=session['user_id']
                )
                flash('New form field added successfully.', 'success')
            except Exception:
                flash_internal_error('Unable to add form field. Please contact administrator.')
            return redirect(url_for('form_management'))

        form_key = (request.form.get('form_key') or '').strip()
        field_key = (request.form.get('field_key') or '').strip()
        config_key = f'{form_key}.{field_key}'
        effective_cfg = get_effective_form_field_configs()
        if config_key not in effective_cfg:
            flash('Invalid form field selection.', 'danger')
            return redirect(url_for('form_management'))

        label = (request.form.get('label') or '').strip() or effective_cfg[config_key]['label']
        field_type = (request.form.get('field_type') or '').strip()
        if field_type not in VALID_DYNAMIC_FIELD_TYPES:
            flash('Invalid field type.', 'danger')
            return redirect(url_for('form_management'))

        is_required = request.form.get('is_required') == 'on'
        options = []
        if effective_cfg[config_key]['type'] == 'select' or field_type == 'select':
            raw_options = request.form.get('options_text', '')
            for line in raw_options.splitlines():
                item = line.strip()
                if not item:
                    continue
                if '|' in item:
                    value, label_text = item.split('|', 1)
                    value = value.strip()
                    label_text = label_text.strip()
                else:
                    value = item.strip()
                    label_text = item.strip()
                if value and label_text:
                    options.append({'value': value, 'label': label_text})

            if not options:
                options = effective_cfg[config_key].get('options', [])

        try:
            models.upsert_form_field_config(
                form_key, field_key, label, field_type, is_required, options, session['user_id']
            )
            flash('Form field updated successfully.', 'success')
        except Exception:
            flash_internal_error('Unable to update form field. Please contact administrator.')
        return redirect(url_for('form_management'))

    effective = get_effective_form_field_configs()
    grouped = {}
    for key, cfg in effective.items():
        fk, field = key.split('.', 1)
        grouped.setdefault(fk, []).append({'form_key': fk, 'field_key': field, **cfg, 'config_key': key})
    for fk in grouped:
        grouped[fk] = sorted(grouped[fk], key=lambda x: x['field_key'])

    return render_template(
        'form_management.html',
        grouped_fields=grouped,
        form_groups=FORM_MANAGEMENT_GROUPS,
        field_types=sorted(VALID_DYNAMIC_FIELD_TYPES),
    )


@app.route('/system-settings', methods=['GET', 'POST'])
@login_required
@role_required('super_admin')
def system_settings():
    if request.method == 'POST':
        updates = {}
        for key, meta in SYSTEM_SETTING_DEFINITIONS.items():
            raw_value = (request.form.get(key) or '').strip()
            if not raw_value:
                flash(f"{meta['label']} is required.", 'warning')
                return redirect(url_for('system_settings'))
            try:
                parsed = int(raw_value)
            except ValueError:
                flash(f"{meta['label']} must be a whole number.", 'warning')
                return redirect(url_for('system_settings'))
            if parsed < int(meta['min']) or parsed > int(meta['max']):
                flash(f"{meta['label']} must be between {meta['min']} and {meta['max']}.", 'warning')
                return redirect(url_for('system_settings'))
            updates[key] = parsed

        try:
            models.upsert_system_settings(updates, session['user_id'])
            if has_request_context() and hasattr(g, '_effective_system_settings'):
                delattr(g, '_effective_system_settings')
            flash('System settings updated successfully.', 'success')
            log_security_event(
                'admin.system_settings_updated',
                severity='info',
                updated_by=session.get('user_id'),
                setting_keys=','.join(sorted(updates.keys())),
            )
        except Exception:
            flash_internal_error('Unable to update system settings. Please contact administrator.')
        return redirect(url_for('system_settings'))

    return render_template(
        'system_settings.html',
        setting_rows=_system_settings_rows(),
    )


@app.route('/users/new', methods=['POST'])
@login_required
@role_required('super_admin')
def user_create():
    try:
        username = request.form.get('username', '').strip()
        # Password is always the system default; users must change it on first login.
        password = _DEFAULT_PASSWORD
        full_name = request.form.get('full_name', '').strip()
        role = (request.form.get('role') or '').strip()
        cvo_office = (request.form.get('cvo_office') or '').strip() or None
        assigned_cvo_id = parse_optional_int(request.form.get('assigned_cvo_id'))
        phone = request.form.get('phone', '').strip() or None
        email = request.form.get('email', '').strip() or None

        if not username or len(username) < 3:
            flash('Username must be at least 3 characters.', 'warning')
            return redirect(url_for('users_list'))
        if not full_name or len(full_name) < 3:
            flash('Officer name must be at least 3 characters.', 'warning')
            return redirect(url_for('users_list'))
        if role not in VALID_USER_ROLES:
            flash('Please select a valid role.', 'warning')
            return redirect(url_for('users_list'))
        if cvo_office and cvo_office not in VALID_CVO_OFFICES:
            flash('Please select a valid office.', 'warning')
            return redirect(url_for('users_list'))
        if not validate_contact(phone):
            flash('Please provide a valid phone number.', 'warning')
            return redirect(url_for('users_list'))
        if not validate_email(email):
            flash('Please provide a valid email address.', 'warning')
            return redirect(url_for('users_list'))

        if role == 'inspector':
            if not cvo_office:
                flash('Office is required for inspector role.', 'warning')
                return redirect(url_for('users_list'))
            if not assigned_cvo_id:
                flash('Please assign inspector to a CVO/DSP.', 'warning')
                return redirect(url_for('users_list'))
        elif role == 'data_entry':
            if not cvo_office:
                flash('Office is required for Data Entry role.', 'warning')
                return redirect(url_for('users_list'))
            assigned_cvo_id = None
        elif role.startswith('cvo_') or role == 'dsp':
            if not cvo_office:
                flash('Office is required for CVO/DSP role.', 'warning')
                return redirect(url_for('users_list'))
            assigned_cvo_id = None
        else:
            assigned_cvo_id = None
            cvo_office = None

        if assigned_cvo_id:
            cvo_user = _get_user_by_id_cached(assigned_cvo_id)
            if not cvo_user or cvo_user.get('role') not in ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Assigned CVO/DSP is invalid.', 'warning')
                return redirect(url_for('users_list'))
        
        models.create_user(username, password, full_name, role, cvo_office, assigned_cvo_id, phone, email)
        flash(f'User {username} created successfully!', 'success')
    except Exception:
        flash_internal_error('Unable to create user. Please contact administrator.')
    
    return redirect(url_for('users_list'))


@app.route('/users/upload', methods=['POST'])
@login_required
@role_required('super_admin')
def users_upload():
    upload = request.files.get('users_file')
    if not upload or not upload.filename:
        flash('Please choose an Excel/CSV file to upload.', 'warning')
        return redirect(url_for('users_list'))

    filename = secure_filename(upload.filename)
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in ('xlsx', 'csv'):
        flash('Only .xlsx or .csv files are allowed for bulk user creation.', 'danger')
        return redirect(url_for('users_list'))

    required_headers = {'username', 'full_name', 'role'}
    rows = []
    try:
        if ext == 'csv':
            content = upload.stream.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            headers = {h.strip().lower() for h in (reader.fieldnames or [])}
            if not required_headers.issubset(headers):
                flash('Missing required columns. Required: username,password,full_name,role', 'danger')
                return redirect(url_for('users_list'))
            for row in reader:
                rows.append({(k or '').strip().lower(): (v or '').strip() for k, v in row.items()})
        else:
            if load_workbook is None:
                flash('Excel support requires openpyxl dependency. Install and retry.', 'danger')
                return redirect(url_for('users_list'))
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                flash('Uploaded file is empty.', 'warning')
                return redirect(url_for('users_list'))
            headers = [str(h).strip().lower() if h is not None else '' for h in all_rows[0]]
            if not required_headers.issubset(set(headers)):
                flash('Missing required columns. Required: username,password,full_name,role', 'danger')
                return redirect(url_for('users_list'))
            for r in all_rows[1:]:
                data = {}
                for idx, col in enumerate(headers):
                    if not col:
                        continue
                    value = r[idx] if idx < len(r) else ''
                    data[col] = str(value).strip() if value is not None else ''
                if any(v for v in data.values()):
                    rows.append(data)
    except Exception:
        app.logger.exception('Unable to parse users bulk upload file')
        flash('Unable to parse upload file. Please verify format and retry.', 'danger')
        return redirect(url_for('users_list'))

    created = 0
    failed = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        username = row.get('username', '').strip()
        # Password column is ignored — every user starts with the system default.
        password = _DEFAULT_PASSWORD
        full_name = row.get('full_name', '').strip()
        role = row.get('role', '').strip().lower()
        cvo_office = row.get('cvo_office', '').strip().lower() or None
        assigned_cvo_username = row.get('assigned_cvo_username', '').strip() or None
        phone = row.get('phone', '').strip() or None
        email = row.get('email', '').strip() or None

        if not username or not full_name or not role:
            failed += 1
            errors.append(f'Row {i}: required values missing.')
            continue
        if role not in VALID_USER_ROLES:
            failed += 1
            errors.append(f'Row {i}: invalid role "{role}".')
            continue
        if cvo_office and cvo_office not in VALID_CVO_OFFICES:
            failed += 1
            errors.append(f'Row {i}: invalid cvo_office "{cvo_office}".')
            continue
        if len(username) < 3:
            failed += 1
            errors.append(f'Row {i}: username must be at least 3 characters.')
            continue
        if len(full_name) < 3:
            failed += 1
            errors.append(f'Row {i}: full_name must be at least 3 characters.')
            continue
        if not validate_contact(phone):
            failed += 1
            errors.append(f'Row {i}: invalid phone.')
            continue
        if not validate_email(email):
            failed += 1
            errors.append(f'Row {i}: invalid email.')
            continue
        if role == 'inspector' and (not cvo_office or not assigned_cvo_username):
            failed += 1
            errors.append(f'Row {i}: inspector requires cvo_office and assigned_cvo_username.')
            continue
        if role == 'data_entry' and not cvo_office:
            failed += 1
            errors.append(f'Row {i}: data_entry requires cvo_office.')
            continue
        if (role.startswith('cvo_') or role == 'dsp') and not cvo_office:
            failed += 1
            errors.append(f'Row {i}: cvo_office is required for {role}.')
            continue
        if role != 'inspector' and assigned_cvo_username:
            failed += 1
            errors.append(f'Row {i}: assigned_cvo_username is allowed only for inspector role.')
            continue

        assigned_cvo_id = None
        if assigned_cvo_username:
            cvo_user = models.get_user_by_username(assigned_cvo_username)
            if not cvo_user or cvo_user.get('role') not in ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                failed += 1
                errors.append(f'Row {i}: assigned_cvo_username "{assigned_cvo_username}" is invalid.')
                continue
            assigned_cvo_id = cvo_user['id']

        try:
            models.create_user(username, password, full_name, role, cvo_office, assigned_cvo_id, phone, email)
            created += 1
        except Exception:
            failed += 1
            app.logger.exception('Bulk user row failed at row %s', i)
            errors.append(f'Row {i}: internal processing error.')

    if created:
        flash(f'Bulk user upload complete. Created: {created}, Failed: {failed}.', 'success')
    if failed:
        preview = '; '.join(errors[:5])
        if len(errors) > 5:
            preview += '; ...'
        flash(f'Upload errors: {preview}', 'warning')

    return redirect(url_for('users_list'))

@app.route('/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@role_required('super_admin')
def user_toggle(user_id):
    try:
        models.toggle_user_status(user_id)
        flash('User status updated.', 'success')
    except Exception:
        flash_internal_error('Unable to update user status. Please contact administrator.')
    return redirect(url_for('users_list'))

@app.route('/users/<int:user_id>/edit', methods=['POST'])
@login_required
@role_required('super_admin')
def user_edit(user_id):
    try:
        full_name = request.form.get('full_name', '').strip()
        role = (request.form.get('role') or '').strip()
        cvo_office = (request.form.get('cvo_office') or '').strip() or None
        assigned_cvo_id = parse_optional_int(request.form.get('assigned_cvo_id'))
        phone = request.form.get('phone', '').strip() or None
        email = request.form.get('email', '').strip() or None
        password = request.form.get('password', '').strip() or None

        if not full_name or len(full_name) < 3:
            flash('Officer name must be at least 3 characters.', 'warning')
            return redirect(url_for('users_list'))
        if role not in VALID_USER_ROLES:
            flash('Please select a valid role.', 'warning')
            return redirect(url_for('users_list'))
        if cvo_office and cvo_office not in VALID_CVO_OFFICES:
            flash('Please select a valid office.', 'warning')
            return redirect(url_for('users_list'))
        if password:
            ok_password, password_error = validate_password_strength(password, 'Password')
            if not ok_password:
                flash(password_error, 'warning')
                return redirect(url_for('users_list'))
        if not validate_contact(phone):
            flash('Please provide a valid phone number.', 'warning')
            return redirect(url_for('users_list'))
        if not validate_email(email):
            flash('Please provide a valid email address.', 'warning')
            return redirect(url_for('users_list'))
        if role == 'inspector':
            if not cvo_office:
                flash('Office is required for inspector role.', 'warning')
                return redirect(url_for('users_list'))
            if not assigned_cvo_id:
                flash('Please assign inspector to a CVO/DSP.', 'warning')
                return redirect(url_for('users_list'))
        elif role == 'data_entry':
            if not cvo_office:
                flash('Office is required for Data Entry role.', 'warning')
                return redirect(url_for('users_list'))
            assigned_cvo_id = None
        elif role.startswith('cvo_') or role == 'dsp':
            if not cvo_office:
                flash('Office is required for CVO/DSP role.', 'warning')
                return redirect(url_for('users_list'))
            assigned_cvo_id = None
        else:
            cvo_office = None
            assigned_cvo_id = None
        if assigned_cvo_id:
            cvo_user = _get_user_by_id_cached(assigned_cvo_id)
            if not cvo_user or cvo_user.get('role') not in ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                flash('Assigned CVO/DSP is invalid.', 'warning')
                return redirect(url_for('users_list'))
        
        models.update_user(user_id, full_name, role, cvo_office, assigned_cvo_id, phone, email, password)
        flash('User updated successfully!', 'success')
    except Exception:
        flash_internal_error('Unable to update user. Please contact administrator.')
    
    return redirect(url_for('users_list'))

@app.route('/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@role_required('super_admin')
def user_reset_password(user_id):
    try:
        # Always resets to the system default; user must change on next login.
        models.set_user_password(user_id, _DEFAULT_PASSWORD)
        models.set_must_change_password(user_id, True)
        flash('Password reset to default. User will be prompted to change it on next login.', 'success')
    except Exception:
        flash_internal_error('Unable to reset password. Please contact administrator.')
    return redirect(url_for('users_list'))

@app.route('/users/<int:user_id>/reset-username', methods=['POST'])
@login_required
@role_required('super_admin')
def user_reset_username(user_id):
    try:
        new_username = request.form.get('new_username', '').strip()
        if not new_username:
            flash('Username cannot be empty.', 'warning')
            return redirect(url_for('users_list'))
        
        if len(new_username) < 3:
            flash('Username must be at least 3 characters.', 'warning')
            return redirect(url_for('users_list'))
        if not re.match(r'^[A-Za-z0-9_.-]+$', new_username):
            flash('Username can only contain letters, numbers, dot, underscore, and hyphen.', 'warning')
            return redirect(url_for('users_list'))
        
        models.set_username(user_id, new_username)
        flash('Username updated successfully.', 'success')
    except Exception as e:
        error_text = str(e).lower()
        if 'unique' in error_text or 'duplicate key' in error_text:
            flash('Username already exists. Choose a different username.', 'danger')
        else:
            flash_internal_error('Unable to update username. Please contact administrator.')
    
    return redirect(url_for('users_list'))


@app.route('/users/<int:user_id>/update-name', methods=['POST'])
@login_required
@role_required('super_admin')
def user_update_name(user_id):
    try:
        full_name = request.form.get('full_name', '').strip()
        if len(full_name) < 3:
            flash('Officer name must be at least 3 characters.', 'warning')
            return redirect(url_for('users_list'))

        models.update_user_full_name(user_id, full_name)
        flash('Officer name updated successfully.', 'success')
    except Exception:
        flash_internal_error('Unable to update officer name. Please contact administrator.')

    return redirect(url_for('users_list'))


@app.route('/users/<int:user_id>/update-contact', methods=['POST'])
@login_required
@role_required('super_admin')
def user_update_contact(user_id):
    try:
        phone = (request.form.get('phone') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        remove_photo = request.form.get('remove_photo') == 'on'
        photo_upload = request.files.get('profile_photo')

        if not validate_contact(phone):
            flash('Please provide a valid phone number.', 'warning')
            return redirect(url_for('users_list'))
        if not validate_email(email):
            flash('Please provide a valid email address.', 'warning')
            return redirect(url_for('users_list'))

        user = _get_user_by_id_cached(user_id)
        if not user:
            flash('User not found.', 'warning')
            return redirect(url_for('users_list'))

        ok_photo, stored_photo_name, photo_error = validate_profile_photo_upload(photo_upload, user_id)
        if not ok_photo:
            flash(photo_error, 'warning')
            return redirect(url_for('users_list'))

        old_photo = user.get('profile_photo')
        photo_changed = False
        try:
            models.update_user_profile_info(
                user_id,
                user.get('full_name'),
                phone,
                email
            )
            if stored_photo_name and photo_upload:
                ensure_upload_dirs()
                saved_ok, save_result = _save_uploaded_file(photo_upload, PROFILE_UPLOAD_DIR, stored_photo_name, 'Profile photo')
                if not saved_ok:
                    flash(save_result, 'danger')
                    return redirect(url_for('users_list'))
                stored_photo_name = save_result
                models.set_user_profile_photo(user_id, stored_photo_name)
                photo_changed = True
            elif remove_photo and old_photo:
                models.set_user_profile_photo(user_id, None)
                photo_changed = True

            if photo_changed and old_photo and old_photo != stored_photo_name:
                delete_profile_photo_file(old_photo)

            if session.get('user_id') == user_id:
                refresh_session_user()
            flash('User contact/profile photo updated.', 'success')
        except Exception:
            if stored_photo_name:
                delete_profile_photo_file(stored_photo_name)
            flash_internal_error('Unable to update contact/photo. Please contact administrator.')
    except Exception:
        flash_internal_error('Unable to update contact/photo. Please contact administrator.')

    return redirect(url_for('users_list'))

@app.route('/users/<int:inspector_id>/map-cvo', methods=['POST'])
@login_required
@role_required('super_admin')
def user_map_cvo(inspector_id):
    try:
        cvo_id_raw = request.form.get('cvo_id', '').strip()
        if not cvo_id_raw:
            flash('Please select a CVO/DSP for mapping.', 'warning')
            return redirect(url_for('users_list'))

        cvo_id = parse_optional_int(cvo_id_raw)
        if not cvo_id:
            flash('Please select a valid CVO/DSP for mapping.', 'warning')
            return redirect(url_for('users_list'))

        cvo_user = _get_user_by_id_cached(cvo_id)
        if not cvo_user or cvo_user.get('role') not in ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
            flash('Selected CVO/DSP mapping is invalid.', 'warning')
            return redirect(url_for('users_list'))

        models.map_inspector_to_cvo(inspector_id, cvo_id)
        flash('Field inspector mapped to CVO/DSP successfully.', 'success')
    except Exception:
        flash_internal_error('Unable to map inspector. Please contact administrator.')

    return redirect(url_for('users_list'))

# ========================================
# API ENDPOINTS
# ========================================

@app.route('/api/inspectors/<int:cvo_id>')
@login_required
def api_inspectors(cvo_id):
    if not _can_access_cvo_scope(cvo_id):
        log_security_event('access.api_inspectors_forbidden', severity='warning', cvo_id=cvo_id)
        return jsonify({'error': 'Forbidden'}), 403
    inspectors = models.get_inspectors_by_cvo(cvo_id)
    return jsonify([{'id': i['id'], 'full_name': i['full_name']} for i in inspectors])

@app.route('/api/stats')
@login_required
def api_stats():
    stats = models.get_dashboard_stats(session['user_role'], session['user_id'], session.get('cvo_office'))
    return jsonify(stats)


@app.route('/api/dashboard-drilldown')
@login_required
def api_dashboard_drilldown():
    metric = request.args.get('metric', '').strip()
    if not metric:
        return jsonify({'items': []})
    rows = models.get_dashboard_drilldown(
        session['user_role'],
        session['user_id'],
        session.get('cvo_office'),
        metric
    )
    accident_detail_map = {}
    petition_ids = [int(p.get('id')) for p in rows if p.get('id')]
    if petition_ids:
        try:
            accident_detail_map = models.get_latest_enquiry_report_accident_details(petition_ids) or {}
        except Exception:
            accident_detail_map = {}
    items = []
    for p in rows:
        petition_id = int(p.get('id') or 0)
        accident_summary = '-'
        if (p.get('petition_type') or '').strip() == 'electrical_accident':
            accident_summary = _format_electrical_accident_summary(accident_detail_map.get(petition_id))
        items.append({
            'id': petition_id,
            'sno': p.get('sno'),
            'petitioner_name': p.get('petitioner_name'),
            'subject': p.get('subject'),
            'status': p.get('status'),
            'received_date': p.get('received_date').strftime('%d/%m/%Y') if p.get('received_date') else '-',
            'accident_summary': accident_summary
        })
    return jsonify({'items': items})


@app.route('/api/dashboard-analytics')
@login_required
def api_dashboard_analytics():
    user_role = session['user_role']
    user_id = session['user_id']
    cvo_office = session.get('cvo_office')
    petitions = get_petitions_for_user_cached(user_id, user_role, cvo_office)
    officer_lookup = {}
    for p in petitions:
        officer_id = p.get('assigned_inspector_id')
        officer_name = (p.get('inspector_name') or '').strip()
        if officer_id and officer_name:
            officer_lookup[int(officer_id)] = officer_name
    dashboard_filter = _extract_dashboard_filters(request.args, officer_lookup)
    filtered_petitions = _apply_dashboard_filters(petitions, dashboard_filter)
    stats = _build_filtered_dashboard_stats(user_role, user_id, petitions, filtered_petitions)
    analytics = _build_dashboard_analytics(filtered_petitions, stats)
    return jsonify({'analytics': analytics, 'summary': analytics.get('summary', {})})


@app.route('/api/petitioner-suggestions')
@login_required
def api_petitioner_suggestions():
    q = _normalize_petitioner_name(request.args.get('q', ''))
    if len(q) < 2:
        return jsonify({'items': []})
    petitions = get_petitions_for_user_cached(session['user_id'], session['user_role'], session.get('cvo_office'))
    counter = Counter()
    q_lower = q.lower()
    for p in petitions:
        pname = _normalize_petitioner_name(p.get('petitioner_name') or '')
        if not pname or pname.lower() in {'anonymous', '-'}:
            continue
        if q_lower in pname.lower():
            counter[pname] += 1
    items = [{'name': name, 'count': count} for name, count in counter.most_common(12)]
    return jsonify({'items': items})


@app.route('/api/petitioner-profile')
@login_required
def api_petitioner_profile():
    name = _normalize_petitioner_name(request.args.get('name', ''))
    if not name:
        return jsonify({'error': 'Petitioner name is required.'}), 400
    petitions = get_petitions_for_user_cached(session['user_id'], session['user_role'], session.get('cvo_office'))
    payload = _build_petitioner_profile_payload(petitions, name)
    return jsonify(payload)


@app.route('/favicon.ico')
def favicon():
    return send_from_directory(
        os.path.join(app.root_path, 'static', 'img'),
        'nigaa-logo.svg',
        mimetype='image/svg+xml',
    )


@app.route('/healthz')
def healthz():
    return jsonify({'status': 'ok'}), 200


@app.route('/healthz/session-store')
def healthz_session_store():
    try:
        payload = _session_store_health_payload()
        return jsonify(payload), 200
    except Exception:
        app.logger.exception('Unable to compute session store health payload')
        return jsonify({'status': 'error', 'store': 'database'}), 503


@app.route('/admin/session-diagnostics')
@login_required
@role_required('super_admin')
def session_diagnostics():
    try:
        health = _session_store_health_payload()
    except Exception:
        health = {
            'store': 'database',
            'status': 'error',
        }
    diagnostics = list(SESSION_DIAGNOSTIC_EVENTS)[:100]
    return render_template(
        'session_diagnostics.html',
        health=health,
        diagnostics=diagnostics,
    )

# ========================================
# CHATBOT API
# ========================================

_CHATBOT_STATUS_LABELS = {
    'received': 'Received',
    'forwarded_to_cvo': 'Forwarded to CVO',
    'sent_for_permission': 'Sent for Permission',
    'permission_approved': 'Permission Approved',
    'permission_rejected': 'Permission Rejected',
    'assigned_to_inspector': 'Assigned to Inspector',
    'sent_back_for_reenquiry': 'Sent Back for Re-enquiry',
    'enquiry_in_progress': 'Enquiry In Progress',
    'enquiry_report_submitted': 'Enquiry Report Submitted',
    'forwarded_to_jmd': 'Forwarded to JMD',
    'forwarded_to_po': 'Forwarded to PO',
    'action_instructed': 'Action Instructed',
    'action_taken': 'Action Taken',
    'lodged': 'Lodged',
    'closed': 'Closed',
}


def _chatbot_format_petitions(petitions):
    out = []
    for p in petitions:
        subj = (p.get('subject') or '')
        out.append({
            'id': p['id'],
            'sno': p.get('sno') or '-',
            'petitioner_name': p.get('petitioner_name') or '-',
            'efile_no': p.get('efile_no') or '-',
            'ereceipt_no': p.get('ereceipt_no') or '-',
            'subject': subj[:90] + ('…' if len(subj) > 90 else ''),
            'status': p.get('status') or '-',
            'status_label': _CHATBOT_STATUS_LABELS.get(p.get('status', ''), (p.get('status') or '-').replace('_', ' ').title()),
            'received_date': str(p.get('received_date') or '-'),
            'petition_type': (p.get('petition_type') or '-').replace('_', ' ').title(),
            'place': p.get('place') or '-',
        })
    return out


def _chatbot_format_petitions_with_date(petitions):
    """Like _chatbot_format_petitions but also includes updated_at."""
    out = []
    for p in petitions:
        subj = (p.get('subject') or '')
        updated = p.get('updated_at')
        if updated:
            try:
                updated = updated.strftime('%d %b %Y, %I:%M %p')
            except Exception:
                updated = str(updated)
        out.append({
            'id': p['id'],
            'sno': p.get('sno') or '-',
            'petitioner_name': p.get('petitioner_name') or '-',
            'efile_no': p.get('efile_no') or '-',
            'ereceipt_no': p.get('ereceipt_no') or '-',
            'subject': subj[:90] + ('…' if len(subj) > 90 else ''),
            'status': p.get('status') or '-',
            'status_label': _CHATBOT_STATUS_LABELS.get(p.get('status', ''), (p.get('status') or '-').replace('_', ' ').title()),
            'received_date': str(p.get('received_date') or '-'),
            'petition_type': (p.get('petition_type') or '-').replace('_', ' ').title(),
            'place': p.get('place') or '-',
            'updated_at': updated or '-',
        })
    return out


@app.route('/api/chatbot', methods=['POST'])
@login_required
def chatbot_api():
    import re as _re
    import random as _random
    try:
        from rapidfuzz import fuzz as _fuzz
        _HAS_FUZZ = True
    except ImportError:
        _HAS_FUZZ = False

    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    if not message:
        return jsonify({'type': 'text', 'text': 'Go ahead, type something! I\'m all ears. 😊'})

    user_id = session['user_id']
    user_role = session.get('user_role', '')
    cvo_office = session.get('cvo_office')
    user_name = session.get('full_name') or 'Officer'
    msg_lower = message.lower().strip()

    # ---- Greetings ----
    greet_patterns = ('hi', 'hello', 'hey', 'namaste', 'good morning', 'good afternoon',
                      'good evening', 'howdy', 'hii', 'helo', 'hai', 'vanakkam', 'namaskar',
                      'sup', 'yo', 'greetings')
    if any(msg_lower == g or msg_lower.startswith(g + ' ') or msg_lower.startswith(g + '!') for g in greet_patterns):
        greet_replies = [
            f"Hey {user_name}! 👋 Great to see you. What can I help you with today?",
            f"Hello {user_name}! 😊 I'm Nigaa, your petition assistant. What do you need?",
            f"Hi there, {user_name}! 👋 Ready to help — try _\"pending\"_, _\"stats\"_, or _\"guide\"_.",
            f"Namaste, {user_name}! 🙏 How can I assist you today?",
            f"Hey! Good to have you here, {user_name}. 😄 What's on your mind?",
        ]
        return jsonify({
            'type': 'text',
            'text': _random.choice(greet_replies) + (
                "\n\n**Quick options:**\n"
                "• ⏳ _\"pending\"_ — petitions needing action\n"
                "• 🔔 _\"updates\"_ — recent activity\n"
                "• 📊 _\"stats\"_ — petition counts\n"
                "• 📋 _\"guide\"_ — workflow steps for your role"
            )
        })

    # ---- Thanks / Appreciation ----
    thanks_words = ('thanks', 'thank you', 'thank u', 'thx', 'ty', 'appreciated',
                    'great job', 'well done', 'nice', 'awesome', 'perfect', 'good bot',
                    'helpful', 'superb', 'excellent', 'brilliant', 'amazing', 'fantastic',
                    'good work', 'great', 'wonderful', 'cheers')
    if any(w in msg_lower for w in thanks_words):
        thanks_replies = [
            "Happy to help! 😊 Anything else you need?",
            "Glad I could assist! 🙌 Let me know if there's anything more.",
            "You're welcome! 😄 I'm always here when you need me.",
            "Anytime! 🤝 That's what I'm here for. Just ask!",
            "Sure thing! 👍 Feel free to ask anything else.",
            "Always a pleasure! 😊 Need anything else?",
        ]
        return jsonify({'type': 'text', 'text': _random.choice(thanks_replies)})

    # ---- Goodbye / Farewell ----
    bye_words = ('bye', 'goodbye', 'good bye', 'see you', 'cya', 'later', 'take care',
                 'ok thanks', 'ok bye', 'got it thanks', 'that is all', "that's all",
                 'no thanks', 'nothing else', 'all good', 'im done', "i'm done")
    if any(w in msg_lower for w in bye_words):
        bye_replies = [
            "Take care! 👋 Come back anytime you need help.",
            "Goodbye! Have a productive day! 😊",
            "See you soon! 🙏 Stay on top of those petitions!",
            "Bye for now! 👋 I'll be right here whenever you need me.",
            "Until next time! 😊 Keep up the great work!",
        ]
        return jsonify({'type': 'text', 'text': _random.choice(bye_replies)})

    # ---- How are you / general small talk ----
    if any(w in msg_lower for w in ('how are you', 'how r u', 'how do you do', 'whats up',
                                     "what's up", 'hows it going', "how's it going", 'all good')):
        how_replies = [
            "Doing great, thanks for asking! 😄 Ready to help you with petitions. What do you need?",
            "All systems go! 🚀 What can I assist you with today?",
            "Running at full speed! ⚡ What's on your plate today?",
            "I'm doing well! 😊 Always happy to help. What do you need?",
        ]
        return jsonify({'type': 'text', 'text': _random.choice(how_replies)})

    # ---- Who are you / intro ----
    if any(w in msg_lower for w in ('who are you', 'what are you', 'who r u',
                                     'introduce yourself', 'tell me about yourself', 'your name')):
        return jsonify({'type': 'text', 'text': (
            "I'm **Nigaa** 🤖 — your smart petition assistant!\n\n"
            "I help officers manage and track petitions right from this chat window. "
            "You can search petitions, check pending work, view stats, and get role-specific workflow guidance — "
            "all without leaving this page.\n\n"
            "Think of me as your digital sidekick! 🦸 Type _\"help\"_ to see everything I can do."
        )})

    # ---- Frustration / Problem ----
    frustration_words = ('not working', "doesn't work", 'doesnt work', 'broken', 'problem',
                         'issue', 'fail', 'failed', 'wrong', 'useless', 'bad bot',
                         'frustrated', 'stuck', 'confused', 'lost', 'no idea', 'dont understand',
                         "don't understand", 'not helpful', 'bad', 'terrible', 'hate this')
    if any(w in msg_lower for w in frustration_words):
        empathy_replies = [
            f"I'm sorry you're having trouble, {user_name}. 😔 Let me help — type _\"help\"_ to see all available commands.",
            "That sounds frustrating. I'm here to help! 🤝 Type _\"help\"_ to see what I can do for you.",
            f"Don't worry, {user_name}! 😊 Let's sort this out. What exactly do you need?",
            "I hear you! Let me try to make this easier. 🙏 What were you looking for?",
        ]
        return jsonify({'type': 'text', 'text': _random.choice(empathy_replies)})

    # ---- Help ----
    if msg_lower in ('help', '?', 'help me', 'commands', 'what can you do', 'options', 'menu'):
        return jsonify({'type': 'help'})

    # ---- Role / Responsibility info ----
    _ROLE_WORDS = ('my role', 'my responsibility', 'my responsibilities', 'my duties',
                   'my job', 'what is my role', 'what should i do', 'role info',
                   'what is my job', 'my work', 'who am i', 'what do i do',
                   'officer duty', 'my function', 'my task', 'my designation')
    _ROLE_DESCRIPTIONS = {
        'inspector': {
            'title': 'Field Inspector',
            'badge': '🔍',
            'color': '#06b6d4',
            'summary': 'You conduct field investigations on petitions assigned to you and submit enquiry reports.',
            'responsibilities': [
                '📥 Check petitions assigned to you under the Assigned tab',
                '🔍 Conduct on-site field investigations and gather evidence',
                '📝 Submit detailed enquiry reports with findings',
                '🔁 Re-investigate cases sent back for re-enquiry promptly',
                '📊 Keep petition status updated as you progress',
            ],
            'key_link': '/petitions?status=assigned_to_inspector',
        },
        'cvo_apspdcl': {
            'title': 'Chief Vigilance Officer (APSPDCL)',
            'badge': '🛡️',
            'color': '#8b5cf6',
            'summary': 'You oversee all petitions under your CVO jurisdiction, assign inspectors, and ensure timely resolution.',
            'responsibilities': [
                '📥 Review all petitions received in your CVO office',
                '👨‍💼 Assign petitions to appropriate field inspectors',
                '🔐 Request permission for sensitive/complex cases',
                '📊 Monitor inspector progress and SLA compliance',
                '⚠️ Escalate overdue cases to the Petitions Officer',
            ],
            'key_link': '/petitions',
        },
        'po': {
            'title': 'Petitions Officer',
            'badge': '⚖️',
            'color': '#6366f1',
            'summary': 'You oversee the entire petition lifecycle, approve/reject permissions, and issue action instructions.',
            'responsibilities': [
                '📥 Review petitions forwarded to you for action',
                '✅ Approve or reject CVO permission requests',
                '📣 Issue action instructions to CMD/CGM officers',
                '📊 Monitor SLA compliance across all CVOs',
                '⚠️ Escalate and resolve overdue petitions',
            ],
            'key_link': '/petitions?status=forwarded_to_po',
        },
        'data_entry': {
            'title': 'Data Entry Officer',
            'badge': '📋',
            'color': '#f59e0b',
            'summary': 'You register new petitions, upload documents, and route them to the correct CVO office.',
            'responsibilities': [
                '➕ Register new petitions with complete petitioner details',
                '📎 Upload scanned E-Receipts and supporting documents',
                '🏢 Route petitions to the correct CVO based on jurisdiction',
                '🔍 Track and verify petition status after submission',
                '✅ Ensure data accuracy and completeness at entry',
            ],
            'key_link': '/petitions/new',
        },
        'super_admin': {
            'title': 'Super Administrator',
            'badge': '👑',
            'color': '#ef4444',
            'summary': 'You have full system access — manage users, monitor all petitions, and ensure system-wide compliance.',
            'responsibilities': [
                '👥 Create and manage officer accounts and role assignments',
                '📊 Monitor all petitions and SLA compliance system-wide',
                '⚠️ Review and escalate overdue petitions across all CVOs',
                '📂 Generate reports and review audit activity',
                '🔧 Configure system settings and help resources',
            ],
            'key_link': '/users',
        },
    }
    _ROLE_DESCRIPTIONS['cvo_apepdcl'] = dict(_ROLE_DESCRIPTIONS['cvo_apspdcl'], title='Chief Vigilance Officer (APEPDCL)')
    _ROLE_DESCRIPTIONS['cvo_apcpdcl'] = dict(_ROLE_DESCRIPTIONS['cvo_apspdcl'], title='Chief Vigilance Officer (APCPDCL)')
    _ROLE_DESCRIPTIONS['dsp'] = dict(_ROLE_DESCRIPTIONS['cvo_apspdcl'], title='DSP Officer', badge='🔒')
    _ROLE_DESCRIPTIONS['cmd_apspdcl'] = {
        'title': 'CMD / CGM Officer',
        'badge': '🏛️',
        'color': '#10b981',
        'summary': 'You execute action instructions issued by the Petitions Officer and report back with actions taken.',
        'responsibilities': [
            '📥 Check action instructions assigned to you',
            '✅ Execute required corrective or disciplinary action',
            '📤 Submit action-taken report with supporting evidence',
            '📊 Track closed cases for compliance records',
        ],
        'key_link': '/petitions?status=action_instructed',
    }
    for _r in ('cmd_apepdcl', 'cmd_apcpdcl', 'cgm_hr_transco'):
        _ROLE_DESCRIPTIONS[_r] = _ROLE_DESCRIPTIONS['cmd_apspdcl']

    if any(phrase in msg_lower for phrase in _ROLE_WORDS):
        role_data = _ROLE_DESCRIPTIONS.get(user_role)
        if role_data:
            return jsonify({
                'type': 'role_info',
                'role': user_role,
                'role_data': role_data,
                'user_name': user_name,
                'suggestions': [
                    {'label': '📋 Guide', 'msg': 'guide'},
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '💡 What Next', 'msg': 'what next'},
                ],
            })
        return jsonify({'type': 'text', 'text': (
            f"You are logged in as **{user_role.replace('_', ' ').title()}**, {user_name}. "
            "Contact your administrator for role-specific guidance."
        )})

    # ---- Keyword sets for main intents ----
    _PENDING_WORDS = ('pending', 'overdue', 'waiting', 'my work', 'action needed',
                      'not closed', 'todo', 'to do', 'due', 'my petitions',
                      'assigned to me', 'open petition', 'outstanding', 'unresolved',
                      'backlog', 'incomplete', 'need to act', 'not done', 'left')
    _UPDATE_WORDS = ('update', 'updates', 'recent', 'activity', 'latest',
                     "what's new", 'whats new', 'notification', 'changes',
                     'modified', 'changed', 'progress', 'moved', 'feed', 'history')
    _GUIDE_WORDS = ('guide', 'how to', 'take action', 'workflow', 'process',
                    'next step', 'what should i do', 'instructions', 'steps',
                    'procedure', 'help me do', 'walkthrough', 'how do i', 'what do i do')
    _STATS_WORDS = ('stats', 'statistics', 'count', 'total', 'summary', 'how many',
                    'petition count', 'numbers', 'overview', 'breakdown', 'tally', 'figure')

    # ---- Pending petitions ----
    if any(w in msg_lower for w in _PENDING_WORDS):
        try:
            petitions = models.get_pending_petitions_for_chatbot(user_id, user_role, cvo_office)
            return jsonify({
                'type': 'pending',
                'petitions': _chatbot_format_petitions(petitions),
                'role': user_role,
                'suggestions': [
                    {'label': '💡 What Next', 'msg': 'what next'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                    {'label': '📋 Guide', 'msg': 'guide'},
                ],
            })
        except Exception:
            app.logger.exception('Chatbot pending error')
            return jsonify({'type': 'text', 'text': 'Hmm, I had trouble fetching pending petitions just now. Mind trying again? 🔄'})

    # ---- Recent updates ----
    if any(w in msg_lower for w in _UPDATE_WORDS):
        try:
            petitions = models.get_recent_updates_for_chatbot(user_id, user_role, cvo_office)
            return jsonify({
                'type': 'updates',
                'petitions': _chatbot_format_petitions_with_date(petitions),
                'role': user_role,
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                    {'label': '📈 Report', 'msg': 'report'},
                ],
            })
        except Exception:
            app.logger.exception('Chatbot updates error')
            return jsonify({'type': 'text', 'text': "Oops! Couldn't load updates right now. Please try again in a moment. 😔"})

    # ---- Action / workflow guide ----
    if any(w in msg_lower for w in _GUIDE_WORDS):
        return jsonify({
            'type': 'action_guide',
            'role': user_role,
            'suggestions': [
                {'label': '⏳ Pending', 'msg': 'pending'},
                {'label': '💡 What Next', 'msg': 'what next'},
            ],
        })

    # ---- Stats ----
    if any(w in msg_lower for w in _STATS_WORDS):
        try:
            stats = models.get_petition_stats_for_chatbot(user_id, user_role, cvo_office)
            return jsonify({
                'type': 'stats',
                'stats': {k: int(v) for k, v in stats.items()},
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '🔔 Updates', 'msg': 'updates'},
                    {'label': '📈 Report', 'msg': 'report'},
                ],
            })
        except Exception:
            app.logger.exception('Chatbot stats error')
            return jsonify({'type': 'text', 'text': 'Stats are taking a moment to load. Give it another shot! 🔄'})

    # ---- Analysis Report / Download ----
    _DOWNLOAD_WORDS = ('download', 'report', 'analysis', 'export', 'analysis report',
                       'download report', 'generate report', 'view report',
                       'excel', 'pdf report', 'csv report', 'see report')
    if any(w in msg_lower for w in _DOWNLOAD_WORDS):
        return jsonify({
            'type': 'download',
            'text': 'Access the full petition analysis with charts, trends, status breakdowns, and export options.',
            'url': '/analysis-report',
            'suggestions': [
                {'label': '📊 Stats', 'msg': 'stats'},
                {'label': '⏳ Pending', 'msg': 'pending'},
            ],
        })

    # ---- Urgent / Overdue / SLA ----
    _URGENT_WORDS = ('urgent', 'overdue', 'sla breach', 'beyond sla', 'critical',
                      'escalate', 'delayed', 'late', 'sla', 'priority cases',
                      'breach', 'violation', 'missed deadline')
    if any(w in msg_lower for w in _URGENT_WORDS):
        if user_role in ('po', 'super_admin'):
            return jsonify({
                'type': 'urgent',
                'message': 'Petitions beyond SLA threshold are flagged for escalation. Review them below.',
                'url': '/petitions?status=beyond_sla',
                'sla_url': '/sla_dashboard',
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                    {'label': '📈 Report', 'msg': 'report'},
                ],
            })
        else:
            return jsonify({
                'type': 'text',
                'text': (
                    "For SLA breach and overdue escalation details, check with your **Petitions Officer** "
                    "or visit the **SLA Dashboard**.\n\n"
                    "Your immediate pending items can be found by asking for _\"pending\"_."
                ),
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📋 Guide', 'msg': 'guide'},
                ],
            })

    # ---- Daily Summary ----
    _SUMMARY_WORDS = ('today', 'daily', 'my day', 'daily report', 'today summary',
                       'day summary', 'what happened today', 'daily summary', 'todays update')
    if any(w in msg_lower for w in _SUMMARY_WORDS):
        try:
            _stats = models.get_petition_stats_for_chatbot(user_id, user_role, cvo_office)
            _pending = models.get_pending_petitions_for_chatbot(user_id, user_role, cvo_office, limit=20)
            _updates = models.get_recent_updates_for_chatbot(user_id, user_role, cvo_office, limit=20)
            _pc = len(_pending)
            _uc = len(_updates)
            if _pc > 0:
                _day_msg = (f"You have **{_pc} pending** item{'s' if _pc != 1 else ''} and "
                            f"**{_uc} update{'s' if _uc != 1 else ''}** today, {user_name}.")
            else:
                _day_msg = (f"Great work, {user_name}! No pending items today. "
                            f"{_uc} update{'s' if _uc != 1 else ''} in your scope.")
            return jsonify({
                'type': 'summary',
                'stats': {k: int(v) for k, v in _stats.items()},
                'pending_count': _pc,
                'updates_count': _uc,
                'message': _day_msg,
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '🔔 Updates', 'msg': 'updates'},
                    {'label': '💡 What Next', 'msg': 'what next'},
                ],
            })
        except Exception:
            app.logger.exception('Chatbot summary error')
            return jsonify({'type': 'text', 'text': 'Could not load your daily summary right now. Try again! 🔄'})

    # ---- What Next / Smart Suggestions ----
    _SUGGEST_WORDS = ('what next', 'suggest', 'recommendation', 'next action',
                       'what should i work on', 'help me prioritize', 'what to do next',
                       'prioritize', 'my priority', 'next steps', 'suggest me', 'advise me')
    if any(w in msg_lower for w in _SUGGEST_WORDS):
        try:
            _pending_list = models.get_pending_petitions_for_chatbot(user_id, user_role, cvo_office)
            _pc2 = len(_pending_list)
            _rd2 = _ROLE_DESCRIPTIONS.get(user_role, {})
            _key_link2 = _rd2.get('key_link', '/petitions')
            _actions2 = []
            if _pc2 > 0:
                _actions2.append({
                    'priority': 'high', 'icon': '🔴',
                    'title': f'{_pc2} Pending Petition{"s" if _pc2 != 1 else ""}',
                    'desc': f'{"These petitions require" if _pc2 > 1 else "This petition requires"} your immediate attention.',
                    'link': _key_link2, 'link_label': 'View Pending →',
                })
            if user_role == 'inspector':
                _actions2 += [
                    {'priority': 'medium', 'icon': '📝', 'title': 'Submit Pending Reports',
                     'desc': 'Complete and submit enquiry reports for in-progress petitions.',
                     'link': '/petitions?status=enquiry_in_progress', 'link_label': 'View In Progress →'},
                    {'priority': 'low', 'icon': '🔁', 'title': 'Check Re-enquiry Cases',
                     'desc': 'Review petitions sent back for re-investigation.',
                     'link': '/petitions?status=sent_back_for_reenquiry', 'link_label': 'View →'},
                ]
            elif user_role in ('cvo_apspdcl', 'cvo_apepdcl', 'cvo_apcpdcl', 'dsp'):
                _actions2 += [
                    {'priority': 'medium', 'icon': '👨\u200d💼', 'title': 'Assign Uninspected Petitions',
                     'desc': 'Route unassigned petitions to appropriate field inspectors.',
                     'link': '/petitions?status=forwarded_to_cvo', 'link_label': 'View →'},
                    {'priority': 'low', 'icon': '📊', 'title': 'Review SLA Compliance',
                     'desc': 'Check enquiry progress and monitor SLA status.',
                     'link': '/sla_dashboard', 'link_label': 'SLA Dashboard →'},
                ]
            elif user_role == 'po':
                _actions2 += [
                    {'priority': 'medium', 'icon': '✅', 'title': 'Approve/Reject Permissions',
                     'desc': 'CVOs are waiting for your decision on permission requests.',
                     'link': '/petitions?status=sent_for_permission', 'link_label': 'Review →'},
                    {'priority': 'medium', 'icon': '📣', 'title': 'Issue Action Instructions',
                     'desc': 'Petitions ready for action instruction to CMD/CGM.',
                     'link': '/petitions?status=forwarded_to_po', 'link_label': 'View →'},
                    {'priority': 'low', 'icon': '⚠️', 'title': 'SLA Overdue Review',
                     'desc': 'Check petitions beyond SLA and escalate as needed.',
                     'link': '/petitions?status=beyond_sla', 'link_label': 'View Overdue →'},
                ]
            elif user_role == 'data_entry':
                _actions2 += [
                    {'priority': 'medium', 'icon': '➕', 'title': 'Register New Petitions',
                     'desc': 'Enter any unregistered petitions into the system.',
                     'link': '/petitions/new', 'link_label': 'Add Petition →'},
                    {'priority': 'low', 'icon': '🔍', 'title': 'Verify Data Accuracy',
                     'desc': 'Review recently registered petitions for completeness.',
                     'link': '/petitions', 'link_label': 'View All →'},
                ]
            elif user_role == 'super_admin':
                _actions2 += [
                    {'priority': 'medium', 'icon': '📊', 'title': 'Review System Analytics',
                     'desc': 'Check overall system performance and compliance metrics.',
                     'link': '/analysis-report', 'link_label': 'View Report →'},
                    {'priority': 'low', 'icon': '👥', 'title': 'User Management',
                     'desc': 'Review pending account requests and role assignments.',
                     'link': '/users', 'link_label': 'Manage Users →'},
                ]
            elif user_role in ('cmd_apspdcl', 'cmd_apepdcl', 'cmd_apcpdcl', 'cgm_hr_transco'):
                _actions2 += [
                    {'priority': 'medium', 'icon': '📤', 'title': 'Submit Action Reports',
                     'desc': 'Report back on actions taken for instructed petitions.',
                     'link': '/petitions?status=action_instructed', 'link_label': 'View →'},
                ]
            if not _actions2:
                _actions2.append({
                    'priority': 'low', 'icon': '✅', 'title': 'All Caught Up!',
                    'desc': 'No immediate actions required. Monitor your dashboard for new items.',
                    'link': '/', 'link_label': 'Go to Dashboard →',
                })
            return jsonify({
                'type': 'suggest',
                'actions': _actions2,
                'user_name': user_name,
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                    {'label': '📋 Guide', 'msg': 'guide'},
                ],
            })
        except Exception:
            app.logger.exception('Chatbot suggest error')
            return jsonify({'type': 'action_guide', 'role': user_role})

    # ---- Search by petitioner name ----
    name_match = _re.search(r'(?:search|find|name|petitioner)[:\s]+(.+)', msg_lower)
    if name_match:
        query = name_match.group(1).strip()
        if len(query) < 2:
            return jsonify({'type': 'text', 'text': 'Could you give me at least 2 characters to search by name? 🔍'})
        try:
            results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='name')
            # If name search finds nothing, automatically try all fields (covers e-office/e-receipt typed after "search")
            if not results:
                results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='all')
            return jsonify({
                'type': 'petitions',
                'petitions': _chatbot_format_petitions(results),
                'query': query,
                'search_type': 'name',
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                ],
            })
        except Exception:
            app.logger.exception('Chatbot search error')
            return jsonify({'type': 'text', 'text': 'Search hit a snag — please try again! 🔄'})

    # ---- Search by E-Office / efile number (keyword prefix) ----
    efile_match = _re.search(r'(?:eoffice|efile|e-office|e-file|file\s*no)[:\s#]*([A-Za-z0-9/_\-\.]+)', msg_lower)
    if efile_match:
        query = efile_match.group(1).strip()
        try:
            results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='efile')
            return jsonify({
                'type': 'petitions',
                'petitions': _chatbot_format_petitions(results),
                'query': query,
                'search_type': 'efile',
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                ],
            })
        except Exception:
            return jsonify({'type': 'text', 'text': 'Search failed. Please try again! 🔄'})

    # ---- Search by E-Receipt number (keyword prefix) ----
    ereceipt_match = _re.search(r'(?:ereceipt|e-receipt|receipt)[:\s#]*([A-Za-z0-9/_\-\.]+)', msg_lower)
    if ereceipt_match:
        query = ereceipt_match.group(1).strip()
        try:
            results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='ereceipt')
            return jsonify({
                'type': 'petitions',
                'petitions': _chatbot_format_petitions(results),
                'query': query,
                'search_type': 'ereceipt',
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                ],
            })
        except Exception:
            return jsonify({'type': 'text', 'text': 'Search failed. Please try again! 🔄'})

    # ---- Search by SNO ----
    sno_match = _re.search(r'(?:sno|serial|vig)[:\s#]*([A-Za-z0-9/_\-]+)', msg_lower)
    if sno_match:
        query = sno_match.group(1).strip()
        try:
            results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='sno')
            return jsonify({
                'type': 'petitions',
                'petitions': _chatbot_format_petitions(results),
                'query': query,
                'search_type': 'sno',
                'suggestions': [
                    {'label': '⏳ Pending', 'msg': 'pending'},
                    {'label': '📊 Stats', 'msg': 'stats'},
                ],
            })
        except Exception:
            return jsonify({'type': 'text', 'text': 'Search failed. Please try again! 🔄'})

    # ---- Smart bare-number detection (no keyword prefix needed) ----
    # E-Receipt: starts with 2+ letters then digits/slash  e.g. "ER2024001", "ER/2024/001"
    bare_ereceipt = _re.match(r'^[a-z]{1,4}[\-/]?\d{4,}', msg_lower)
    # E-Office: slash-separated path  e.g. "VIG/HQ/2024/01", "vig/cor/2025/100"
    bare_efile = _re.match(r'^[a-z0-9]+(?:/[a-z0-9]+){2,}', msg_lower)

    if bare_ereceipt and not bare_efile:
        query = message.strip()
        try:
            results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='ereceipt')
            if not results:
                results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='all')
            return jsonify({'type': 'petitions', 'petitions': _chatbot_format_petitions(results),
                            'query': query, 'search_type': 'ereceipt'})
        except Exception:
            return jsonify({'type': 'text', 'text': 'Search failed. Please try again! 🔄'})

    if bare_efile:
        query = message.strip()
        try:
            results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='efile')
            if not results:
                results = models.search_petitions(user_id, user_role, cvo_office, query, search_type='all')
            return jsonify({'type': 'petitions', 'petitions': _chatbot_format_petitions(results),
                            'query': query, 'search_type': 'efile'})
        except Exception:
            return jsonify({'type': 'text', 'text': 'Search failed. Please try again! 🔄'})

    # ---- Fuzzy intent matching (rapidfuzz) — catches typos like "pendin", "stas", "updtes" ----
    if _HAS_FUZZ:
        _FUZZY_MAP = [
            ('pending', _PENDING_WORDS),
            ('updates', _UPDATE_WORDS),
            ('guide',   _GUIDE_WORDS),
            ('stats',   _STATS_WORDS),
        ]
        best_intent_name = None
        best_score = 0
        for word in msg_lower.split():
            for intent_name, keywords in _FUZZY_MAP:
                for kw in keywords:
                    if ' ' not in kw:  # single-word keywords only
                        score = _fuzz.ratio(word, kw)
                        if score > best_score:
                            best_score = score
                            best_intent_name = intent_name
        if best_score >= 78 and best_intent_name:
            if best_intent_name == 'pending':
                try:
                    petitions = models.get_pending_petitions_for_chatbot(user_id, user_role, cvo_office)
                    return jsonify({'type': 'pending', 'petitions': _chatbot_format_petitions(petitions), 'role': user_role})
                except Exception:
                    pass
            elif best_intent_name == 'updates':
                try:
                    petitions = models.get_recent_updates_for_chatbot(user_id, user_role, cvo_office)
                    return jsonify({'type': 'updates', 'petitions': _chatbot_format_petitions_with_date(petitions), 'role': user_role})
                except Exception:
                    pass
            elif best_intent_name == 'guide':
                return jsonify({'type': 'action_guide', 'role': user_role})
            elif best_intent_name == 'stats':
                try:
                    stats = models.get_petition_stats_for_chatbot(user_id, user_role, cvo_office)
                    return jsonify({'type': 'stats', 'stats': {k: int(v) for k, v in stats.items()}})
                except Exception:
                    pass

    # ---- Generic fallback: try full-text search ----
    if len(message) >= 3:
        try:
            results = models.search_petitions(user_id, user_role, cvo_office, message, search_type='all')
            if results:
                return jsonify({'type': 'petitions', 'petitions': _chatbot_format_petitions(results),
                                'query': message, 'search_type': 'all'})
        except Exception:
            pass

    # ---- Final conversational fallback ----
    fallback_replies = [
        (
            f"Hmm, I'm not quite sure what you meant, {user_name}. 🤔\n\n"
            "Here's what I can help with:\n"
            "• **\"pending\"** — petitions waiting for your action\n"
            "• **\"updates\"** — recent activity and status changes\n"
            "• **\"guide\"** — step-by-step workflow for your role\n"
            "• **\"search Ravi Kumar\"** — find petitions by name\n"
            "• **\"stats\"** — petition statistics\n"
            "• **\"help\"** — full command list"
        ),
        (
            "I didn't quite catch that! 😅 Here are some things to try:\n\n"
            "• _\"pending\"_ to see what needs your attention\n"
            "• _\"search Ravi Kumar\"_ to find a petition\n"
            "• _\"stats\"_ for a quick overview\n"
            "• _\"help\"_ for all commands"
        ),
        (
            f"Not sure about that one, {user_name}! 🤖 "
            "Try **\"help\"** to see everything I can do, or ask me to search a petition by name."
        ),
    ]
    return jsonify({'type': 'text', 'text': _random.choice(fallback_replies)})


# ========================================
# RUN
# ========================================

if __name__ == '__main__':
    app.run(debug=config.DEBUG, host=config.HOST, port=config.PORT)
